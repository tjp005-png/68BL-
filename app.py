import sqlite3
import pandas as pd
from io import BytesIO
from datetime import date, timedelta, datetime
from contextlib import closing
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file
from openpyxl.styles import Font, Alignment, Border, Side
import pdfplumber
import re
import zipfile
import base64
import json
import os
import uuid
import time
from collections import defaultdict

import stu_services 

app = Flask(__name__)

# ==========================================
#        CONFIGURATION & HELPER VARS
# ==========================================
TARGET_ROUTING_CODE = "BL"
PERMITTED_CODES = {'CBP', 'CNO', 'CBPS', 'CNOS', 'CNIA', 'CCS', 'CCSS', 'D23'}

# MULTI-USER SYNC TRACKER
SCHEDULE_UPDATES = {'GLOBAL': 0}

def safe_xml(text):
    return str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

def get_db_connection():
    conn = sqlite3.connect('database.db', timeout=15)
    conn.row_factory = sqlite3.Row 
    conn.execute('PRAGMA journal_mode=WAL;')
    return conn

def column_exists(cursor, table_name, column_name):
    cursor.execute(f"PRAGMA table_info({table_name})")
    columns = [row['name'] for row in cursor.fetchall()]
    return column_name in columns

def upgrade_db():
    with closing(get_db_connection()) as conn:
        cursor = conn.cursor()
        
        # 1. TRUCK LOGS
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS truck_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                truck_id TEXT, profile_number TEXT, manifest_number TEXT, load_number TEXT,
                gross_weight REAL, test_assigned TEXT, test_status TEXT, exit_weight REAL,
                net_weight REAL, cell_location TEXT, grid_location TEXT
            )
        ''')
        if not column_exists(cursor, 'truck_logs', 'lab_results'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN lab_results TEXT')
        if not column_exists(cursor, 'truck_logs', 'manifest_weight'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN manifest_weight REAL')
        if not column_exists(cursor, 'truck_logs', 'manifest_units'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN manifest_units TEXT')
        if not column_exists(cursor, 'truck_logs', 'date_received'):
            cursor.execute('ALTER TABLE truck_logs ADD COLUMN date_received TEXT')
            cursor.execute("UPDATE truck_logs SET date_received = ? WHERE date_received IS NULL", (date.today().isoformat(),))
        if not column_exists(cursor, 'truck_logs', 'rejection_reason'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN rejection_reason TEXT')
        if not column_exists(cursor, 'truck_logs', 'extra_fees'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN extra_fees TEXT DEFAULT "None"')
        if not column_exists(cursor, 'truck_logs', 'sales_order'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN sales_order TEXT DEFAULT ""')
        if not column_exists(cursor, 'truck_logs', 'time_in'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN time_in TEXT DEFAULT ""')
        if not column_exists(cursor, 'truck_logs', 'time_out'): cursor.execute('ALTER TABLE truck_logs ADD COLUMN time_out TEXT DEFAULT ""')

        # 2. MASTER PROFILES 
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS profiles (
                profile_number TEXT PRIMARY KEY, generator TEXT, 
                waste_description TEXT, win_code TEXT, voc_percentage REAL DEFAULT 0.0,
                special_handling TEXT
            )
        ''')
        if not column_exists(cursor, 'profiles', 'waste_description'): cursor.execute('ALTER TABLE profiles ADD COLUMN waste_description TEXT')
        if not column_exists(cursor, 'profiles', 'win_code'): cursor.execute('ALTER TABLE profiles ADD COLUMN win_code TEXT')
        if not column_exists(cursor, 'profiles', 'special_handling'): cursor.execute('ALTER TABLE profiles ADD COLUMN special_handling TEXT')
        if not column_exists(cursor, 'profiles', 'ph_range'): cursor.execute('ALTER TABLE profiles ADD COLUMN ph_range TEXT')
        if not column_exists(cursor, 'profiles', 'physical_appearance'): cursor.execute('ALTER TABLE profiles ADD COLUMN physical_appearance TEXT')
        if not column_exists(cursor, 'profiles', 'flash_point'): cursor.execute('ALTER TABLE profiles ADD COLUMN flash_point TEXT')
        if not column_exists(cursor, 'profiles', 'expiration_date'): cursor.execute('ALTER TABLE profiles ADD COLUMN expiration_date TEXT')
        if not column_exists(cursor, 'profiles', 'expiration_date'):
            cursor.execute('ALTER TABLE profiles ADD COLUMN expiration_date TEXT')
        if not column_exists(cursor, 'profiles', 'status'):
            cursor.execute('ALTER TABLE profiles ADD COLUMN status TEXT DEFAULT "A"')

        # 3. DAILY SCHEDULE 
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS daily_schedule (
                id INTEGER PRIMARY KEY AUTOINCREMENT, schedule_date TEXT, start_time TEXT,
                end_time TEXT, profile_number TEXT, load_count INTEGER, generator TEXT,
                waste_type TEXT, sales_order TEXT, routing_code TEXT, scheduler_initials TEXT, special_notes TEXT
            )
        ''')
        if not column_exists(cursor, 'daily_schedule', 'voc_level'): cursor.execute('ALTER TABLE daily_schedule ADD COLUMN voc_level REAL')
        if not column_exists(cursor, 'daily_schedule', 'is_unscheduled'): cursor.execute('ALTER TABLE daily_schedule ADD COLUMN is_unscheduled INTEGER DEFAULT 0')
        if not column_exists(cursor, 'daily_schedule', 'order_index'): cursor.execute('ALTER TABLE daily_schedule ADD COLUMN order_index INTEGER DEFAULT 0')
        if not column_exists(cursor, 'daily_schedule', 'series_id'): cursor.execute('ALTER TABLE daily_schedule ADD COLUMN series_id TEXT')
        if not column_exists(cursor, 'daily_schedule', 'is_pinned'): cursor.execute('ALTER TABLE daily_schedule ADD COLUMN is_pinned INTEGER DEFAULT 0')
        if not column_exists(cursor, 'daily_schedule', 'order_index'):
            cursor.execute('ALTER TABLE daily_schedule ADD COLUMN order_index INTEGER DEFAULT 0')
        if not column_exists(cursor, 'daily_schedule', 'is_pinned'):
            cursor.execute('ALTER TABLE daily_schedule ADD COLUMN is_pinned INTEGER DEFAULT 0')

        # 4. CREATE STU DRUM INVENTORY TABLE
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS drum_inventory (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                track_no TEXT,
                inb_prof TEXT,
                process_type TEXT,
                weight REAL,
                ph REAL,
                age REAL,
                voc_ppm REAL,
                voc_weight REAL,
                import_date TEXT
            )
        ''')
        if not column_exists(cursor, 'drum_inventory', 'manifest'): cursor.execute('ALTER TABLE drum_inventory ADD COLUMN manifest TEXT')

        # 5. DRUM SAMPLING COMPLIANCE TABLES
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS compliance_tracker (
                profile TEXT PRIMARY KEY,
                last_voc_test_date DATE,
                drums_since_last_test INTEGER
            );
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_log (
                log_id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp DATETIME,
                profile TEXT,
                drum_id TEXT,
                manifest TEXT,
                trigger_reason TEXT,
                chemist_name TEXT,
                voc_result REAL
            );
        ''')

        # 6. DRUM LAB QUEUE
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS drum_lab_queue (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT, drum_id TEXT, profile TEXT, manifest TEXT,
                tests_required TEXT, status TEXT DEFAULT 'PENDING',
                fp_result TEXT, ph_result REAL, voc_result REAL, notes TEXT
            )
        ''')
        if not column_exists(cursor, 'drum_lab_queue', 'flashpoint'): cursor.execute('ALTER TABLE drum_lab_queue ADD COLUMN flashpoint TEXT')
        if not column_exists(cursor, 'drum_lab_queue', 'cyanide'): cursor.execute('ALTER TABLE drum_lab_queue ADD COLUMN cyanide TEXT')
        if not column_exists(cursor, 'drum_lab_queue', 'sulfide'): cursor.execute('ALTER TABLE drum_lab_queue ADD COLUMN sulfide TEXT')
        if not column_exists(cursor, 'drum_lab_queue', 'oxidation'): cursor.execute('ALTER TABLE drum_lab_queue ADD COLUMN oxidation TEXT')
        
        conn.commit()

upgrade_db()


# ==========================================
#        ROUTES (PORTAL / HUB)
# ==========================================

@app.route('/')
def portal_hub(): return render_template('index.html')

@app.route('/receiving')
def home():
    today_str = date.today().isoformat() 
    with closing(get_db_connection()) as conn:
        pending_trucks = conn.execute("SELECT * FROM truck_logs WHERE exit_weight IS NULL AND test_status != 'REJECTED' ORDER BY id DESC").fetchall()
        flagged_query = conn.execute("SELECT DISTINCT profile_number FROM daily_schedule WHERE schedule_date = ? AND is_unscheduled = 1", (today_str,)).fetchall()
        scheduled_query = conn.execute("SELECT DISTINCT profile_number FROM daily_schedule WHERE schedule_date = ?", (today_str,)).fetchall()
        
    return render_template('receiving_log.html', pending_trucks=pending_trucks, flagged_profiles=[row['profile_number'] for row in flagged_query], scheduled_profiles=[row['profile_number'] for row in scheduled_query], today_str=today_str)

@app.route('/api/search_profiles')
def search_profiles():
    query = request.args.get('q', '').strip() 
    with closing(get_db_connection()) as conn:
        results = conn.execute('SELECT profile_number FROM profiles WHERE profile_number LIKE ? LIMIT 50', ('%' + query + '%',)).fetchall()
    return jsonify([{'value': row['profile_number'], 'text': row['profile_number']} for row in results])

@app.route('/api/get_profile_details/<path:profile_number>')
def get_profile_details(profile_number):
    clean_profile = profile_number.strip().upper()
    
    with closing(get_db_connection()) as conn:
        profile = conn.execute('SELECT * FROM profiles WHERE TRIM(UPPER(profile_number)) = ?', (clean_profile,)).fetchone()
    
    if profile:
        p_dict = dict(profile)
        
        voc_raw = str(p_dict.get('voc_percentage', '')).strip()
        if voc_raw in ['?', '', 'None']:
            p_dict['voc_percentage'] = 'TBD'
        else:
            p_dict['voc_percentage'] = voc_raw
            
        exp_raw = str(p_dict.get('expiration_date', '')).strip().lower()
        if exp_raw == 'none':
            p_dict['expiration_date'] = ''
            
        notes = p_dict.get('special_handling') or ""
        notes = re.sub(r'(?i)\bNOT CERCLA\b', '', notes)
        notes = re.sub(r'(?i)\bCERCLA\b', '', notes)
        notes = re.sub(r'^[,\s]+|[,\s]+$', '', notes) 
        p_dict['special_handling'] = notes.strip()
        
        gen = p_dict.get('generator')
        p_dict['generator'] = str(gen).strip() if gen and str(gen).strip().lower() != 'none' else ''
            
        return jsonify(p_dict)
        
    return jsonify({'error': 'Profile not found'}), 404

@app.route('/api/auto_sync_profiles', methods=['POST'])
def auto_sync_profiles():
    date_str = request.form.get('schedule_date')
    updates_made = False
    
    if date_str:
        with closing(get_db_connection()) as conn:
            schedules = conn.execute('SELECT id, profile_number, voc_level, generator FROM daily_schedule WHERE schedule_date = ?', (date_str,)).fetchall()
            
            for s in schedules:
                prof_num = str(s['profile_number']).strip().upper()
                prof = conn.execute('''
                    SELECT voc_percentage, generator, win_code 
                    FROM profiles 
                    WHERE TRIM(UPPER(profile_number)) = ?
                ''', (prof_num,)).fetchone()
                
                if prof:
                    try:
                        val = float(prof['voc_percentage'])
                        new_voc = str(int(val)) if val > 0 else '0'
                    except (ValueError, TypeError):
                        new_voc = 'TBD'
                        
                    # CRITICAL: Only update the database IF the values are actually different
                    if new_voc != str(s['voc_level']) or str(prof['generator']) != str(s['generator']):
                        conn.execute('''
                            UPDATE daily_schedule 
                            SET voc_level = ?, generator = ?, routing_code = ?
                            WHERE id = ?
                        ''', (new_voc, prof['generator'], prof['win_code'], s['id']))
                        updates_made = True
                        
            if updates_made:
                conn.commit()
                
        # Only ping the connected users if we actually changed something
        if updates_made:
            SCHEDULE_UPDATES[date_str] = time.time()
            
            
    return jsonify({'updated': updates_made})

@app.route('/refresh_schedule_data', methods=['POST'])
def refresh_schedule_data():
    date_str = request.form.get('schedule_date')
    
    if date_str:
        with closing(get_db_connection()) as conn:
            # 1. Grab all scheduled loads for the current view
            schedules = conn.execute('SELECT id, profile_number FROM daily_schedule WHERE schedule_date = ?', (date_str,)).fetchall()
            
            for s in schedules:
                prof_num = str(s['profile_number']).strip().upper()
                
                # 2. Look up the latest data in the Master Profiles table
                prof = conn.execute('''
                    SELECT voc_percentage, generator, win_code 
                    FROM profiles 
                    WHERE TRIM(UPPER(profile_number)) = ?
                ''', (prof_num,)).fetchone()
                
                if prof:
                    # 3. Safely parse the updated VOC using our new logic
                    try:
                        val = float(prof['voc_percentage'])
                        new_voc = str(int(val)) if val > 0 else '0'
                    except (ValueError, TypeError):
                        new_voc = 'TBD'
                        
                    # 4. Update the existing schedule entry
                    conn.execute('''
                        UPDATE daily_schedule 
                        SET voc_level = ?, generator = ?, routing_code = ?
                        WHERE id = ?
                    ''', (new_voc, prof['generator'], prof['win_code'], s['id']))
                    
            conn.commit()
            
        # 5. Push the visual update to all connected users
        SCHEDULE_UPDATES[date_str] = time.time()
        SCHEDULE_UPDATES['GLOBAL'] = time.time()
        
    return redirect(url_for('schedule_portal', date=date_str))

@app.route('/api/check_duplicate')
def check_duplicate():
    check_date_str = request.args.get('date')
    profile = request.args.get('profile', '').strip().upper()
    try: repeat_weeks = int(request.args.get('repeat', 1))
    except: repeat_weeks = 1
    
    if not check_date_str or not profile:
        return jsonify({'exists': False})
        
    try: base_date = datetime.strptime(check_date_str, '%Y-%m-%d')
    except: return jsonify({'exists': False})

    with closing(get_db_connection()) as conn:
        days_checked = 0
        current_date = base_date
        
        while days_checked < repeat_weeks:
            is_weekend = current_date.weekday() >= 5
            is_first_day = (days_checked == 0)
            
            if not is_weekend or is_first_day:
                target_date = current_date.strftime('%Y-%m-%d')
                existing = conn.execute('SELECT id FROM daily_schedule WHERE schedule_date = ? AND profile_number = ?', (target_date, profile)).fetchone()
                
                if existing:
                    return jsonify({'exists': True, 'conflict_date': target_date})
                    
                days_checked += 1
                
            current_date += timedelta(days=1)
            
        return jsonify({'exists': False})

@app.route('/api/check_truck_duplicate')
def check_truck_duplicate():
    manifest = request.args.get('manifest', '').strip()
    load = request.args.get('load', '').strip()
    check_date = request.args.get('date', date.today().isoformat())

    with closing(get_db_connection()) as conn:
        existing = conn.execute('''
            SELECT load_number, manifest_number 
            FROM truck_logs 
            WHERE date_received = ? AND test_status != 'REJECTED' 
            AND (manifest_number = ? OR load_number = ?)
        ''', (check_date, manifest, load)).fetchone()

        if existing:
            return jsonify({'exists': True, 'load': existing['load_number'], 'manifest': existing['manifest_number']})
        return jsonify({'exists': False})

@app.route('/api/check_schedule_updates')
def check_schedule_updates():
    check_date = request.args.get('date')
    return jsonify({
        'date_updated': SCHEDULE_UPDATES.get(check_date, 0),
        'global_updated': SCHEDULE_UPDATES.get('GLOBAL', 0)
    })

@app.route('/submit_truck', methods=['POST'])
def submit_truck():
    from datetime import datetime
    
    # We removed truck_id from the UI, so we default it to empty string to prevent errors
    truck_id = request.form.get('truck_id', '') 
    profile_number = request.form.get('profile_number', '').strip().upper()
    manifest_number = request.form.get('manifest_number')
    load_number = request.form.get('load_number')
    
    try: gross_weight = float(request.form.get('gross_weight', 0))
    except: gross_weight = 0.0
    
    received_date = request.form.get('date_received')
    if not received_date:
        received_date = date.today().isoformat()
        
    # Capture the exact check-in time
    time_in = datetime.now().strftime('%H:%M')
    
    with closing(get_db_connection()) as conn:
        profile = conn.execute('SELECT * FROM profiles WHERE profile_number = ?', (profile_number,)).fetchone()
        previous_trucks = conn.execute('SELECT COUNT(*) FROM truck_logs WHERE profile_number = ?', (profile_number,)).fetchone()[0]
        
        # Auto-fetch the Sales Order from today's schedule
        schedule_entry = conn.execute('''
            SELECT sales_order FROM daily_schedule 
            WHERE profile_number = ? AND schedule_date = ?
        ''', (profile_number, received_date)).fetchone()
        sales_order = schedule_entry['sales_order'] if schedule_entry else 'UNSCHEDULED'
        
        test_assigned = 'FINGERPRINT' 
        try: voc_percentage = float(profile['voc_percentage']) if profile and profile['voc_percentage'] is not None else 0.0
        except: voc_percentage = 0.0

        is_las_profile = False
        is_asbestos = False
        
        if profile:
            p_dict = dict(profile)
            win_code = str(p_dict.get('win_code', '')).strip().upper()
            
            if 'CNIA' in win_code or 'CNIA' in profile_number:
                is_asbestos = True
                
            
            # --- PURE EXPIRATION DATE & STATUS LAS LOGIC ---
            raw_exp = str(p_dict.get('expiration_date') or '').strip().lower()
            clean_exp = re.sub(r'[^a-z0-9]', '', raw_exp)
            prof_status = str(p_dict.get('status') or '').strip().upper()
            
            is_las_profile = False
            
            # 1. No Date -> Check the Status!
            if clean_exp in ['nodate', '', 'blank']:
                if prof_status == 'A':
                    is_las_profile = False  # Active + No Date = Safe
                else:
                    is_las_profile = True   # Not Active + No Date = LAS
                    
            # 2. "None" -> Safe, never expires
            elif clean_exp in ['none', 'nan', 'nat', 'null', 'na', 'tbd', '0', 'false']:
                is_las_profile = False
                
            # 3. Has a Date -> Expired Date ALWAYS triggers LAS
            else:
                if any(char.isdigit() for char in raw_exp):
                    try:
                        import pandas as pd
                        exp_date = pd.to_datetime(raw_exp, errors='coerce')
                        if pd.notna(exp_date) and exp_date < datetime.now():
                            is_las_profile = True
                    except: 
                        pass

        if is_asbestos or profile_number == 'BLCBPNONEB':
            is_las_profile = False

        if is_las_profile and previous_trucks == 0:
            test_assigned = 'LAS'

        if voc_percentage >= 50:
            if (previous_trucks + 1) % 10 == 0: test_assigned += " + VOC TEST"
            
        # Added sales_order and time_in to the INSERT statement
        conn.execute('''
            INSERT INTO truck_logs (
                truck_id, profile_number, manifest_number, load_number, 
                gross_weight, test_assigned, test_status, date_received, 
                sales_order, time_in
            )
            VALUES (?, ?, ?, ?, ?, ?, 'WEIGHED IN', ?, ?, ?)
        ''', (truck_id, profile_number, manifest_number, load_number, gross_weight, test_assigned, received_date, sales_order, time_in))
        conn.commit()
        
    return redirect(url_for('home'))

@app.route('/checkout_truck', methods=['POST'])
def checkout_truck():
    from datetime import datetime
    
    log_id = request.form.get('log_id')
    extra_fees_list = request.form.getlist('extra_fees')
    extra_fees = ", ".join(extra_fees_list) if extra_fees_list else "None"
    
    # NEW: Capture the exact check-out time
    time_out = datetime.now().strftime('%H:%M')
    
    try:
        exit_wt_raw = request.form.get('exit_weight', '0').replace(',', '')
        exit_weight = float(exit_wt_raw) if exit_wt_raw.strip() != '' else 0.0
        man_wt_raw = request.form.get('manifest_weight', '0').replace(',', '')
        manifest_weight = float(man_wt_raw) if man_wt_raw.strip() != '' else 0.0
    except: exit_weight, manifest_weight = 0.0, 0.0

    cell_location = request.form.get('cell_location', '')
    grid_location = request.form.get('grid_location', '')
    manifest_units = request.form.get('manifest_units', 'Pounds')

    with closing(get_db_connection()) as conn:
        truck = conn.execute('SELECT gross_weight FROM truck_logs WHERE id = ?', (log_id,)).fetchone()
        gross_weight = float(truck['gross_weight']) if truck and truck['gross_weight'] else 0.0
        
        # Keeps your critical weight validation!
        if exit_weight >= gross_weight and exit_weight > 0: 
            return "Critical Error: Tare weight cannot be greater than or equal to Gross weight.", 400
            
        net_weight_tons = (gross_weight - exit_weight) / 2000.0
        
        # NEW: Added time_out to the UPDATE statement
        conn.execute('''
            UPDATE truck_logs 
            SET exit_weight = ?, net_weight = ?, cell_location = ?, grid_location = ?,
                manifest_weight = ?, manifest_units = ?, extra_fees = ?, test_status = 'COMPLETED', time_out = ? 
            WHERE id = ?
        ''', (exit_weight, net_weight_tons, cell_location, grid_location, manifest_weight, manifest_units, extra_fees, time_out, log_id))
        conn.commit()
        
    return redirect(url_for('home'))

@app.route('/reject_truck', methods=['POST'])
def reject_truck():
    with closing(get_db_connection()) as conn:
        conn.execute("UPDATE truck_logs SET exit_weight = NULL, net_weight = 0, test_status = 'REJECTED', rejection_reason = ? WHERE id = ?", (request.form.get('rejection_reason').strip(), request.form.get('log_id')))
        conn.commit()
    return redirect(url_for('home'))

@app.route('/edit_truck/<int:log_id>', methods=['POST'])
def edit_truck(log_id):
    manifest_number = request.form.get('manifest_number', '').strip()
    load_number = request.form.get('load_number', '').strip()
    profile_number = request.form.get('profile_number', '').strip().upper()
    time_in = request.form.get('time_in', '').strip()
    time_out = request.form.get('time_out', '').strip()
    
    try: gross_weight = float(request.form.get('gross_weight', '0').replace(',', ''))
    except: gross_weight = 0.0
    
    source = request.form.get('source', 'reports')
    
    with closing(get_db_connection()) as conn:
        if source == 'receiving':
            # --- ACTIVE TRUCK EDIT (RECEIVING SCREEN) ---
            conn.execute('''
                UPDATE truck_logs 
                SET manifest_number = ?, load_number = ?, profile_number = ?, gross_weight = ?, time_in = ?
                WHERE id = ?
            ''', (manifest_number, load_number, profile_number, gross_weight, time_in, log_id))
            conn.commit()
            return redirect(url_for('home'))
            
        else:
            # --- COMPLETED TRUCK EDIT (REPORTS SCREEN) ---
            exit_weight_raw = request.form.get('exit_weight', '0')
            try: exit_weight = float(exit_weight_raw.replace(',', '')) if exit_weight_raw.strip() else 0.0
            except: exit_weight = 0.0
            
            cell_location = request.form.get('cell_location', '')
            grid_location = request.form.get('grid_location', '')
            date_received = request.form.get('date_received', date.today().isoformat())
            
            net_weight_tons = (gross_weight - exit_weight) / 2000.0 if exit_weight > 0 else 0.0
            
            conn.execute('''
                UPDATE truck_logs 
                SET manifest_number = ?, load_number = ?, profile_number = ?, 
                    gross_weight = ?, exit_weight = ?, net_weight = ?, cell_location = ?, grid_location = ?,
                    time_in = ?, time_out = ?
                WHERE id = ?
            ''', (manifest_number, load_number, profile_number, gross_weight, exit_weight, net_weight_tons, cell_location, grid_location, time_in, time_out, log_id))
            conn.commit()
            return redirect(url_for('reports', date=date_received))

@app.route('/delete_truck/<int:log_id>', methods=['POST'])
def delete_truck(log_id):
    date_received = request.form.get('date_received', date.today().isoformat())
    with closing(get_db_connection()) as conn:
        conn.execute('DELETE FROM truck_logs WHERE id = ?', (log_id,))
        conn.commit()
    return redirect(url_for('reports', date=date_received))

@app.route('/export')
def export_excel():
    selected_date = request.args.get('date', date.today().isoformat())

    with closing(get_db_connection()) as conn:
        # Pull all completed/weighed-out trucks for the selected date
        trucks_raw = conn.execute('''
            SELECT t.*, p.voc_percentage
            FROM truck_logs t
            LEFT JOIN profiles p ON t.profile_number = p.profile_number
            WHERE t.date_received = ? AND t.test_status != 'REJECTED' AND t.exit_weight IS NOT NULL
            ORDER BY CAST(t.load_number AS INTEGER) ASC
        ''', (selected_date,)).fetchall()

    # Lists to hold data for each specific Excel tab
    wmu35_data = []
    non_voc_data = []
    gallons_data = []

    for row in trucks_raw:
        t = dict(row)
        wmu = str(t.get('cell_location', '')).strip().upper()
        
        # Safely parse VOC
        try: voc = float(t.get('voc_percentage'))
        except (ValueError, TypeError): voc = 0.0
            
        # Safely calculate Net LBS from Gross and Tare
        try:
            gross = float(t.get('gross_weight') or 0)
            tare = float(t.get('exit_weight') or 0)
            lbs = gross - tare
        except ValueError:
            lbs = 0.0
            
        # Do the Tonnage and VOC math
        tons = lbs / 2000.0 if lbs else 0.0
        voc_x_wt = voc * tons
        
        # Build the standard base row (Removed EXTRA FEES & RECEIVED, moved Load # to front)
        base_row = {
            'Weighmaster Load No.': t.get('load_number', ''),
            'VOCs (ppm)': voc,
            'Weight (Pounds)': lbs,
            'Weight (TONs)': round(tons, 2),
            'VOCSxWt.': round(voc_x_wt, 2),
            'MANIFEST #': t.get('manifest_number', ''),
            'APPROVAL #': t.get('profile_number', ''),
            'WMU': wmu
        }

        # ---------------------------------------------------------
        # SHEET 1: WMU 35 VOC TRACKING (All 35 loads)
        # ---------------------------------------------------------
        if wmu.startswith('35'):
            wmu35_data.append(base_row.copy())
            
        # ---------------------------------------------------------
        # SHEET 2: DAILY NON-VOC'S (WMU 31, STU, BAY, and WMU 35 0-VOCs)
        # ---------------------------------------------------------
        # Added 'BAY' to the check list
        is_stu = wmu.startswith('34') or wmu.startswith('STU') or 'BAY' in wmu or wmu in ['CCS', 'CCSF', 'CCSM']
        
        if wmu.startswith('31') or is_stu or (wmu.startswith('35') and voc == 0):
            non_voc_row = base_row.copy()
            # Rename columns to match the exact format of the Non-VOC sheet
            non_voc_row['Pounds'] = non_voc_row.pop('Weight (Pounds)')
            non_voc_row['TONs'] = non_voc_row.pop('Weight (TONs)')
            non_voc_row['VOCS/Wt.'] = non_voc_row.pop('VOCSxWt.')
            non_voc_data.append(non_voc_row)
            
        # ---------------------------------------------------------
        # SHEET 3: GALLONS CALCULATOR (WMU 31 loads only)
        # ---------------------------------------------------------
        if wmu.startswith('31'):
            # Convert LBS to Gallons (Assuming water weight ~8.34 lbs/gal)
            gallons = (lbs / 8.34) if lbs else 0.0 
            gallons_data.append({
                'SPECIFIC GRAVITY': 1.0,  # Defaulting to 1.0, can be adjusted
                'NET (LBS) WEIGHT OF LOAD': lbs,
                'GALLONS PER LOAD': round(gallons, 2),
                'WEIGHT TICKET #': t.get('load_number', ''),
                'WMU #': wmu
            })

    # Generate Excel File in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        # Helper function to generate sheets with Totals and auto-width
        def write_formatted_sheet(data, sheet_name, columns, sum_cols):
            df = pd.DataFrame(data, columns=columns) if data else pd.DataFrame(columns=columns)
            
            # Append a Totals Row at the bottom
            if not df.empty and sum_cols:
                totals = {col: '' for col in columns}
                totals[columns[0]] = 'TOTAL'
                for col in sum_cols:
                    if col in df.columns:
                        totals[col] = df[col].sum()
                df = pd.concat([df, pd.DataFrame([totals])], ignore_index=True)
            
            # Write to Excel
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Auto-adjust column widths for a polished look
            worksheet = writer.sheets[sheet_name]
            for idx, col in enumerate(df.columns):
                max_len = max(df[col].astype(str).map(len).max(), len(str(col))) + 2
                worksheet.column_dimensions[chr(65 + idx)].width = min(max_len, 30)
        
        # 1. Write WMU 35 Sheet
        write_formatted_sheet(wmu35_data, 'WMU 35 VOC TRACKING', 
                    ['Weighmaster Load No.', 'VOCs (ppm)', 'Weight (Pounds)', 'Weight (TONs)', 'VOCSxWt.', 'MANIFEST #', 'APPROVAL #', 'WMU'],
                    ['Weight (Pounds)', 'Weight (TONs)', 'VOCSxWt.'])
                    
        # 2. Write NON-VOC Sheet
        write_formatted_sheet(non_voc_data, "DAILY NON-VOC'S REPORT", 
                    ['Weighmaster Load No.', 'VOCs (ppm)', 'Pounds', 'TONs', 'VOCS/Wt.', 'MANIFEST #', 'APPROVAL #', 'WMU'],
                    ['Pounds', 'TONs', 'VOCS/Wt.'])
                    
        # 3. Write GALLONS Sheet
        write_formatted_sheet(gallons_data, 'GALLONS CALCULATOR', 
                    ['SPECIFIC GRAVITY', 'NET (LBS) WEIGHT OF LOAD', 'GALLONS PER LOAD', 'WEIGHT TICKET #', 'WMU #'],
                    ['NET (LBS) WEIGHT OF LOAD', 'GALLONS PER LOAD'])

    output.seek(0)
    
    # Return the dynamic multi-sheet workbook to the user
    return send_file(
        output, 
        as_attachment=True, 
        download_name=f"Daily_Tracking_Logs_{selected_date}.xlsx", 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
@app.route('/reports')
def reports():
    today_str = date.today().isoformat()
    selected_date = request.args.get('date', today_str)
    
    with closing(get_db_connection()) as conn:
        all_profiles = conn.execute('SELECT profile_number, voc_percentage FROM profiles').fetchall()
        voc_dict = {p['profile_number']: p['voc_percentage'] for p in all_profiles}
        
        sched_data = conn.execute("SELECT profile_number, SUM(load_count) as sched_loads FROM daily_schedule WHERE schedule_date = ? GROUP BY profile_number", (selected_date,)).fetchall()
        schedule_dict = {row['profile_number']: row['sched_loads'] for row in sched_data}
        total_scheduled = sum(schedule_dict.values())
        
        # Pulls truck counts and exact tonnage sent to Unit 35
        actual_data = conn.execute('''
            SELECT 
                t.profile_number, 
                COUNT(t.id) as total_trucks, 
                SUM(CASE WHEN t.cell_location LIKE '35%' THEN t.net_weight ELSE 0 END) as total_tons, 
                MAX(p.voc_percentage) as profile_voc 
            FROM truck_logs t 
            LEFT JOIN profiles p ON t.profile_number = p.profile_number 
            WHERE t.exit_weight IS NOT NULL AND t.date_received = ? 
            GROUP BY t.profile_number
        ''', (selected_date,)).fetchall()
        
        truck_logs_raw = conn.execute('''
            SELECT t.*, p.generator
            FROM truck_logs t
            LEFT JOIN profiles p ON t.profile_number = p.profile_number
            WHERE t.date_received = ? AND t.test_status != 'REJECTED'
            ORDER BY CAST(t.load_number AS INTEGER) ASC
        ''', (selected_date,)).fetchall()
        truck_logs = [dict(row) for row in truck_logs_raw]
        
    master_report, grand_trucks, grand_tons = [], 0, 0.0
    
    # --- NEW: TONNAGE-WEIGHTED TRACKING VARIABLES ---
    voc_x_tons_sum = 0.0
    unit_35_total_tons = 0.0
    
    actual_profiles = set()
    for a in actual_data:
        prof = a['profile_number']
        actual_profiles.add(prof)
        sched = schedule_dict.get(prof, 0)
        
        try:
            safe_voc = float(a['profile_voc'])
        except (ValueError, TypeError):
            safe_voc = 0.0
            
        unit_35_tons = a['total_tons'] or 0.0
        
        master_report.append({
            'profile_number': prof, 
            'total_trucks': a['total_trucks'], 
            'total_tons': unit_35_tons, 
            'avg_voc_ppm': safe_voc, 
            'scheduled': sched, 
            'variance': a['total_trucks'] - sched
        })
        
        grand_trucks += a['total_trucks']
        grand_tons += unit_35_tons
        
        # -------------------------------------------------------------
        # EXACT EXCEL MATH: (VOC x Tons) / Total Tons
        # -------------------------------------------------------------
        if unit_35_tons > 0: 
            voc_x_tons_sum += (safe_voc * unit_35_tons) 
            unit_35_total_tons += unit_35_tons
            
    for s in sched_data:
        prof = s['profile_number']
        if prof not in actual_profiles: 
            try:
                safe_voc = float(voc_dict.get(prof, 0.0))
            except (ValueError, TypeError):
                safe_voc = 0.0
                
            master_report.append({'profile_number': prof, 'total_trucks': 0, 'total_tons': 0.0, 'avg_voc_ppm': safe_voc, 'scheduled': s['sched_loads'], 'variance': -s['sched_loads']})
            
    master_report.sort(key=lambda x: (x['scheduled'], x['total_trucks']), reverse=True)
            
    # Calculate Final Tonnage-Weighted Average
    final_avg_voc = (voc_x_tons_sum / unit_35_total_tons) if unit_35_total_tons > 0 else 0.0
    
    grand_totals = {
        'grand_trucks': grand_trucks, 
        'grand_tons': grand_tons, 
        'grand_avg_voc_ppm': final_avg_voc, 
        'grand_scheduled': total_scheduled
    }
    
    trucks_by_profile = defaultdict(list)
    for t in truck_logs:
        trucks_by_profile[t['profile_number']].append(t)

    # ---------------------------------------------------------
    # GAP FILLING LOGIC FOR PRINTED COVER SHEET (SAFEGUARDED)
    # ---------------------------------------------------------
    filled_trucks = []
    if truck_logs:
        def get_load_int(t):
            try: return int(t['load_number'])
            except: return -1
            
        numbered_trucks = [t for t in truck_logs if get_load_int(t) > 0]
        
        if numbered_trucks:
            truck_dict = {get_load_int(t): dict(t) for t in numbered_trucks}
            sorted_load_nums = sorted(truck_dict.keys())
            
            for idx, current_num in enumerate(sorted_load_nums):
                # 1. Add the actual truck
                filled_trucks.append(truck_dict[current_num])
                
                # 2. Check the gap to the NEXT truck
                if idx < len(sorted_load_nums) - 1:
                    next_num = sorted_load_nums[idx + 1]
                    gap = next_num - current_num - 1
                    
                    if 0 < gap <= 20:
                        # Safe to fill small missing gaps individually
                        for i in range(current_num + 1, next_num):
                            filled_trucks.append({
                                'load_number': str(i),
                                'manifest_number': '---',
                                'profile_number': 'VOID / NOT ISSUED',
                                'sales_order': '',
                                'generator': 'Awaiting Entry / Missing Ticket',
                                'gross_weight': None,
                                'exit_weight': None,
                                'net_weight': None,
                                'cell_location': '---',
                                'grid_location': '---',
                                'time_in': '',
                                'time_out': '',
                                'test_status': 'VOID'
                            })
                    elif gap > 20:
                        # DANGER: Huge gap detected (likely a test, typo, or late entry). 
                        # Print ONE summary row to prevent crashing the server.
                        filled_trucks.append({
                            'load_number': f"{current_num + 1} ➔ {next_num - 1}",
                            'manifest_number': '⚠️ LARGE GAP',
                            'profile_number': 'SYSTEM SAFEGUARD',
                            'sales_order': '',
                            'generator': 'Multiple load numbers skipped to prevent system lag.',
                            'gross_weight': None,
                            'exit_weight': None,
                            'net_weight': None,
                            'cell_location': '---',
                            'grid_location': '---',
                            'time_in': '',
                            'time_out': '',
                            'test_status': 'VOID'
                        })
                        
        # Add any letters/non-numbers (like "12A") at the very end just in case
        for t in truck_logs:
            if get_load_int(t) == -1:
                filled_trucks.append(dict(t))

    return render_template('reports.html', 
                           report_data=master_report, 
                           grand_totals=grand_totals, 
                           selected_date=selected_date, 
                           today_str=today_str, 
                           trucks_by_profile=trucks_by_profile, 
                           all_trucks=filled_trucks)

@app.route('/update_lab', methods=['POST'])
def update_lab():
    with closing(get_db_connection()) as conn:
        conn.execute("UPDATE truck_logs SET lab_results = ?, test_status = 'LAB COMPLETED' WHERE id = ?", (request.form.get('lab_results'), request.form.get('log_id')))
        conn.commit()
    return redirect(url_for('home'))

@app.route('/chemist')
def chemist_dashboard():
    with closing(get_db_connection()) as conn:
        pending_lab_trucks = conn.execute("SELECT * FROM truck_logs WHERE (test_assigned LIKE '%FINGERPRINT%' OR test_assigned LIKE '%VOC TEST%') AND test_status = 'WEIGHED IN'").fetchall()
    return render_template('chemist.html', pending_trucks=pending_lab_trucks)

# --- SCHEDULE & RECURRING ROUTES ---
@app.route('/schedule')
def schedule_portal():
    selected_date = request.args.get('date', date.today().isoformat())
    with closing(get_db_connection()) as conn:
        
        # Order by order_index so manual drag sequences are locked down
        daily_loads_raw = conn.execute('''
            SELECT ds.*, p.expiration_date, p.status AS profile_status, p.special_handling 
            FROM daily_schedule ds
            LEFT JOIN profiles p ON ds.profile_number = p.profile_number
            WHERE ds.schedule_date = ? 
            ORDER BY ds.order_index ASC, ds.id ASC
        ''', (selected_date,)).fetchall()
        
        total_loads_query = conn.execute('SELECT SUM(load_count) FROM daily_schedule WHERE schedule_date = ?', (selected_date,)).fetchone()
        total_loads = total_loads_query[0] if total_loads_query[0] else 0

    daily_loads = []
    for row in daily_loads_raw:
        load = dict(row)
        notes = load['special_notes'] or ""
        
        win_code = str(load.get('routing_code', '')).strip().upper()
        profile_num = str(load.get('profile_number', '')).strip().upper()
        is_asbestos = ('CNIA' in win_code) or ('CNIA' in profile_num)
        
        # --- PURE EXPIRATION DATE & STATUS LAS LOGIC ---
        is_las = False 
        raw_exp = str(load.get('expiration_date') or '').strip().lower()
        clean_exp = re.sub(r'[^a-z0-9]', '', raw_exp)
        prof_status = str(load.get('profile_status') or '').strip().upper()
        special_handling = str(load.get('special_handling') or '').strip().upper()
        
        # CNIA profiles with NONE expiration dates do not trigger LAS
        is_asbestos_trigger = is_asbestos
        if is_asbestos_trigger and clean_exp in ['nodate', '', 'blank', 'none', 'nan', 'nat', 'null', 'na', 'tbd', '0', 'false']:
            is_asbestos_trigger = False

        # 0. Failsafe: Written LAS or Asbestos triggers LAS automatically
        if 'LAS' in special_handling or 'LAS' in str(notes).upper() or is_asbestos_trigger:
            is_las = True
        # 1. No Date Check
        elif clean_exp in ['nodate', '', 'blank']:
            if prof_status == 'A':
                is_las = False
            else:
                is_las = True
        # 2. Never Expires
        elif clean_exp in ['none', 'nan', 'nat', 'null', 'na', 'tbd', '0', 'false']:
            is_las = False  
        # 3. Explicit Date Expiry Parsing
        else:
            if any(char.isdigit() for char in raw_exp):
                try:
                    exp_date = pd.to_datetime(raw_exp, errors='coerce')
                    if pd.notna(exp_date) and exp_date < datetime.now():
                        is_las = True
                except:
                    pass

        if profile_num == 'BLCBPNONEB':
            is_las = False

        load['is_las'] = is_las
        
        voc_val = str(load.get('voc_level', '')).strip()
        load['voc_level'] = 'TBD' if voc_val in ['None', '', '?', 'TBD'] else voc_val
        
        # Clean display notes
        notes = re.sub(r'(?i)\bNOT CERCLA\b', '', notes)
        notes = re.sub(r'(?i)\bCERCLA\b', '', notes)
        notes = re.sub(r'(?i)\bLAS\b', '', notes)
        notes = re.sub(r'^[,\s]+|[,\s]+$', '', notes) 
        notes = re.sub(r',\s*,', ',', notes)
        load['clean_notes'] = notes.strip()
        
        daily_loads.append(load)

    standard_loads = []
    blcb_loads = []
    
    for load in daily_loads:
        try: load['order_index'] = int(load.get('order_index') or 0)
        except: load['order_index'] = 0
            
        try: load['id'] = int(load.get('id') or 0)
        except: load['id'] = 0
        
        try: load['load_count'] = int(load.get('load_count') or 1)
        except: load['load_count'] = 1

        try: load['is_pinned'] = int(load.get('is_pinned') or 0)
        except: load['is_pinned'] = 0

        if str(load['profile_number']).strip().upper() == 'BLCBPNONEB':
            blcb_loads.append(load)
        else:
            standard_loads.append(load)

    # --- BULLETPROOF SORT SEQUENCE ---
    standard_loads.sort(key=lambda x: (
        x['order_index'] != 0,  # 1. Inbox rows (0) stay locked on top
        not x.get('is_las', False) if x['order_index'] == 0 else x['order_index'], # 2. Inbox defaults to LAS first
        -x['is_pinned'],        # 3. Pinned profiles are anchored stably above unpinned items
        -x['load_count'] if x['order_index'] == 0 else 0, # 4. Secondary sort by load count
        -x['id']                # 5. Fallback safety
    ))

    return render_template('schedule.html', daily_loads=standard_loads, blcb_loads=blcb_loads, selected_date=selected_date, total_loads=total_loads)

@app.route('/update_schedule_order', methods=['POST'])
def update_schedule_order():
    order_data = request.json.get('order', [])
    date_str = request.args.get('date')
    new_timestamp = time.time()
    
    with closing(get_db_connection()) as conn:
        for item in order_data:
            # Shift order by 1 to lock it into Bucket A (Pinned)
            new_order = item['order'] + 1 
            conn.execute('UPDATE daily_schedule SET order_index = ? WHERE id = ?', (new_order, item['id']))
            
            # --- THE SERIES_ID OVERRIDE HAS BEEN DELETED FROM HERE! ---
            
        conn.commit()
        
    if date_str:
        SCHEDULE_UPDATES[date_str] = new_timestamp
    SCHEDULE_UPDATES['GLOBAL'] = new_timestamp
        
    return jsonify({
        'status': 'success',
        'new_date_timestamp': new_timestamp,
        'new_global_timestamp': new_timestamp
    })

@app.route('/toggle_pin/<int:schedule_id>', methods=['POST'])
def toggle_pin(schedule_id):
    date_str = request.args.get('date')
    with closing(get_db_connection()) as conn:
        row = conn.execute('SELECT is_pinned FROM daily_schedule WHERE id = ?', (schedule_id,)).fetchone()
        if row:
            # Flip it: if 1 make it 0, if 0 make it 1
            new_status = 0 if row['is_pinned'] == 1 else 1
            conn.execute('UPDATE daily_schedule SET is_pinned = ? WHERE id = ?', (new_status, schedule_id))
            conn.commit()

    if date_str: SCHEDULE_UPDATES[date_str] = time.time()
    SCHEDULE_UPDATES['GLOBAL'] = time.time()
    return jsonify({'status': 'success'})

@app.route('/clear_all_pins', methods=['POST'])
def clear_all_pins():
    date_str = request.form.get('schedule_date')
    if date_str:
        with closing(get_db_connection()) as conn:
            # Set everyone back to Unpinned (0) for the day
            conn.execute('UPDATE daily_schedule SET is_pinned = 0 WHERE schedule_date = ?', (date_str,))
            conn.commit()
            
        SCHEDULE_UPDATES[date_str] = time.time()
        SCHEDULE_UPDATES['GLOBAL'] = time.time()
        
    return redirect(url_for('schedule_portal', date=date_str))

@app.route('/add_schedule', methods=['POST'])
def add_schedule():
    selected_dates_raw = request.form.get('selected_dates') # Pass comma separated string (e.g. "2026-05-21, 2026-05-22")
    start_date_str = request.form.get('schedule_date')     # Fallback date input
    
    dates_to_schedule = []
    if selected_dates_raw:
        dates_to_schedule = [d.strip() for d in selected_dates_raw.split(',') if d.strip()]
    elif start_date_str:
        dates_to_schedule = [start_date_str]
        
    if not dates_to_schedule:
        return redirect(url_for('schedule_portal'))

    recurrence = request.form.get('recurrence', 'none')
    occurrences = int(request.form.get('occurrences', 1)) if recurrence != 'none' else 1
    
    final_dates = []
    if selected_dates_raw:
        final_dates = dates_to_schedule
    else:
        # Legacy recurrence engine fallback if calendar picker wasn't used
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d')
            for i in range(occurrences):
                if recurrence == 'daily':
                    current_date = start_date + timedelta(days=i)
                elif recurrence == 'weekly':
                    current_date = start_date + timedelta(weeks=i)
                else:
                    current_date = start_date
                final_dates.append(current_date.strftime('%Y-%m-%d'))
        except:
            final_dates = [start_date_str]

    with closing(get_db_connection()) as conn:
        for date_str in final_dates:
            generator_val = request.form.get('generator') or request.form.get('customer_name') or ''
            conn.execute('''
                INSERT INTO daily_schedule (
                    schedule_date, start_time, end_time, profile_number, load_count, 
                    customer_name, generator, waste_type, sales_order, routing_code, scheduler_initials, special_notes, voc_level, order_index, is_pinned
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, 0)
            ''', (
                date_str, 'TBD', 'TBD', 
                request.form.get('profile_number'), int(request.form.get('load_count', 1)),
                generator_val, generator_val, request.form.get('waste_type', 'WASTE PICKUP'),
                request.form.get('sales_order'), request.form.get('routing_code'),
                request.form.get('scheduler_initials'), request.form.get('special_notes'),
                request.form.get('voc_level', 0)
            ))
        conn.commit()
        
    if dates_to_schedule:
        for date_str in dates_to_schedule:
            SCHEDULE_UPDATES[date_str] = time.time()
    SCHEDULE_UPDATES['GLOBAL'] = time.time()
        
    return redirect(url_for('schedule_portal', date=dates_to_schedule[0] if dates_to_schedule else start_date_str))
@app.route('/delete_schedule/<int:schedule_id>', methods=['POST'])
def delete_schedule(schedule_id):
    apply_to_series = request.form.get('apply_to_series')
    schedule_date = request.form.get('schedule_date')
    
    with closing(get_db_connection()) as conn:
        if apply_to_series:
            row = conn.execute("SELECT series_id FROM daily_schedule WHERE id = ?", (schedule_id,)).fetchone()
            if row and row['series_id']:
                conn.execute("DELETE FROM daily_schedule WHERE series_id = ?", (row['series_id'],))
                SCHEDULE_UPDATES['GLOBAL'] = time.time()
            else:
                conn.execute("DELETE FROM daily_schedule WHERE id = ?", (schedule_id,))
        else:
            conn.execute("DELETE FROM daily_schedule WHERE id = ?", (schedule_id,))
            
        conn.commit()
        
    if schedule_date:
        SCHEDULE_UPDATES[schedule_date] = time.time()
        
    return redirect(url_for('schedule_portal', date=schedule_date))

@app.route('/edit_schedule/<int:id>', methods=['POST'])
def edit_schedule(id):
    new_date = request.form.get('schedule_date')
    notes = request.form.get('special_notes', '')
    initials = request.form.get('scheduler_initials', '')
    voc_level = request.form.get('voc_level', 'TBD') 
    apply_to_series = request.form.get('apply_to_series') == 'on'
    
    profile_number = request.form.get('profile_number', '').upper()
    waste_type = request.form.get('waste_type', '')
    sales_order = request.form.get('sales_order', '')
    
    try: load_count = int(request.form.get('load_count') or 1)
    except: load_count = 1

    with closing(get_db_connection()) as conn:
        old_entry = conn.execute('SELECT schedule_date, series_id, order_index FROM daily_schedule WHERE id = ?', (id,)).fetchone()
        fallback_date = old_entry['schedule_date'] if old_entry else date.today().isoformat()
        
        if not new_date:
            new_date = fallback_date

        # FIXED: Only send to the top (0) if the date was changed. Otherwise, leave it exactly where it is!
        new_order_index = old_entry['order_index'] if (old_entry and new_date == fallback_date) else 0

        if apply_to_series and old_entry and old_entry['series_id']:

            conn.execute('''
                UPDATE daily_schedule 
                SET special_notes = ?, scheduler_initials = ?, voc_level = ?, profile_number = ?, load_count = ?, waste_type = ?, sales_order = ?, order_index = ?
                WHERE series_id = ?
            ''', (notes, initials, voc_level, profile_number, load_count, waste_type, sales_order, new_order_index, old_entry['series_id']))
            
            conn.execute('''
                UPDATE daily_schedule 
                SET schedule_date = ?, order_index = ?
                WHERE id = ?
            ''', (new_date, new_order_index, id))
        else:
            conn.execute('''
                UPDATE daily_schedule 
                SET schedule_date = ?, special_notes = ?, scheduler_initials = ?, voc_level = ?, profile_number = ?, load_count = ?, waste_type = ?, sales_order = ?, order_index = ?
                WHERE id = ?
            ''', (new_date, notes, initials, voc_level, profile_number, load_count, waste_type, sales_order, new_order_index, id))
            
        SCHEDULE_UPDATES['GLOBAL'] = time.time()
        SCHEDULE_UPDATES[new_date] = time.time()
        if new_date != fallback_date:
            SCHEDULE_UPDATES[fallback_date] = time.time() 
            
        conn.commit()

    return redirect(url_for('schedule_portal', date=new_date))

@app.route('/reset_schedule_sort', methods=['POST'])
def reset_schedule_sort():
    if request.is_json:
        date_str = request.json.get('schedule_date')
    else:
        date_str = request.form.get('schedule_date')
        
    if date_str:
        with closing(get_db_connection()) as conn:
            daily_loads_raw = conn.execute('''
                SELECT ds.id, ds.load_count, ds.profile_number, ds.routing_code, ds.order_index, ds.is_pinned, p.expiration_date, p.status AS profile_status 
                FROM daily_schedule ds
                LEFT JOIN profiles p ON ds.profile_number = p.profile_number
                WHERE ds.schedule_date = ? 
            ''', (date_str,)).fetchall()
            
            pinned_trucks = []
            unpinned_trucks = []
            
            for row in daily_loads_raw:
                t = dict(row)
                try: t['load_count'] = int(t.get('load_count') or 1)
                except: t['load_count'] = 1
                try: t['order_index'] = int(t.get('order_index') or 0)
                except: t['order_index'] = 0
                
                # --- CALCULATE LAS FOR UNPINNED ITEMS ---
                is_las = False
                win_code = str(t.get('routing_code', '')).strip().upper()
                profile_num = str(t.get('profile_number', '')).strip().upper()
                is_asbestos = ('CNIA' in win_code) or ('CNIA' in profile_num)
                raw_exp = str(t.get('expiration_date') or '').strip().lower()
                
                import re
                clean_exp = re.sub(r'[^a-z0-9]', '', raw_exp)
                prof_status = str(t.get('profile_status') or '').strip().upper()
                
                if clean_exp in ['nodate', '', 'blank']:
                    if prof_status != 'A': is_las = True
                elif clean_exp in ['none', 'nan', 'nat', 'null', 'na', 'tbd', '0', 'false']:
                    is_las = False  
                else:
                    if any(char.isdigit() for char in raw_exp):
                        try:
                            import pandas as pd
                            exp_date = pd.to_datetime(raw_exp, errors='coerce')
                            if pd.notna(exp_date) and exp_date < datetime.now(): is_las = True
                        except: pass

                # CNIA profiles with NONE expiration dates do not trigger LAS
                is_asbestos_trigger = is_asbestos
                if is_asbestos_trigger and clean_exp in ['nodate', '', 'blank', 'none', 'nan', 'nat', 'null', 'na', 'tbd', '0', 'false']:
                    is_asbestos_trigger = False

                if is_asbestos_trigger:
                    is_las = True

                if profile_num == 'BLCBPNONEB':
                    is_las = False

                t['is_las'] = is_las
                
                # SPLIT INTO BUCKETS BASED ON REAL PIN STATUS
                if t.get('is_pinned') == 1:
                    pinned_trucks.append(t)
                else:
                    unpinned_trucks.append(t)
            
            # 1. Sort the PINNED trucks by exactly where you manually dragged them
            pinned_trucks.sort(key=lambda x: x['order_index'])
            
            # 2. Sort the UNPINNED trucks purely by Math (LAS first, then High Loads)
            unpinned_trucks.sort(key=lambda x: (not x['is_las'], -x['load_count']))
            
            # 3. FIXED: Interleave them! Lock Pinned trucks in their exact slots, flow unpinned around them.
            layout = {}
            for pt in pinned_trucks:
                pref = pt['order_index']
                if pref <= 0: pref = 1
                while pref in layout:
                    pref += 1
                layout[pref] = pt
                
            unpinned_idx = 0
            for slot in range(1, len(daily_loads_raw) + 1):
                if slot not in layout and unpinned_idx < len(unpinned_trucks):
                    layout[slot] = unpinned_trucks[unpinned_idx]
                    unpinned_idx += 1
                    
            next_slot = max(layout.keys()) + 1 if layout else 1
            while unpinned_idx < len(unpinned_trucks):
                layout[next_slot] = unpinned_trucks[unpinned_idx]
                next_slot += 1
                unpinned_idx += 1
                
            combined_list = [layout[k] for k in sorted(layout.keys())]
            
            # 4. Save the new visual layout
            for index, truck in enumerate(combined_list, start=1):
                conn.execute('UPDATE daily_schedule SET order_index = ? WHERE id = ?', (index, truck['id']))
                
            conn.commit()
        
        SCHEDULE_UPDATES[date_str] = time.time()
        SCHEDULE_UPDATES['GLOBAL'] = time.time()
        
    if request.is_json:
        return jsonify({'status': 'success'})
        
    return redirect(url_for('schedule_portal', date=date_str))

@app.route('/toggle_unscheduled/<int:id>', methods=['POST'])
def toggle_unscheduled(id):
    date_str = request.form.get('schedule_date')
    
    with closing(get_db_connection()) as conn:
        row = conn.execute('SELECT is_unscheduled FROM daily_schedule WHERE id = ?', (id,)).fetchone()
        
        if row:
            # Flip it: if 1 make it 0, if 0 make it 1
            new_status = 0 if row['is_unscheduled'] == 1 else 1
            conn.execute('UPDATE daily_schedule SET is_unscheduled = ? WHERE id = ?', (new_status, id))
            conn.commit()

    if date_str:
        SCHEDULE_UPDATES[date_str] = time.time()
    SCHEDULE_UPDATES['GLOBAL'] = time.time()

    return redirect(url_for('schedule_portal', date=date_str))

@app.route('/api/search_schedule')
def search_schedule():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])

    with closing(get_db_connection()) as conn:
        results = conn.execute('''
            SELECT schedule_date, profile_number, sales_order, generator, load_count
            FROM daily_schedule
            WHERE profile_number LIKE ? OR sales_order LIKE ? OR generator LIKE ?
            ORDER BY schedule_date DESC
            LIMIT 50
        ''', (f'%{query}%', f'%{query}%', f'%{query}%')).fetchall()

    return jsonify([dict(r) for r in results])

@app.route('/api/check_schedule_duplicate', methods=['POST'])
def check_schedule_duplicate():
    profile_number = request.json.get('profile_number', '').strip().upper()
    routing_code = request.json.get('routing_code', '').strip().upper()
    sales_order = request.json.get('sales_order', '').strip().upper()
    selected_dates_raw = request.json.get('selected_dates', '')
    
    # Extract targeted dates from multi-date calendar layout
    dates = [d.strip() for d in selected_dates_raw.split(',') if d.strip()]
    if not dates and request.json.get('schedule_date'):
        dates = [request.json.get('schedule_date')]
        
    with closing(get_db_connection()) as conn:
        for d in dates:
            if not routing_code and not sales_order:
                # STRICT PROFILE-LEVEL CHECK
                dup = conn.execute('''
                    SELECT id FROM daily_schedule 
                    WHERE schedule_date = ? AND UPPER(profile_number) = ?
                ''', (d, profile_number)).fetchone()
            else:
                dup = conn.execute('''
                    SELECT id FROM daily_schedule 
                    WHERE schedule_date = ? AND UPPER(profile_number) = ? 
                      AND UPPER(routing_code) = ? AND UPPER(sales_order) = ?
                ''', (d, profile_number, routing_code, sales_order)).fetchone()
            if dup:
                return jsonify({'duplicate': True, 'date': d})
                
    return jsonify({'duplicate': False})

# --- 11. THE STU PORTAL & SAMPLING INTEGRATION ---

@app.route('/stu/inventory')
def stu_inventory():
    category = request.args.get('category', 'All')
    with closing(get_db_connection()) as conn:
        queries = {
            'Decon': "SELECT * FROM drum_inventory WHERE process_type IN ('direct land haz', 'directlandasbes', 'asbestos') OR inb_prof = 'cnia' ORDER BY age DESC",
            'Solidification': "SELECT * FROM drum_inventory WHERE process_type = 'solidify normal' ORDER BY age DESC",
            'T_Drums': "SELECT * FROM drum_inventory WHERE process_type IN ('stabsolids 1', 'stabsolids 2') ORDER BY age DESC",
            'TL_Drums': "SELECT * FROM drum_inventory WHERE process_type IN ('stabsolids 3', 'stabsolids 5') ORDER BY age DESC",
            'Special_Handling': "SELECT * FROM drum_inventory WHERE process_type IN ('stabsolids 6', 'stabsolids 8', 'stabsolids 9', 'stabsolids 11', 'stabsolids 12', 'stabsolids 13', 'stabsolids 14') ORDER BY age DESC",
            'All': "SELECT * FROM drum_inventory ORDER BY age DESC"
        }
        drums = conn.execute(queries.get(category, queries['All'])).fetchall()
        last_upload_row = conn.execute('SELECT MAX(import_date) FROM drum_inventory').fetchone()
        last_upload = last_upload_row[0] if last_upload_row else 'NO DATA'
    return render_template('stu_inventory.html', drums=drums, category=category, last_upload=last_upload)

@app.route('/stu/sampling', methods=['GET', 'POST'])
def stu_sampling():
    if request.method == 'POST':
        if 'label_pdf' not in request.files: return "No file uploaded", 400
        file = request.files['label_pdf']
        if file.filename == '': return "No file selected", 400

        file_bytes = file.read()
        
        raw_drums = stu_services.parse_drum_labels_from_pdf(BytesIO(file_bytes))
        if not raw_drums:
            return "<script>alert('No valid drums found in PDF.'); window.history.back();</script>"

        if not os.path.exists('temp_uploads'):
            os.makedirs('temp_uploads')
            
        temp_filename = f"{uuid.uuid4().hex}.pdf"
        temp_filepath = os.path.join('temp_uploads', temp_filename)
        
        with open(temp_filepath, 'wb') as f:
            f.write(file_bytes)
        
        return render_template('stu_sampling.html', drums=raw_drums, pdf_filename=temp_filename, raw_drums_json=json.dumps(raw_drums))
        
    return render_template('stu_sampling.html', drums=None)

@app.route('/generate_sampling_packet', methods=['POST'])
def generate_sampling_packet():
    pdf_filename = request.form.get('pdf_filename')
    selected_drums_json = request.form.get('selected_drums')

    if not pdf_filename or not selected_drums_json:
        return f"Error: Data dropped.", 400

    temp_filepath = os.path.join('temp_uploads', pdf_filename)
    try:
        with open(temp_filepath, 'rb') as f:
            file_bytes = f.read()
    except FileNotFoundError:
        return "Temporary file lost. Please start over and re-upload the PDF.", 400

    raw_drums = json.loads(selected_drums_json)
    job_name = datetime.now().strftime("%m-%d-%Y_%H%M")

    with closing(get_db_connection()) as conn:
        picklist_data, total_samples = stu_services.process_drums(conn, raw_drums)
        
        for d in raw_drums:
            today_str = date.today().isoformat()
            
            existing = conn.execute("SELECT id FROM drum_inventory WHERE track_no = ?", (d['drum_id'],)).fetchone()
            if not existing:
                conn.execute('''
                    INSERT INTO drum_inventory (track_no, inb_prof, manifest, process_type, weight, ph, age, voc_ppm, voc_weight, import_date) 
                    VALUES (?, ?, ?, 'PENDING SAMPLING', 0, 0, 0, 0, 0, ?)
                ''', (d['drum_id'], d['profile'], d['manifest'], today_str))
                
        for p in picklist_data:
            if p.get('is_sampled') == 'Yes':
                
                existing_lab = conn.execute("SELECT id FROM drum_lab_queue WHERE drum_id = ? AND status != 'COMPLETED'", (p['drum_id'],)).fetchone()
                if not existing_lab:
                    conn.execute('''
                        INSERT INTO drum_lab_queue (job_id, drum_id, profile, manifest, tests_required, status)
                        VALUES (?, ?, ?, ?, ?, 'PENDING')
                    ''', (job_name, p['drum_id'], p['profile'], p['manifest'], p.get('voc_testing_trigger', 'FingerPrint')))
                
        conn.commit()

    memory_zip = BytesIO()
    with zipfile.ZipFile(memory_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        picklist_buffer = BytesIO()
        stu_services.create_pdf_report(picklist_buffer, f"Drum Sampling Pick List - {job_name}", picklist_data, total_samples)
        zf.writestr(f"Picklist_{job_name}.pdf", picklist_buffer.getvalue())

        lab_buffer = BytesIO()
        stu_services.create_lab_sheet_pdf(lab_buffer, job_name, picklist_data)
        zf.writestr(f"LabBenchSheet_{job_name}.pdf", lab_buffer.getvalue())

        df = pd.DataFrame(picklist_data)
        excel_io = BytesIO()
        with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
            cols = ["sample_num", "drum_id", "manifest", "manifest_line", "profile", "display_profile", "waste_code", "is_sampled"]
            exist_cols = [c for c in cols if c in df.columns]
            df[exist_cols].to_excel(writer, index=False, sheet_name='Data')
        zf.writestr(f"Data_{job_name}.xlsx", excel_io.getvalue())

        annotated_buffer = BytesIO()
        stu_services.create_annotated_pdf(BytesIO(file_bytes), annotated_buffer, picklist_data)
        if annotated_buffer.getvalue():
            zf.writestr(f"LABELS_{job_name}_Marked.pdf", annotated_buffer.getvalue())

    if os.path.exists(temp_filepath):
        os.remove(temp_filepath)

    memory_zip.seek(0)
    return send_file(memory_zip, mimetype='application/zip', as_attachment=True, download_name=f"Drum_Sampling_{job_name}.zip")

@app.route('/upload_vpi', methods=['POST'])
def upload_vpi():
    if 'vpi_file' not in request.files: return "No file uploaded", 400
    file = request.files['vpi_file']
    if file.filename == '': return "No file selected", 400
    try:
        df = pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]
        
        required = ['Track No', 'Process Type', 'Weight', 'pH', 'Inb Prof', 'Age', 'Type']
        if not all(col in df.columns for col in required): 
            return "Error: Uploaded CSV is missing required columns from WIN.", 400
            
        df['Process Type'] = df['Process Type'].astype(str).str.strip().str.lower()
        df['Inb Prof'] = df['Inb Prof'].astype(str).str.strip().str.lower()
        df['Type'] = df['Type'].astype(str).str.strip().str.lower() 
        
        df = df[~df['Process Type'].isin(['put pile', '=', 'nan', ''])]
        df = df[~df['Type'].str.contains('cm|dt', na=False)]
        
        df['Weight'] = pd.to_numeric(df['Weight'], errors='coerce').fillna(0)
        df['pH'] = pd.to_numeric(df['pH'], errors='coerce').fillna(0)
        df['Age'] = pd.to_numeric(df['Age'], errors='coerce').fillna(0)
        
        df = df.drop_duplicates(subset=['Track No'], keep='last')
        
        with closing(get_db_connection()) as conn:
            conn.execute('DELETE FROM drum_inventory')
            profiles_df = pd.read_sql_query("SELECT LOWER(profile_number) as profile_number, voc_percentage FROM profiles", conn)
            
            # Text values mapped from DB are safely inside the database block now
            voc_dict = dict(zip(profiles_df['profile_number'], profiles_df['voc_percentage']))
            df['voc_ppm'] = df['Inb Prof'].map(voc_dict).fillna(0)
            
            # Convert text flags like 'TBD' safely to numeric 0
            df['voc_ppm'] = pd.to_numeric(df['voc_ppm'], errors='coerce').fillna(0)
            df['voc_weight'] = df['Weight'] * df['voc_ppm']
            
            cleaned_data = list(zip(df['Track No'], df['Inb Prof'], df['Process Type'], df['Weight'], df['pH'], df['Age'], df['voc_ppm'], df['voc_weight'], [date.today().isoformat()]*len(df)))
            conn.executemany("INSERT INTO drum_inventory (track_no, inb_prof, process_type, weight, ph, age, voc_ppm, voc_weight, import_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)", cleaned_data)
            conn.commit()
    except Exception as e: 
        return f"Critical Error processing file: {e}", 500
    
    return redirect(url_for('stu_inventory'))


@app.route('/export_stu', methods=['POST'])
def export_stu():
    selected_ids = request.form.getlist('selected_drums')
    category = request.form.get('category_name', 'Custom')
    if not selected_ids: return "No drums selected for export", 400
        
    placeholders = ','.join('?' for _ in selected_ids)
    with closing(get_db_connection()) as conn:
        df = pd.read_sql_query(f"SELECT track_no, inb_prof, process_type, weight, ph, age, voc_ppm, voc_weight FROM drum_inventory WHERE id IN ({placeholders})", conn, params=selected_ids)
        
    if df.empty: return "No data found", 400
    df.rename(columns={'track_no': 'Track No', 'inb_prof': 'Inb Prof', 'process_type': 'Process Type', 'weight': 'Weight', 'ph': 'pH', 'age': 'Age', 'voc_ppm': 'VOC', 'voc_weight': 'VOC Weight'}, inplace=True)
    df.insert(0, ' ', '')
    
    output = BytesIO()
    clean_category = category.replace('_', ' ')
    date_str = date.today().strftime('%m-%d-%Y')
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1, sheet_name=f"{category[:31]}")
        worksheet = writer.sheets[f"{category[:31]}"]
        last_row = len(df) + 2  
        
        title_cell = worksheet.cell(row=1, column=1, value=f"{clean_category} - Generated On - {date.today().strftime('%m/%d/%Y')}")
        worksheet.merge_cells('A1:I1')
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num in range(1, 10):
            worksheet.cell(row=2, column=col_num).font = Font(size=16, bold=True, underline='double')
            worksheet.cell(row=2, column=col_num).alignment = Alignment(horizontal='center')
            
        worksheet.column_dimensions['A'].width = 6
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r in range(3, last_row + 1): 
            worksheet.cell(row=r, column=1).border = thin_border
            worksheet.cell(row=r, column=5).number_format = '0.00' 
            worksheet.cell(row=r, column=8).number_format = '0.00' 
            worksheet.cell(row=r, column=9).number_format = '0.00' 
            
        worksheet.cell(row=last_row + 2, column=3, value="Total Weight (lbs):").font = Font(bold=True)
        wt_tot = worksheet.cell(row=last_row + 2, column=4, value=f"=SUM(E3:E{last_row})")
        wt_tot.font = Font(bold=True)
        wt_tot.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 3, column=3, value="Total Weight (tons):").font = Font(bold=True)
        wt_tons = worksheet.cell(row=last_row + 3, column=4, value=f"=D{last_row + 2}/2000")
        wt_tons.font = Font(bold=True)
        wt_tons.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 2, column=7, value="Total VOC Weight:").font = Font(bold=True)
        voc_tot = worksheet.cell(row=last_row + 2, column=8, value=f"=SUM(I3:I{last_row})")
        voc_tot.font = Font(bold=True)
        voc_tot.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 3, column=7, value="Average VOC:").font = Font(bold=True)
        voc_avg = worksheet.cell(row=last_row + 3, column=8, value=f"=AVERAGE(H3:H{last_row})")
        voc_avg.font = Font(bold=True)
        voc_avg.number_format = '0.00'
        
        for col_idx, col in enumerate(worksheet.columns, start=1):
            if col_idx == 1: continue 
            max_length = 0
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            col_letter = worksheet.cell(row=2, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = max_length + 3

    output.seek(0)
    export_filename = f"{clean_category} {date_str}.xlsx"
    return send_file(output, as_attachment=True, download_name=export_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/api/parse_profile_pdf', methods=['POST'])
def parse_profile_pdf():
    if 'pdf_file' not in request.files: return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['pdf_file']
    if file.filename == '': return jsonify({'error': 'No file selected'}), 400

    extracted_data = {'profile_number': '', 'generator': '', 'waste_description': '', 'win_code': '', 'special_handling': ''}
    try:
        with pdfplumber.open(file) as pdf:
            full_text = "".join([page.extract_text() + "\n" for page in pdf.pages[:2]])
            prof_match = re.search(r'Clean Harbors Profile No\.\s*([A-Z0-9]+)', full_text)
            if prof_match: extracted_data['profile_number'] = prof_match.group(1).strip()
            gen_match = re.search(r'GENERATOR NAME:\s*(.*?)\n', full_text)
            if gen_match: extracted_data['generator'] = gen_match.group(1).strip()
            desc_match = re.search(r'CUSTOMER WASTE DESCRIPTION:\s*(.*?)\n', full_text)
            if desc_match: extracted_data['waste_description'] = desc_match.group(1).strip()
            form_code_match = re.search(r'SPECIFY THE FORM CODE.*?([A-Z][0-9]{3})', full_text, re.IGNORECASE | re.DOTALL)
            if form_code_match: extracted_data['win_code'] = form_code_match.group(1).strip()
        return jsonify(extracted_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/add_master_profile', methods=['POST'])
def add_master_profile():
    with closing(get_db_connection()) as conn:
        conn.execute('''
            REPLACE INTO profiles (profile_number, generator, waste_description, win_code, voc_percentage, special_handling, ph_range, physical_appearance, flash_point, expiration_date)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (request.form.get('profile_number').upper(), request.form.get('generator', ''), request.form.get('waste_description', ''), request.form.get('win_code', ''), request.form.get('voc_percentage', 0.0), request.form.get('special_handling', ''), request.form.get('ph_range', ''), request.form.get('physical_appearance', ''), request.form.get('flash_point', ''), request.form.get('expiration_date', '')))
        conn.commit()
    return redirect(url_for('approvals_portal'))

@app.route('/chemist/drums')
def chemist_drums():
    with closing(get_db_connection()) as conn:
        raw_queue = conn.execute("SELECT * FROM drum_lab_queue WHERE status != 'COMPLETED'").fetchall()
        
        def sort_priority(drum):
            test = str(drum['tests_required']).upper()
            if 'VOC' in test or 'CP1' in test or 'RECERT' in test:
                return 0 
            return 1 
            
        sorted_queue = sorted(raw_queue, key=sort_priority)
        
        jobs = defaultdict(list)
        for d in sorted_queue:
            jobs[d['job_id']].append(dict(d))
            
    return render_template('chemist_drums.html', jobs=jobs)

@app.route('/chemist/drums/update', methods=['POST'])
def update_drum_lab():
    job_id = request.form.get('job_id')
    drum_db_ids = request.form.getlist('drum_db_id')
    
    with closing(get_db_connection()) as conn:
        for d_id in drum_db_ids:
            status = request.form.get(f'status_{d_id}')
            notes = request.form.get(f'notes_{d_id}', '')
            flashpoint = request.form.get(f'flashpoint_{d_id}', '')
            cyanide = request.form.get(f'cyanide_{d_id}', '')
            sulfide = request.form.get(f'sulfide_{d_id}', '')
            oxidation = request.form.get(f'oxidation_{d_id}', '')
            
            try: ph_result = float(request.form.get(f'ph_{d_id}', 0))
            except: ph_result = 0.0
                
            try: voc_result = float(request.form.get(f'voc_{d_id}', 0))
            except: voc_result = 0.0
            
            conn.execute('''
                UPDATE drum_lab_queue 
                SET status = ?, ph_result = ?, voc_result = ?, flashpoint = ?, 
                    cyanide = ?, sulfide = ?, oxidation = ?, notes = ?
                WHERE id = ?
            ''', (status, ph_result, voc_result, flashpoint, cyanide, sulfide, oxidation, notes, d_id))
                    
        conn.commit()
    return redirect(url_for('chemist_drums'))

@app.route('/waste_acceptance')
def waste_acceptance():
    with closing(get_db_connection()) as conn:
        completed_labs = conn.execute("SELECT * FROM drum_lab_queue WHERE status = 'COMPLETED'").fetchall()
        
        staging_data = []
        for lab in completed_labs:
            related_drums = conn.execute('''
                SELECT * FROM drum_inventory 
                WHERE manifest = ? AND inb_prof = ? AND process_type = 'PENDING SAMPLING'
            ''', (lab['manifest'], lab['profile'])).fetchall()
            
            if related_drums:
                staging_data.append({
                    'lab_result': dict(lab),
                    'related_drums': [dict(d) for d in related_drums],
                    'drum_count': len(related_drums)
                })
                
    return render_template('waste_acceptance.html', staging_data=staging_data)

@app.route('/final_code', methods=['POST'])
def final_code():
    lab_id = request.form.get('lab_id')
    manifest = request.form.get('manifest')
    profile = request.form.get('profile')
    ph_val = request.form.get('ph_val')
    voc_val = request.form.get('voc_val')
    
    with closing(get_db_connection()) as conn:
        conn.execute('''
            UPDATE drum_inventory 
            SET process_type = 'TESTED', ph = ?, voc_ppm = ?
            WHERE manifest = ? AND inb_prof = ? AND process_type = 'PENDING SAMPLING'
        ''', (ph_val, voc_val, manifest, profile))
        
        conn.execute("UPDATE drum_lab_queue SET status = 'FINAL CODED' WHERE id = ?", (lab_id,))
        conn.commit()
        
    return redirect(url_for('waste_acceptance'))        

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)