from flask import Blueprint, render_template, request, redirect, url_for, jsonify, send_from_directory, send_file
import os
from werkzeug.utils import secure_filename
from datetime import date
from contextlib import closing
import pdfplumber
import re
import time
from database import get_db_connection
from shared_state import SCHEDULE_UPDATES

approvals_bp = Blueprint('approvals_bp', __name__)

@approvals_bp.route('/approvals')
def approvals_portal():
    selected = request.args.get('selected_profile')
    if selected:
        return redirect(url_for('approvals_bp.approvals_profiles', selected_profile=selected))
    return render_template('approvals.html', show_view='log')

@approvals_bp.route('/approvals/profiles')
def approvals_profiles():
    return render_template('approvals.html', show_view='profiles')

@approvals_bp.route('/waste_acceptance')
def waste_acceptance():
    return redirect(url_for('stu_bp.stu_hub', view='pipeline'))

@approvals_bp.route('/waste_acceptance/checklist/<job_id>')
def waste_acceptance_checklist(job_id):
    with closing(get_db_connection()) as conn:
        # Fetch all labs for this job_id, joining with profile_wvi and profiles
        labs = conn.execute('''
            SELECT q.*, 
                   w.ph_min, w.ph_max, w.sulfides, w.cyanide, w.free_liquids, w.flashpoint, w.voc_ppm, w.color AS w_color,
                   w.treatment_information, w.notes_revisions, w.physical_description, w.handling_instruction,
                   w.generator_name, w.waste_name,
                   p.win_code, p.generator AS p_generator, p.waste_description AS p_waste_description,
                   p.ph_range AS p_ph_range, p.flash_point AS p_flash_point, p.voc_percentage AS p_voc_percentage,
                   p.special_handling AS p_special_handling, p.cyanide AS p_cyanide, p.sulfide AS p_sulfide,
                   p.free_liquids AS p_free_liquids, p.physical_appearance AS p_physical_appearance,
                   p.treatment_recipe AS p_treatment_recipe, p.color AS p_color
            FROM drum_lab_queue q
            LEFT JOIN profile_wvi w ON TRIM(UPPER(q.profile)) = TRIM(UPPER(w.profile))
            LEFT JOIN profiles p ON TRIM(UPPER(q.profile)) = TRIM(UPPER(p.profile_number))
            WHERE q.job_id = ?
        ''', (job_id,)).fetchall()
        
        # Fetch all related physical drums in drum_inventory for this job
        related_drums = conn.execute('''
            SELECT * FROM drum_inventory 
            WHERE job_id = ? AND process_type = 'PENDING SAMPLING'
        ''', (job_id,)).fetchall()
        
    labs_list = []
    for row in labs:
        lab = dict(row)
        
        # --- FALLBACK TO MASTER PROFILE REGISTRY (EXTRACTED PDF DATA) ---
        if not lab.get('generator_name') and lab.get('p_generator'):
            lab['generator_name'] = lab['p_generator']
        if not lab.get('waste_name') and lab.get('p_waste_description'):
            lab['waste_name'] = lab['p_waste_description']
        if not lab.get('flashpoint') and lab.get('p_flash_point'):
            lab['flashpoint'] = lab['p_flash_point']
            
        # Set color
        lab['color'] = lab.get('w_color') or lab.get('p_color') or ''
        
        if lab.get('ph_min') is None and lab.get('ph_max') is None and lab.get('p_ph_range'):
            import re
            ph_str = str(lab['p_ph_range']).strip().upper()
            if "7 (NEUTRAL)" in ph_str or ph_str == "7" or "7 NEUTRAL" in ph_str:
                lab['ph_min'] = 4.0
                lab['ph_max'] = 10.0
            else:
                m = re.findall(r'(\d+\.?\d*)', ph_str)
                if len(m) >= 2:
                    try:
                        lab['ph_min'] = float(m[0])
                        lab['ph_max'] = float(m[1])
                    except:
                        pass
                elif len(m) == 1:
                    try:
                        lab['ph_min'] = float(m[0])
                        lab['ph_max'] = float(m[0])
                    except:
                        pass
        if lab.get('voc_ppm') is None and lab.get('p_voc_percentage') is not None:
            try:
                val = float(lab['p_voc_percentage'])
                lab['voc_ppm'] = val
            except:
                pass
        if not lab.get('sulfides') and lab.get('p_sulfide'):
            lab['sulfides'] = lab['p_sulfide']
        lab['sulfide'] = lab.get('sulfides')
        if not lab.get('cyanide') and lab.get('p_cyanide'):
            lab['cyanide'] = lab['p_cyanide']
        if not lab.get('free_liquids') and lab.get('p_free_liquids'):
            lab['free_liquids'] = lab['p_free_liquids']
        if not lab.get('physical_description') and lab.get('p_physical_appearance'):
            lab['physical_description'] = lab['p_physical_appearance']
        if not lab.get('treatment_information') and lab.get('p_treatment_recipe'):
            lab['treatment_information'] = lab['p_treatment_recipe']
        if not lab.get('handling_instruction') and lab.get('p_special_handling'):
            lab['handling_instruction'] = lab['p_special_handling']
            
        labs_list.append(lab)
        
    drums_list = [dict(d) for d in related_drums]
    
    # Associate physical drums with their representing lab queue sample
    associated_track_nos = set()
    for lab in labs_list:
        lab_manifest = str(lab['manifest']).strip().upper()
        lab_profile = str(lab['profile']).strip().upper()
        
        batch_drums = []
        for drum in drums_list:
            drum_manifest = str(drum['manifest']).strip().upper() if drum.get('manifest') else ""
            drum_profile = str(drum['inb_prof']).strip().upper() if drum.get('inb_prof') else ""
            
            if drum_manifest == lab_manifest and drum_profile == lab_profile:
                batch_drums.append(drum)
                associated_track_nos.add(drum['track_no'])
                
        lab['batch_drums'] = batch_drums
        
    # Filter out unsampled drums (e.g., asbestos or other low-risk profiles)
    unsampled_drums = [d for d in drums_list if d['track_no'] not in associated_track_nos]
        
    return render_template('waste_acceptance.html', 
                           job_id=job_id, 
                           labs=labs_list, 
                           related_drums=drums_list,
                           unsampled_drums=unsampled_drums)

@approvals_bp.route('/waste_acceptance/mark_coded', methods=['POST'])
def waste_acceptance_mark_coded():
    data = request.get_json() or {}
    lab_id = data.get('lab_id')
    coded = data.get('coded') # 1 or 0
    
    if lab_id is None or coded is None:
        return jsonify({'error': 'Missing lab_id or coded state'}), 400
        
    with closing(get_db_connection()) as conn:
        conn.execute('''
            UPDATE drum_lab_queue 
            SET coded_in_win = ? 
            WHERE id = ?
        ''', (int(coded), lab_id))
        
        # Fetch job_id to emit websocket update
        lab_row = conn.execute('SELECT job_id FROM drum_lab_queue WHERE id = ?', (lab_id,)).fetchone()
        job_id = lab_row['job_id'] if lab_row else None
        
        conn.commit()
        
        if job_id:
            from shared_state import socketio
            socketio.emit('drum_update', {'job_id': job_id})
            
    return jsonify({'success': True})

@approvals_bp.route('/waste_acceptance/finalize_load', methods=['POST'])
def waste_acceptance_finalize_load():
    job_id = request.form.get('job_id')
    if not job_id:
        return "Missing Load Number (job_id)", 400
        
    with closing(get_db_connection()) as conn:
        # Get all completed labs for this job_id
        labs = conn.execute('''
            SELECT * FROM drum_lab_queue 
            WHERE job_id = ? AND status = 'COMPLETED'
        ''', (job_id,)).fetchall()
        
        # 1. Update the drum_inventory drums that match the lab's manifest and profile
        for lab in labs:
            conn.execute('''
                UPDATE drum_inventory 
                SET process_type = 'TESTED', 
                    ph = ?, 
                    voc_ppm = ?, 
                    voc_weight = weight * ?
                WHERE job_id = ? AND manifest = ? AND inb_prof = ? AND process_type = 'PENDING SAMPLING'
            ''', (lab['ph_result'], lab['voc_result'], lab['voc_result'], job_id, lab['manifest'], lab['profile']))
            
        # 2. Update ALL remaining drums for this job_id that are PENDING SAMPLING to TESTED
        conn.execute('''
            UPDATE drum_inventory
            SET process_type = 'TESTED'
            WHERE job_id = ? AND process_type = 'PENDING SAMPLING'
        ''', (job_id,))
        
        # 3. Update the drum_lab_queue status to 'FINAL CODED' for this job_id
        conn.execute('''
            UPDATE drum_lab_queue 
            SET status = 'FINAL CODED' 
            WHERE job_id = ?
        ''', (job_id,))
        
        conn.commit()
        
        # Emit a socket update for the completed load
        from shared_state import socketio
        socketio.emit('drum_update', {'job_id': job_id})
        
    return redirect(url_for('stu_bp.stu_hub', view='pipeline'))

@approvals_bp.route('/api/parse_profile_pdf', methods=['POST'])
def parse_profile_pdf():
    if 'pdf_file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    file = request.files['pdf_file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
        
    extracted_data = {
        'profile_number': '', 'generator_name': '', 'waste_description': '', 
        'special_handling': '', 'physical_appearance': '',
        'epa_id': '', 'dot_description': '', 'state_waste_code': '', 
        'federal_waste_code': '', 'ph_range': '', 'flash_point': '',
        'cyanide': 'No', 'sulfide': 'No', 'free_liquids': 'No', 'ldr_required': 'No',
        'color': '', 'shipping_container_type': 'Containerized'
    }
    try:
        import pdfplumber
        import re
        import math
        
        def get_distance(x1, y1, x2, y2):
            return math.sqrt((x1-x2)**2 + (y1-y2)**2)

        def find_nearest_word(page, cm, max_dist=100):
            cx = (cm['x0'] + cm['x1']) / 2
            cy = (cm['top'] + cm['bottom']) / 2
            words = page.extract_words()
            closest = None
            min_d = max_dist
            for w in words:
                wx = (w['x0'] + w['x1']) / 2
                wy = (w['top'] + w['bottom']) / 2
                dist = get_distance(cx, cy, wx, wy)
                if dist < min_d:
                    min_d = dist
                    closest = w
            return closest, min_d

        with pdfplumber.open(file) as pdf:
            full_text = "".join([page.extract_text() + "\n" for page in pdf.pages if page.extract_text()])
            
            # Profile Number (supports suffixes like -B, -A, -1, etc.)
            prof_match = re.search(r'Profile No\.\s*([A-Z0-9\-_]+)', full_text, re.IGNORECASE)
            if not prof_match:
                prof_match = re.search(r'(?:Profile\s+No\.|Profile\s+#|Profile\s+Number[:\.\s]*)\s*([A-Z0-9\-_]+)', full_text, re.IGNORECASE)
            if prof_match:
                extracted_data['profile_number'] = prof_match.group(1).strip()
                
            # Generator Name
            gen_match = re.search(r'GENERATOR\s+NAME:\s*(.*?)(?:\n|$)', full_text, re.IGNORECASE)
            if gen_match:
                extracted_data['generator_name'] = gen_match.group(1).strip()
                
            # EPA ID
            epa_match = re.search(r'EPA\s+ID\s*(?:#/REGISTRATION\s*#)?\s*([A-Z0-9]+)', full_text, re.IGNORECASE)
            if epa_match:
                extracted_data['epa_id'] = epa_match.group(1).strip()
                
            # Waste Description
            desc_match = re.search(r'CUSTOMER\s+WASTE\s+DESCRIPTION:\s*(.*?)(?:\n|$)', full_text, re.IGNORECASE)
            if desc_match:
                extracted_data['waste_description'] = desc_match.group(1).strip()
                
            # DOT Description
            dot_match = re.search(r'DOT/TDG\s+PROPER\s+SHIPPING\s+NAME:\s*\n\s*(.*?)(?:\n|$)', full_text, re.IGNORECASE)
            if dot_match:
                extracted_data['dot_description'] = dot_match.group(1).strip()
                
            # Color will be extracted using spatial logic later
            
            # State and Federal Waste Codes will be extracted from page 2 using spatial logic
            
            # LDR Required
            ldr_cat_match = re.search(r'LDR\s+CATEGORY:\s*(.*?)(?:\n|$)', full_text, re.IGNORECASE)
            if ldr_cat_match:
                cat_text = ldr_cat_match.group(1).strip().lower()
                if "not subject" in cat_text:
                    extracted_data['ldr_required'] = "No"
                else:
                    extracted_data['ldr_required'] = "Yes"
            else:
                if len(pdf.pages) >= 3:
                    p3 = pdf.pages[2]
                    checkmarks = [img for img in p3.images if 8 < img['width'] < 12 and 8 < img['height'] < 12]
                    checkmarks = sorted(checkmarks, key=lambda c: c['top'])
                    if checkmarks:
                        closest, d = find_nearest_word(p3, checkmarks[0])
                        if closest:
                            ans = closest['text'].strip().capitalize()
                            extracted_data['ldr_required'] = "Yes" if ans == "Yes" else "No"

            # Checkmarks on page 1 for Physical Appearance, pH, Flash Point, Free Liquids
            p1 = pdf.pages[0]
            p1_checkmarks = [img for img in p1.images if 8 < img['width'] < 12 and 8 < img['height'] < 12]
            
            physical_state = "Solid"
            free_liquids = "No"
            ph_val = ""
            fp_val = ""
            
            for cm in p1_checkmarks:
                # 1. Physical State Checkbox (y-coordinate/top is roughly between 250 and 350)
                if 200 < cm['top'] < 360:
                    line_words = [w['text'] for w in p1.extract_words() if abs(w['top'] - cm['top']) < 8 and w['x0'] >= cm['x0'] - 2]
                    line_text = " ".join(line_words).lower()
                    if "solid without free liquid" in line_text:
                        physical_state = "Solid"
                        free_liquids = "No"
                    elif "liquid" in line_text or "sludge" in line_text:
                        physical_state = "Liquid"
                        if "free liquid" in line_text:
                            free_liquids = "Yes"
                    elif "powder" in line_text or "monolithic" in line_text:
                        physical_state = "Solid"
                    
                    # Shipping Container Type Check
                    if "container" in line_text or "drum" in line_text:
                        extracted_data['shipping_container_type'] = "Containerized"
                    elif "bulk solid" in line_text:
                        extracted_data['shipping_container_type'] = "Bulk Solid"
                    elif "bulk liquid" in line_text:
                        extracted_data['shipping_container_type'] = "Bulk Liquid"
                
                # 2. pH Checkbox (top is roughly between 380 and 550, x is strictly in the pH column)
                elif 380 < cm['top'] < 550 and 80 < cm['x0'] < 140:
                    line_words_raw = [w for w in p1.extract_words() if abs(w['top'] - cm['top']) < 5 and w['x0'] >= cm['x0'] - 2 and w['x0'] < cm['x0'] + 60]
                    line_words_raw = sorted(line_words_raw, key=lambda x: x['x0'])
                    line_text = " ".join([w['text'] for w in line_words_raw])
                    m = re.search(r'(<=?\s*\d+(?:\.\d+)?|\d+(?:\.\d+)?\s*-\s*\d+(?:\.\d+)?|\d+\s*\(Neutral\)|>=?\s*\d+(?:\.\d+)?)', line_text)
                    if m:
                        ph_val = m.group(1)

                # 3. Flash Point Checkbox (top is roughly between 380 and 550, x is strictly in the Flash Point column)
                elif 380 < cm['top'] < 550 and cm['x0'] < 50:
                    line_words_raw = [w for w in p1.extract_words() if abs(w['top'] - cm['top']) < 5 and w['x0'] >= cm['x0'] - 2 and w['x0'] < cm['x0'] + 60]
                    line_words_raw = sorted(line_words_raw, key=lambda x: x['x0'])
                    line_text = " ".join([w['text'] for w in line_words_raw])
                    m = re.search(r'(<=?\s*\d+|73\s*-\s*100|\d+\s*-\s*\d+|\d+\s*-\s*\d+|>\s*\d+)', line_text)
                    if m:
                        fp_val = m.group(1)

            extracted_data['physical_appearance'] = physical_state
            extracted_data['free_liquids'] = free_liquids
            extracted_data['ph_range'] = ph_val if ph_val else "7 (Neutral)"
            
            if physical_state == "Solid" and not fp_val:
                extracted_data['flash_point'] = "Not Required"
            elif fp_val:
                extracted_data['flash_point'] = fp_val.replace(' ', '') if '>' in fp_val or '<' in fp_val else fp_val
            else:
                extracted_data['flash_point'] = ">140"

            # Cyanide and Sulfide (page 2 checkmarks)
            cyanide = "No"
            sulfide = "No"
            if len(pdf.pages) >= 2:
                p2 = pdf.pages[1]
                p2_checkmarks = [img for img in p2.images if 8 < img['width'] < 12 and 8 < img['height'] < 12]
                for cm in p2_checkmarks:
                    line_words = sorted([w for w in p2.extract_words() if abs(w['top'] - cm['top']) < 8], key=lambda x: x['x0'])
                    line_text = " ".join([w['text'] for w in line_words]).lower()
                    if "cyanide" in line_text:
                        if cm['x0'] < 500:
                            cyanide = "Yes"
                    elif "sulfide" in line_text:
                        if cm['x0'] < 500:
                            sulfide = "Yes"
                            
                # Extract EPA and State Waste Codes from Page 2
                p2_words = p2.extract_words()
                
                # USEPA HAZARDOUS WASTE?
                usepa_word = next((w for w in p2_words if "USEPA" in w['text'].upper()), None)
                if usepa_word:
                    # Find the word "WASTE?" or similar on the same line to know where the label ends
                    waste_word = next((w for w in p2_words if "WASTE" in w['text'].upper() and w['x0'] > usepa_word['x0'] and abs(w['top'] - usepa_word['top']) < 10), usepa_word)
                    # Extract text strictly to the right of the label
                    line_words = sorted([w for w in p2_words if abs(w['top'] - usepa_word['top']) < 10 and w['x0'] > waste_word['x1'] + 5], key=lambda x: x['x0'])
                    fed_codes = [w['text'].strip() for w in line_words if w['text'].strip().upper() not in ['YES', 'NO', '?', 'APPLY']]
                    if fed_codes:
                        extracted_data['federal_waste_code'] = " ".join(fed_codes)
                        
                # STATE WASTE CODES
                # Explicitly ignore "Texas Waste Code" based on user request, only look for generic "State Waste Code"
                state_code_word = next((w for w in p2_words if "WASTE" in w['text'].upper() and "CODE" in "".join([ww['text'].upper() for ww in p2_words if abs(ww['top'] - w['top']) < 10]) and "STATE" in "".join([ww['text'].upper() for ww in p2_words if abs(ww['top'] - w['top']) < 10]) and "TEXAS" not in "".join([ww['text'].upper() for ww in p2_words if abs(ww['top'] - w['top']) < 10])), None)
                
                if state_code_word:
                    # Find the word "CODE" or "CODES" on the same line to know where the label ends
                    code_word = next((w for w in p2_words if "CODE" in w['text'].upper() and w['x0'] >= state_code_word['x0'] and abs(w['top'] - state_code_word['top']) < 10), state_code_word)
                    
                    # Find words either to the right of the label on the same line, or on the line directly below (within 15px)
                    line_words = []
                    for w in p2_words:
                        v_diff = w['top'] - state_code_word['top']
                        if abs(v_diff) < 3: # Same line
                            if w['x0'] > code_word['x1'] + 5:
                                line_words.append(w)
                        elif 3 <= v_diff < 15: # Line below
                            if w['x0'] >= state_code_word['x0'] - 60:
                                line_words.append(w)
                                
                    line_words = sorted(line_words, key=lambda x: x['x0'])
                    st_codes = [w['text'].strip() for w in line_words if w['text'].strip().upper() not in ['YES', 'NO', 'CODE', 'CODES', 'TEXAS', 'STATE', 'APPLY', '?']]
                    if st_codes:
                        extracted_data['state_waste_code'] = " ".join(st_codes)

            extracted_data['cyanide'] = cyanide
            extracted_data['sulfide'] = sulfide
            
            # Extract Color from Page 1 using spatial logic
            color_word = next((w for w in p1.extract_words() if w['text'].strip().upper() == 'COLOR'), None)
            if color_word:
                cx0, cx1, cbottom = color_word['x0'] - 20, color_word['x1'] + 60, color_word['bottom']
                color_vals = [w for w in p1.extract_words() if cbottom < w['top'] < cbottom + 50 and cx0 <= w['x0'] <= cx1]
                if color_vals:
                    color_vals.sort(key=lambda x: (x['top'], x['x0']))
                    extracted_data['color'] = " ".join([w['text'] for w in color_vals]).strip()

            
        return jsonify(extracted_data)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@approvals_bp.route('/add_master_profile', methods=['POST'])
def add_master_profile():
    profile_number = (request.form.get('profile_number') or '').strip().upper()
    if not profile_number:
        return "Profile Number is required", 400

    generator = (request.form.get('generator') or request.form.get('generator_name') or '').strip().upper()
    epa_id = request.form.get('epa_id', '').strip().upper()
    waste_description = request.form.get('waste_description', '').strip().upper()
    win_code = request.form.get('win_code', '').strip().upper()
    special_handling = request.form.get('special_handling', '').strip().upper()
    ph_range = request.form.get('ph_range', '').strip().upper()
    physical_appearance = request.form.get('physical_appearance', '').strip()
    flash_point = request.form.get('flash_point', '').strip().upper()
    expiration_date = request.form.get('expiration_date', '').strip()
    
    ldr_required = request.form.get('ldr_required', 'No').strip()
    ldr_option_str = request.form.get('ldr_option', '').strip()
    ldr_option = int(ldr_option_str) if ldr_option_str.isdigit() else None
    state_waste_code = request.form.get('state_waste_code', '').strip().upper()
    federal_waste_code = request.form.get('federal_waste_code', '').strip().upper()
    dot_description = request.form.get('dot_description', '').strip().upper()
    cyanide = request.form.get('cyanide', 'No').strip()
    sulfide = request.form.get('sulfide', 'No').strip()
    free_liquids = request.form.get('free_liquids', 'No').strip()
    lab_number = request.form.get('lab_number', '').strip().upper()
    color = request.form.get('color', '').strip().upper()
    shipping_container_type = request.form.get('shipping_container_type', 'Containerized').strip()

    raw_voc_input = str(request.form.get('voc_percentage', '')).strip().upper()
    if raw_voc_input in ['TBD', '?', 'NONE', '']:
        voc_pct = 'TBD'
    else:
        try:
            voc_pct = float(raw_voc_input)
        except (ValueError, TypeError):
            voc_pct = 'TBD'

    status = request.form.get('status', 'S').strip().upper()
    treatment_recipe = request.form.get('treatment_recipe', '').strip()

    import time
    from shared_state import MASTER_EXCEL_PATH
    excel_mtime = None
    try:
        if os.path.exists(MASTER_EXCEL_PATH):
            excel_mtime = os.path.getmtime(MASTER_EXCEL_PATH)
    except Exception:
        pass
    if not excel_mtime:
        excel_mtime = time.time()

    with closing(get_db_connection()) as conn:
        existing = conn.execute('SELECT profile_number FROM profiles WHERE TRIM(UPPER(profile_number)) = ?', (profile_number,)).fetchone()
        if existing:
            conn.execute('''
                UPDATE profiles 
                SET generator = ?, waste_description = ?, win_code = ?, voc_percentage = ?, 
                    special_handling = ?, ph_range = ?, physical_appearance = ?, flash_point = ?, 
                    expiration_date = ?, epa_id = ?, ldr_required = ?, ldr_option = ?, state_waste_code = ?, 
                    federal_waste_code = ?, dot_description = ?, cyanide = ?, sulfide = ?, 
                    free_liquids = ?, status = ?, lab_number = ?, color = ?, treatment_recipe = ?,
                    shipping_container_type = ?, last_synced_mtime = ?
                WHERE TRIM(UPPER(profile_number)) = ?
            ''', (generator, waste_description, win_code, voc_pct, 
                  special_handling, ph_range, physical_appearance, flash_point, 
                  expiration_date, epa_id, ldr_required, ldr_option, state_waste_code, 
                  federal_waste_code, dot_description, cyanide, sulfide, 
                  free_liquids, status, lab_number, color, treatment_recipe,
                  shipping_container_type, excel_mtime, profile_number))
        else:
            conn.execute('''
                INSERT INTO profiles (profile_number, generator, waste_description, win_code, voc_percentage, 
                                      special_handling, ph_range, physical_appearance, flash_point, expiration_date, 
                                      epa_id, status, ldr_required, ldr_option, state_waste_code, federal_waste_code, 
                                      dot_description, cyanide, sulfide, free_liquids, lab_number, color, treatment_recipe,
                                      shipping_container_type, last_synced_mtime)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (profile_number, generator, waste_description, win_code, voc_pct, 
                  special_handling, ph_range, physical_appearance, flash_point, expiration_date, 
                  epa_id, status, ldr_required, ldr_option, state_waste_code, federal_waste_code, 
                  dot_description, cyanide, sulfide, free_liquids, lab_number, color, treatment_recipe,
                  shipping_container_type, excel_mtime))
        conn.commit()
        
    return redirect(url_for('approvals_bp.approvals_portal', selected_profile=profile_number))

@approvals_bp.route('/api/auto_sync_profiles', methods=['POST'])
def auto_sync_profiles():
    date_str = request.form.get('schedule_date')
    updates_made = False
    
    if date_str:
        with closing(get_db_connection()) as conn:
            schedules = conn.execute('SELECT id, profile_number, voc_level, generator FROM daily_schedule WHERE schedule_date = ?', (date_str,)).fetchall()
            
            for s in schedules:
                prof_num = str(s['profile_number']).strip().upper()
                from database import ensure_profile_exists
                try:
                    ensure_profile_exists(conn, prof_num)
                except Exception as sync_err:
                    pass
                prof = conn.execute('''
                    SELECT voc_percentage, generator, win_code 
                    FROM profiles 
                    WHERE TRIM(UPPER(profile_number)) = ?
                ''', (prof_num,)).fetchone()
                
                if prof:
                    voc_raw = prof['voc_percentage']
                    if voc_raw is not None and str(voc_raw).strip() not in ['', 'None', 'NAN', 'NULL']:
                        try:
                            val = float(voc_raw)
                            new_voc = str(int(round(val)))
                        except (ValueError, TypeError):
                            new_voc = 'TBD'
                    else:
                        new_voc = 'TBD'
                        
                    # CRITICAL: Only update the database IF the values are actually different
                    if new_voc != str(s['voc_level']) or str(prof['generator']) != str(s['generator']):
                        conn.execute('''
                            UPDATE daily_schedule 
                            SET voc_level = ?, generator = ?, routing_code = ?
                            WHERE id = ?
                        ''', (new_voc, prof['generator'], prof['win_code'], s['id']))
                        updates_made = True
                        
            if updates_made:
                conn.commit()
                
        # Only ping the connected users if we actually changed something
        if updates_made:
            SCHEDULE_UPDATES[date_str] = time.time()
            
    return jsonify({'updated': updates_made})

from shared_state import UPLOADS_DIR
UPLOAD_FOLDER = UPLOADS_DIR

@approvals_bp.route('/api/profile/search')
def api_profile_search():
    query = request.args.get('q', '').strip()
    if not query:
        return jsonify([])
        
    with closing(get_db_connection()) as conn:
        like_query = f"%{query}%"
        rows = conn.execute('''
            SELECT p.*, 
                   w.verification_procedures, 
                   w.sample_procedures, 
                   w.unloading_instructions, 
                   w.notes_revisions, 
                   w.treatment_information
            FROM profiles p
            LEFT JOIN profile_wvi w ON TRIM(UPPER(p.profile_number)) = TRIM(UPPER(w.profile))
            WHERE (p.profile_number LIKE ? OR p.generator LIKE ? OR p.lab_number LIKE ? OR p.win_code LIKE ? OR p.epa_id LIKE ?) AND p.status != 'NOT FOUND'
            ORDER BY p.profile_number ASC
            LIMIT 50
        ''', (like_query, like_query, like_query, like_query, like_query)).fetchall()
        
    return jsonify([dict(r) for r in rows])

@approvals_bp.route('/api/profile/delete', methods=['POST'])
def api_profile_delete():
    data = request.json
    profile_number = data.get('profile_number')
    if not profile_number:
        return jsonify({'success': False, 'error': 'No profile number provided.'}), 400
        
    with closing(get_db_connection()) as conn:
        conn.execute('DELETE FROM profiles WHERE profile_number = ?', (profile_number,))
        conn.commit()
    
    return jsonify({'success': True})

@approvals_bp.route('/api/profile/<path:profile_number>/history')
def api_profile_history(profile_number):
    clean_profile = str(profile_number).strip().upper()
    with closing(get_db_connection()) as conn:
        drum_loads = conn.execute('''
            SELECT 
                manifest, 
                job_id, 
                import_date, 
                SUM(COALESCE(weight, 0)) AS weight, 
                COALESCE(process_type, 'Containerized') AS process_type, 
                MAX(ph) AS ph,
                MAX(voc_ppm) AS voc_ppm,
                'Drum' AS load_source
            FROM drum_inventory
            WHERE TRIM(UPPER(inb_prof)) = ?
            GROUP BY COALESCE(job_id, manifest), import_date
        ''', (clean_profile,)).fetchall()

        truck_loads = conn.execute('''
            SELECT 
                COALESCE(manifest_number, '---') AS manifest,
                COALESCE(truck_id, load_number, '---') AS job_id,
                COALESCE(date_received, time_in, '---') AS import_date,
                COALESCE(net_weight, CASE WHEN (gross_weight > 0 AND exit_weight > 0) THEN (gross_weight - exit_weight) ELSE gross_weight END, 0) AS weight,
                COALESCE(container_type, shipping_mode, 'Yellow Entry') AS process_type,
                measured_ph AS ph,
                measured_voc AS voc_ppm,
                'Yellow / Bulk' AS load_source
            FROM truck_logs
            WHERE TRIM(UPPER(profile_number)) = ?
              AND UPPER(COALESCE(test_status, '')) != 'VOID'
        ''', (clean_profile,)).fetchall()

        all_loads = [dict(l) for l in drum_loads] + [dict(l) for l in truck_loads]
        all_loads.sort(key=lambda x: str(x.get('import_date') or ''), reverse=True)

    return jsonify(all_loads)

@approvals_bp.route('/api/profile/<path:profile_number>/drum_history')
def api_profile_drum_history(profile_number):
    profile_clean = str(profile_number).strip().upper()
    with closing(get_db_connection()) as conn:
        drums = conn.execute('''
            SELECT q.drum_id, q.job_id, q.manifest, q.tests_required, q.status, MAX(i.import_date) AS import_date
            FROM drum_lab_queue q
            LEFT JOIN drum_inventory i ON TRIM(UPPER(q.drum_id)) = TRIM(UPPER(i.track_no))
            WHERE TRIM(UPPER(q.profile)) = ?
            GROUP BY q.id
            ORDER BY import_date DESC, q.id DESC
        ''', (profile_clean,)).fetchall()
    return jsonify([dict(d) for d in drums])

@approvals_bp.route('/api/profile/<path:profile_number>/sync', methods=['POST'])
def api_profile_sync(profile_number):
    profile_clean = str(profile_number).strip().upper()
    with closing(get_db_connection()) as conn:
        # Clear last_synced_mtime to force a fresh sync
        conn.execute('UPDATE profiles SET last_synced_mtime = NULL WHERE TRIM(UPPER(profile_number)) = ?', (profile_clean,))
        conn.commit()
        
        # Trigger ensure_profile_exists
        from database import ensure_profile_exists
        row = ensure_profile_exists(conn, profile_clean)
        if not row:
            return jsonify({'error': 'Profile not found in Excel.'}), 404
            
    return jsonify(dict(row))

@approvals_bp.route('/api/profile/<path:profile_number>/upload', methods=['POST'])
def api_profile_upload(profile_number):
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    files = request.files.getlist('file')
    if not files or all(file.filename == '' for file in files):
        return jsonify({'error': 'No selected files'}), 400
        
    uploaded_files = []
    with closing(get_db_connection()) as conn:
        for file in files:
            if file and file.filename != '':
                import time
                filename = secure_filename(file.filename)
                
                # Create a subfolder for the profile
                profile_dir_name = secure_filename(profile_number)
                profile_dir_path = os.path.join(UPLOAD_FOLDER, profile_dir_name)
                if not os.path.exists(profile_dir_path):
                    os.makedirs(profile_dir_path)
                
                # Append a timestamp to the original filename to prevent overwrites
                save_name = f"{int(time.time())}_{filename}"
                relative_path = f"{profile_dir_name}/{save_name}"
                file_path = os.path.join(profile_dir_path, save_name)
                
                file.save(file_path)
                
                # Real-time backup to network I: drive if mounted (runs with retries to prevent WinError 32 file lock crashes)
                try:
                    import shutil
                    import time
                    i_dirs = [
                        r"I:\Buttonwillow\LAB\Operations App\uploads",
                        r"I:\Buttonwillow\LAB\Operations App\uploads_backup"
                    ]
                    if os.path.exists(r"I:\\"):
                        for i_dir in i_dirs:
                            net_dest = os.path.join(i_dir, relative_path)
                            os.makedirs(os.path.dirname(net_dest), exist_ok=True)
                            for attempt in range(3):
                                try:
                                    shutil.copy2(file_path, net_dest)
                                    break
                                except Exception as copy_attempt_err:
                                    time.sleep(0.3)
                except Exception as net_e:
                    print(f"Real-time network upload backup warning: {net_e}")

                conn.execute('''
                    INSERT INTO profile_attachments (profile_number, filename, file_path)
                    VALUES (?, ?, ?)
                ''', (profile_number, filename, relative_path))
                uploaded_files.append(filename)
        conn.commit()
        
    return jsonify({'success': True, 'filenames': uploaded_files})

@approvals_bp.route('/api/profile/attachment/<int:attachment_id>/delete', methods=['POST'])
def api_delete_attachment(attachment_id):
    with closing(get_db_connection()) as conn:
        attachment = conn.execute('SELECT file_path FROM profile_attachments WHERE id = ?', (attachment_id,)).fetchone()
        if not attachment:
            return jsonify({'error': 'Attachment not found'}), 404
        
        file_path = attachment['file_path']
        full_path = os.path.join(UPLOAD_FOLDER, file_path)
        
        # Delete from database
        conn.execute('DELETE FROM profile_attachments WHERE id = ?', (attachment_id,))
        conn.commit()
        
        # Delete physical local file
        if os.path.exists(full_path):
            try:
                os.remove(full_path)
            except Exception as e:
                print(f"Error deleting file {full_path}: {e}")
                
        # Also remove from network backup if mounted
        try:
            i_uploads_dir = os.environ.get("I_DRIVE_UPLOADS_DIR", r"I:\Buttonwillow\LAB\Operations App\uploads_backup")
            net_full = os.path.join(i_uploads_dir, file_path)
            if os.path.exists(net_full):
                os.remove(net_full)
        except Exception:
            pass
                
    return jsonify({'success': True})

@approvals_bp.route('/api/profile/<path:profile_number>/attachments')
def api_profile_attachments(profile_number):
    with closing(get_db_connection()) as conn:
        attachments = conn.execute('''
            SELECT id, filename, file_path, upload_date 
            FROM profile_attachments
            WHERE profile_number = ?
            ORDER BY upload_date DESC
        ''', (profile_number,)).fetchall()
    return jsonify([dict(a) for a in attachments])

# ==========================================
# VOC ANALYZER & SULFIDE STATISTICAL TESTING LOGS
# ==========================================

T_TABLE_90_CONFIDENCE = {
    1: 3.078,
    2: 1.886,
    3: 1.638,
    4: 1.533
}

@approvals_bp.route('/api/profile/<path:profile_number>/voc_analyzer', methods=['GET', 'POST'])
def api_profile_voc_analyzer(profile_number):
    profile_clean = str(profile_number).strip().upper()
    with closing(get_db_connection()) as conn:
        if request.method == 'POST':
            data = request.json or request.form
            cp1_number = str(data.get('cp1_number', '')).strip().upper()
            try:
                voc_val = float(data.get('voc_analyzer_value'))
            except (ValueError, TypeError):
                return jsonify({'error': 'Invalid VOC Analyzer reading'}), 400
                
            orig_voc = data.get('original_voc_value')
            try:
                orig_voc = float(orig_voc) if orig_voc is not None else None
            except:
                orig_voc = None
                
            tested_by = str(data.get('tested_by', 'Lab Chemist')).strip()
            notes = str(data.get('notes', '')).strip()
            
            conn.execute('''
                INSERT INTO voc_analyzer_logs (
                    profile_number, cp1_number, voc_analyzer_value, original_voc_value, tested_by, notes
                ) VALUES (?, ?, ?, ?, ?, ?)
            ''', (profile_clean, cp1_number, voc_val, orig_voc, tested_by, notes))
            conn.commit()
            return jsonify({'success': True})
            
        else:
            logs = conn.execute('''
                SELECT id, cp1_number, voc_analyzer_value, original_voc_value, tested_by, test_date, notes
                FROM voc_analyzer_logs
                WHERE TRIM(UPPER(profile_number)) = ?
                ORDER BY test_date DESC
            ''', (profile_clean,)).fetchall()
            return jsonify([dict(l) for l in logs])


def calculate_sulfide_summary_stats(logs):
    import math
    tot_samples = []
    react_samples = []
    
    for l in logs:
        # Check single sample columns
        t_val_item = l.get('total_sulfide')
        r_val_item = l.get('reactive_sulfide')
        
        if t_val_item is not None and str(t_val_item).strip() != '':
            try: tot_samples.append(float(t_val_item))
            except: pass
            if r_val_item is not None and str(r_val_item).strip() != '':
                try: react_samples.append(float(r_val_item))
                except: pass
        elif l.get('total_sulfide_samples'):
            # Legacy bulk array handling
            try:
                raw_tot = json.loads(l['total_sulfide_samples']) if isinstance(l['total_sulfide_samples'], str) else l['total_sulfide_samples']
                tot_samples.extend([float(x) for x in raw_tot if x is not None])
            except: pass
            try:
                raw_react = json.loads(l['reactive_sulfide_samples']) if isinstance(l['reactive_sulfide_samples'], str) else l['reactive_sulfide_samples']
                react_samples.extend([float(x) for x in raw_react if x is not None])
            except: pass

    n = len(tot_samples)
    if n == 0:
        return {
            'n': 0, 'df': 0, 't_value': 1.638,
            'total_sulfide': {'mean': 0.0, 'stddev': 0.0, 'ci90': 0.0},
            'reactive_sulfide': {'mean': 0.0, 'stddev': 0.0, 'ci90': 0.0},
            'reactive_pass': False,
            'approval_status': 'AWAITING_DATA',
            'tests_remaining': 5
        }
        
    df = max(1, n - 1)
    t_val = T_TABLE_90_CONFIDENCE.get(df, 1.476)
    
    def calc(samples):
        cnt = len(samples)
        if cnt == 0: return 0.0, 0.0, 0.0
        mean = sum(samples) / float(cnt)
        stddev = math.sqrt(sum((x - mean)**2 for x in samples) / float(cnt - 1)) if cnt > 1 else 0.0
        ci90 = mean + (stddev * t_val)
        return round(mean, 2), round(stddev, 2), round(ci90, 2)
        
    tot_mean, tot_stddev, tot_ci90 = calc(tot_samples)
    react_mean, react_stddev, react_ci90 = calc(react_samples)
    
    is_approved = (n >= 5) and (len(react_samples) == 0 or react_ci90 <= 500.0)
    reactive_pass = 1 if is_approved else 0
    approval_status = "APPROVED" if is_approved else ("IN_PROGRESS" if n < 5 else "FAILED")
    tests_remaining = max(0, 5 - n)
    
    return {
        'n': n,
        'df': df,
        't_value': t_val,
        'total_sulfide': {'mean': tot_mean, 'stddev': tot_stddev, 'ci90': tot_ci90},
        'reactive_sulfide': {'mean': react_mean, 'stddev': react_stddev, 'ci90': react_ci90},
        'reactive_pass': bool(reactive_pass),
        'approval_status': approval_status,
        'tests_remaining': tests_remaining
    }

@approvals_bp.route('/api/profile/<path:profile_number>/sulfide_log', methods=['GET', 'POST'])
def api_profile_sulfide_log(profile_number):
    profile_clean = str(profile_number).strip().upper()
    import json, math
    with closing(get_db_connection()) as conn:
        if request.method == 'POST':
            data = request.json or request.form
            cp1_number = str(data.get('cp1_number', '')).strip().upper()
            lab_number = str(data.get('lab_number', '')).strip().upper()
            weight_ticket = str(data.get('weight_ticket', '')).strip().upper()
            tested_by = str(data.get('tested_by', 'Lab Chemist')).strip()
            notes = str(data.get('notes', '')).strip()
            test_date_input = data.get('test_date')
            
            tot_val = data.get('total_sulfide')
            react_val = data.get('reactive_sulfide')
            
            # Single sample validation
            if tot_val is None or str(tot_val).strip() == '':
                # Check legacy array input fallback
                tot_raw = data.get('total_sulfide_samples', [])
                if not tot_raw:
                    return jsonify({'error': 'Total Sulfide (mg/kg) is required.'}), 400
                tot_val = tot_raw[0] if isinstance(tot_raw, list) else float(tot_raw)
                
            tot_num = float(tot_val)
            react_num = float(react_val) if (react_val is not None and str(react_val).strip() != '') else None
            
            if test_date_input:
                conn.execute('''
                    INSERT INTO sulfide_testing_logs (
                        profile_number, cp1_number, lab_number, weight_ticket,
                        total_sulfide, reactive_sulfide, tested_by, test_date, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    profile_clean, cp1_number, lab_number, weight_ticket,
                    tot_num, react_num, tested_by, test_date_input, notes
                ))
            else:
                conn.execute('''
                    INSERT INTO sulfide_testing_logs (
                        profile_number, cp1_number, lab_number, weight_ticket,
                        total_sulfide, reactive_sulfide, tested_by, notes
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (
                    profile_clean, cp1_number, lab_number, weight_ticket,
                    tot_num, react_num, tested_by, notes
                ))
            conn.commit()
            
            # Fetch all logs to compute summary
            rows = conn.execute('''
                SELECT * FROM sulfide_testing_logs
                WHERE TRIM(UPPER(profile_number)) = ?
                ORDER BY test_date ASC, id ASC
            ''', (profile_clean,)).fetchall()
            logs = [dict(r) for r in rows]
            stats = calculate_sulfide_summary_stats(logs)
            
            return jsonify({
                'success': True,
                'stats': stats,
                'logs': logs
            })
        else:
            rows = conn.execute('''
                SELECT * FROM sulfide_testing_logs
                WHERE TRIM(UPPER(profile_number)) = ?
                ORDER BY test_date ASC, id ASC
            ''', (profile_clean,)).fetchall()
            logs = [dict(r) for r in rows]
            stats = calculate_sulfide_summary_stats(logs)
            return jsonify({'logs': logs, 'stats': stats})


@approvals_bp.route('/api/profile/sulfide_log/<int:log_id>', methods=['PUT', 'DELETE'])
def api_profile_sulfide_log_item(log_id):
    with closing(get_db_connection()) as conn:
        log = conn.execute('SELECT * FROM sulfide_testing_logs WHERE id = ?', (log_id,)).fetchone()
        if not log:
            return jsonify({'error': 'Log entry not found'}), 404
            
        profile_clean = log['profile_number']
        
        if request.method == 'DELETE':
            conn.execute('DELETE FROM sulfide_testing_logs WHERE id = ?', (log_id,))
            conn.commit()
        elif request.method == 'PUT':
            data = request.json or request.form
            cp1_number = str(data.get('cp1_number', '')).strip().upper()
            lab_number = str(data.get('lab_number', '')).strip().upper()
            weight_ticket = str(data.get('weight_ticket', '')).strip().upper()
            tested_by = str(data.get('tested_by', log['tested_by'] or 'Lab Chemist')).strip()
            notes = str(data.get('notes', '')).strip()
            test_date_input = data.get('test_date')
            
            tot_val = data.get('total_sulfide')
            react_val = data.get('reactive_sulfide')
            
            tot_num = float(tot_val) if (tot_val is not None and str(tot_val).strip() != '') else log['total_sulfide']
            react_num = float(react_val) if (react_val is not None and str(react_val).strip() != '') else None
            
            conn.execute('''
                UPDATE sulfide_testing_logs
                SET cp1_number = ?, lab_number = ?, weight_ticket = ?,
                    total_sulfide = ?, reactive_sulfide = ?, tested_by = ?,
                    test_date = COALESCE(?, test_date), notes = ?
                WHERE id = ?
            ''', (
                cp1_number, lab_number, weight_ticket,
                tot_num, react_num, tested_by, test_date_input, notes, log_id
            ))
            conn.commit()
            
        # Re-fetch updated logs & stats for profile
        rows = conn.execute('''
            SELECT * FROM sulfide_testing_logs
            WHERE TRIM(UPPER(profile_number)) = ?
            ORDER BY test_date ASC, id ASC
        ''', (profile_clean,)).fetchall()
        logs = [dict(r) for r in rows]
        stats = calculate_sulfide_summary_stats(logs)
        return jsonify({'success': True, 'stats': stats, 'logs': logs})

@approvals_bp.route('/uploads/profiles/<path:filename>')
def serve_profile_upload(filename):
    clean_fn = filename.replace('\\', '/').strip('/')
    basename = os.path.basename(clean_fn)
    
    # Candidate search paths in priority order (I: Drive FIRST!)
    i_drive_primary = os.environ.get("I_DRIVE_UPLOADS_DIR", r"I:\Buttonwillow\LAB\Operations App\uploads")
    i_drive_backup = r"I:\Buttonwillow\LAB\Operations App\uploads_backup"
    i_drive_base = r"I:\Buttonwillow\LAB\Operations App"
    wvi_base = r"I:\Buttonwillow\WAP\WVI"
    
    from shared_state import UPLOADS_DIR, APP_DIR
    local_dir = UPLOADS_DIR
    fallback_local = os.path.join(APP_DIR, 'uploads', 'profiles')

    candidate_paths = [
        os.path.join(i_drive_primary, clean_fn),
        os.path.join(i_drive_backup, clean_fn),
        os.path.join(local_dir, clean_fn),
        os.path.join(fallback_local, clean_fn),
        os.path.join(i_drive_base, clean_fn),
        
        # Basename fallbacks in upload folders
        os.path.join(i_drive_primary, basename),
        os.path.join(i_drive_backup, basename),
        os.path.join(local_dir, basename),
        os.path.join(fallback_local, basename),
        
        # WVI Folder Fallbacks for profile documents
        os.path.join(wvi_base, basename),
        os.path.join(wvi_base, "2026 WVI", basename),
        os.path.join(wvi_base, "2025 WVI", basename),
        os.path.join(wvi_base, "2024 WVI", basename),
        os.path.join(wvi_base, "2023 WVI", basename),
    ]

    for path in candidate_paths:
        if os.path.exists(path) and os.path.isfile(path):
            # Auto-restore/cache to local folder if missing locally
            try:
                local_dest = os.path.join(local_dir, clean_fn)
                if not os.path.exists(local_dest):
                    os.makedirs(os.path.dirname(local_dest), exist_ok=True)
                    import shutil
                    shutil.copy2(path, local_dest)
            except Exception:
                pass
            return send_file(path)

    return jsonify({'error': f'Attachment file {filename} not found on I: drive or local server.'}), 404

@approvals_bp.route('/release_las_truck', methods=['POST'])
def release_las_truck():
    from shared_state import socketio
    log_id = request.form.get('log_id')
    measured_ph = request.form.get('measured_ph')
    measured_voc = request.form.get('measured_voc')
    measured_sulfides = request.form.get('measured_sulfides', 'Negative')
    measured_cyanide = request.form.get('measured_cyanide', 'Negative')
    measured_free_liquids = request.form.get('measured_free_liquids', 'No')
    measured_flashpoint = request.form.get('measured_flashpoint', '')
    notes = request.form.get('notes', '')

    # Safely convert to float
    try:
        ph_val = float(measured_ph) if measured_ph else None
    except ValueError:
        ph_val = None

    try:
        voc_val = float(measured_voc) if measured_voc else None
    except ValueError:
        voc_val = None

    with closing(get_db_connection()) as conn:
        # Get the truck details to find profile number and date_received
        truck = conn.execute('SELECT profile_number, date_received FROM truck_logs WHERE id = ?', (log_id,)).fetchone()
        if not truck:
            return "Truck log not found", 404
        
        profile_number = truck['profile_number']
        received_date = truck['date_received'] or date.today().isoformat()

        # Find the profile details
        from database import ensure_profile_exists
        ensure_profile_exists(conn, profile_number)
        profile = conn.execute('SELECT win_code FROM profiles WHERE TRIM(UPPER(profile_number)) = TRIM(UPPER(?))', (profile_number,)).fetchone()
        
        is_ccs = False
        if profile_number and 'CCS' in profile_number.upper():
            is_ccs = True
        elif profile and profile['win_code'] and 'CCS' in str(profile['win_code']).upper():
            is_ccs = True

        # CCS check
        if is_ccs:
            if voc_val is not None and voc_val > 500:
                voc_pass_fail = 'FAIL'
            else:
                voc_pass_fail = 'PASS'
        else:
            voc_pass_fail = 'N/A'

        release_note = f"Released by Waste Acceptance. {notes}".strip() if notes else "Released by Waste Acceptance."

        conn.execute('''
            UPDATE truck_logs
            SET measured_ph = ?,
                measured_voc = ?,
                measured_sulfides = ?,
                measured_cyanide = ?,
                measured_free_liquids = ?,
                measured_flashpoint = ?,
                voc_pass_fail = ?,
                lab_results = ?,
                test_status = 'RELEASED'
            WHERE id = ?
        ''', (ph_val, voc_val, measured_sulfides, measured_cyanide, measured_free_liquids, measured_flashpoint, voc_pass_fail, release_note, log_id))
        
        conn.commit()

        # Emit websocket updates to keep queues in sync
        socketio.emit('lab_update', {'date': received_date})
        socketio.emit('truck_update', {'date': received_date})

    return redirect(url_for('stu_bp.stu_hub', view='pipeline'))

# --- WASTE ACCEPTANCE ACTIVE REVIEWS LOG API ---

@approvals_bp.route('/api/waste_acceptance/log', methods=['GET'])
def get_waste_acceptance_log():
    archived = request.args.get('archived', '0')
    archived_val = 1 if archived == '1' else 0
    with closing(get_db_connection()) as conn:
        logs = conn.execute('''
            SELECT w.*, COALESCE(w.generator_requestor, p.generator) AS generator 
            FROM waste_acceptance_log w
            LEFT JOIN profiles p ON TRIM(UPPER(w.profile_number)) = TRIM(UPPER(p.profile_number))
            WHERE COALESCE(w.is_archived, 0) = ?
            ORDER BY w.last_updated DESC
        ''', (archived_val,)).fetchall()
    return jsonify([dict(row) for row in logs])

@approvals_bp.route('/api/waste_acceptance/log/add', methods=['POST'])
def add_waste_acceptance_log():
    data = request.get_json() or {}
    profile_number = data.get('profile_number', '').strip().upper()
    if not profile_number:
        return jsonify({'error': 'Profile number is required'}), 400
        
    with closing(get_db_connection()) as conn:
        # Check if the profile already exists in the log (active or archived)
        existing = conn.execute('''
            SELECT id, is_archived FROM waste_acceptance_log 
            WHERE TRIM(UPPER(profile_number)) = ?
        ''', (profile_number,)).fetchone()
        
        if existing:
            if existing['is_archived'] == 1:
                # Reactivate the archived profile and reset status/notes
                conn.execute('''
                    UPDATE waste_acceptance_log
                    SET is_archived = 0,
                        status = 'Needs Review',
                        assigned_to = '',
                        notes = '',
                        date_added = CURRENT_TIMESTAMP,
                        last_updated = CURRENT_TIMESTAMP
                    WHERE id = ?
                ''', (existing['id'],))
                conn.commit()
                return jsonify({'success': True})
            else:
                return jsonify({'error': f"Profile {profile_number} is already active in the review log."}), 400
        
        try:
            conn.execute('''
                INSERT INTO waste_acceptance_log (profile_number, status, assigned_to, notes, date_added)
                VALUES (?, 'Needs Review', '', '', CURRENT_TIMESTAMP)
            ''', (profile_number,))
            conn.commit()
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'error': str(e)}), 400

@approvals_bp.route('/api/waste_acceptance/log/update', methods=['POST'])
def update_waste_acceptance_log():
    data = request.get_json() or {}
    log_id = data.get('id')
    status = data.get('status')
    assigned_to = data.get('assigned_to')
    notes = data.get('notes')
    generator_requestor = data.get('generator_requestor')
    profile_number = data.get('profile_number')
    expiration_date = data.get('expiration_date')
    
    if profile_number is not None:
        profile_number = profile_number.strip().upper()
        if not profile_number:
            return jsonify({'success': False, 'error': 'Profile number cannot be empty'}), 400
            
    if not log_id:
        return jsonify({'success': False, 'error': 'Log ID is required'}), 400
        
    with closing(get_db_connection()) as conn:
        try:
            conn.execute('''
                UPDATE waste_acceptance_log 
                SET status = COALESCE(?, status),
                    assigned_to = COALESCE(?, assigned_to),
                    notes = COALESCE(?, notes),
                    generator_requestor = COALESCE(?, generator_requestor),
                    profile_number = COALESCE(?, profile_number),
                    expiration_date = COALESCE(?, expiration_date),
                    last_updated = CURRENT_TIMESTAMP
                WHERE id = ?
            ''', (status, assigned_to, notes, generator_requestor, profile_number, expiration_date, log_id))
            
            row_prof = conn.execute('SELECT profile_number FROM waste_acceptance_log WHERE id = ?', (log_id,)).fetchone()
            prof_key = str(row_prof['profile_number'] if row_prof else profile_number or '').strip().upper()
            if prof_key:
                excel_mtime = None
                try:
                    from shared_state import MASTER_EXCEL_PATH
                    if os.path.exists(MASTER_EXCEL_PATH):
                        excel_mtime = os.path.getmtime(MASTER_EXCEL_PATH)
                except Exception:
                    pass
                if not excel_mtime:
                    import time
                    excel_mtime = time.time()

                conn.execute('''
                    UPDATE profiles 
                    SET expiration_date = CASE WHEN ? IS NOT NULL AND TRIM(?) != '' THEN ? ELSE expiration_date END,
                        status = CASE WHEN ? IN ('Recertified', 'Complete', 'Released', 'ACTIVE') THEN 'ACTIVE' ELSE status END,
                        last_synced_mtime = ?
                    WHERE TRIM(UPPER(profile_number)) = ?
                ''', (expiration_date, expiration_date, expiration_date, status, excel_mtime, prof_key))
            
            conn.commit()
            
            # Emit schedule update so schedule UI refreshes LAS badges immediately
            try:
                from shared_state import socketio
                socketio.emit('schedule_update', {'date': 'GLOBAL'})
            except Exception:
                pass
                
            return jsonify({'success': True})
        except Exception as e:
            return jsonify({'success': False, 'error': 'Profile number must be unique in the log.'}), 400

@approvals_bp.route('/api/waste_acceptance/log/delete', methods=['POST'])
def delete_waste_acceptance_log():
    data = request.get_json() or {}
    log_id = data.get('id')
    if not log_id:
        return jsonify({'error': 'Log ID is required'}), 400
        
    with closing(get_db_connection()) as conn:
        conn.execute('DELETE FROM waste_acceptance_log WHERE id = ?', (log_id,))
        conn.commit()
    return jsonify({'success': True})

@approvals_bp.route('/api/waste_acceptance/log/archive', methods=['POST'])
def archive_waste_acceptance_log():
    data = request.get_json() or {}
    log_id = data.get('id')
    is_archived = data.get('is_archived', 1)
    if not log_id:
        return jsonify({'error': 'Log ID is required'}), 400
        
    archived_val = 1 if is_archived else 0
    with closing(get_db_connection()) as conn:
        conn.execute('''
            UPDATE waste_acceptance_log 
            SET is_archived = ?, last_updated = CURRENT_TIMESTAMP 
            WHERE id = ?
        ''', (archived_val, log_id))
        conn.commit()
    return jsonify({'success': True})

LDR_OPTIONS_MAP = {
    1: 'Not subject to LDR',
    2: 'This is subject to LDR.',
    3: 'Alternate Debris Standard',
    4: 'Meets LDR Standards',
    5: 'Alternate Soil Std-meets std. (with listed and characteristic hazardous waste)',
    6: 'Alternate Soil Std-meets std. (with characteristic hazardous waste only)',
    7: 'Alternate Soil Std-meets std. (with listed hazardous waste only)',
    8: 'Alternate Soil Std-does not meet std. (with listed and characteristic hazardous waste)',
    9: 'Alternate Soil Std-does not meet std. (with listed hazardous waste only)',
    10: 'Alternate Soil Std-does not meet std. (with characteristic hazardous waste only)',
    11: 'Labpack Alternate Standard',
    12: 'Subject to a variance or exemption',
    13: 'Partially meets LDR standards',
    14: 'Treatment Technology & Process certification',
    15: 'Contaminated Soil Treatment Tech. And Process Cert.',
    16: 'Concentration-based Certification for combustion residues',
    17: 'Waste treated to remove characteristic, but not UHCs',
    18: 'Waste treated to remove characteristic and UHCs',
    19: 'Treated Debris under 268.45',
    20: 'Generator chooses not to make the determination of whether the waste must be treated.'
}

@approvals_bp.route('/api/profile/<path:profile_number>/wvi')
def export_wvi_excel(profile_number):
    import io
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import send_file
    import re
    
    profile_clean = str(profile_number).strip().upper()
    
    # 1. Fetch data from profiles and profile_wvi tables
    with closing(get_db_connection()) as conn:
        p_row = conn.execute('SELECT * FROM profiles WHERE TRIM(UPPER(profile_number)) = ?', (profile_clean,)).fetchone()
        w_row = conn.execute('SELECT * FROM profile_wvi WHERE TRIM(UPPER(profile)) = ?', (profile_clean,)).fetchone()
        
    if not p_row and not w_row:
        return "Profile not found in database", 404
        
    win_code = str(p_row['win_code'] if (p_row and p_row['win_code']) else '').strip().upper()
    lab_num_clean = str(w_row['lab_num'] if (w_row and w_row['lab_num']) else (p_row['lab_number'] if (p_row and p_row['lab_number']) else '')).strip().upper()
    is_monolith = "MONOLITH" in lab_num_clean or "MONO" in lab_num_clean
        
    # Helper to parse pH range (like "7.0 to 12.4" or "7.0-12.0")
    def parse_ph_range(ph_range):
        if not ph_range:
            return None, None
            
        ph_str = str(ph_range).strip().upper()
        if "7 (NEUTRAL)" in ph_str or ph_str == "7":
            return 4.0, 10.0
            
        m = re.findall(r'(\d+\.?\d*)', str(ph_range))
        if len(m) >= 2:
            try:
                return float(m[0]), float(m[1])
            except:
                pass
        elif len(m) == 1:
            try:
                return float(m[0]), float(m[0])
            except:
                pass
        return None, None

    # Helper to convert/clean voc_percentage to voc_ppm
    def get_voc_ppm(percentage):
        if percentage is None:
            return None
        try:
            val = float(percentage)
            return val
        except:
            return None

    # Combine data from tables
    combined = {
        'profile': profile_clean,
        'generator_name': w_row['generator_name'] if (w_row and w_row['generator_name']) else (p_row['generator'] if p_row else ''),
        'waste_name': w_row['waste_name'] if (w_row and w_row['waste_name']) else (p_row['waste_description'] if p_row else ''),
        'physical_description': w_row['physical_description'] if (w_row and w_row['physical_description']) else (p_row['physical_appearance'] if p_row else ''),
        'ldr': w_row['ldr'] if (w_row and w_row['ldr']) else (p_row['ldr_required'] if p_row else ''),
        'state_waste_codes': w_row['state_waste_codes'] if (w_row and w_row['state_waste_codes']) else (p_row['state_waste_code'] if p_row else ''),
        'federal_waste_codes': w_row['federal_waste_codes'] if (w_row and w_row['federal_waste_codes']) else (p_row['federal_waste_code'] if p_row else ''),
        'dot_description': w_row['dot_description'] if (w_row and w_row['dot_description']) else (p_row['dot_description'] if p_row else ''),
        'handling_instruction': w_row['handling_instruction'] if (w_row and w_row['handling_instruction']) else '',
        'sample_procedures': w_row['sample_procedures'] if (w_row and w_row['sample_procedures']) else '',
        'verification_procedures': (
            w_row['verification_procedures'] if (w_row and w_row['verification_procedures'] and "FINGER" not in str(w_row['verification_procedures']).upper())
            else ("Visual" if (win_code == 'CNIA' or is_monolith) else "VERIFY FINGERPRINT RESULTS AND PHYSICAL DESCRIPTION WITH PROFILE")
        ),
        'ph_min': w_row['ph_min'] if (w_row and w_row['ph_min'] is not None) else None,
        'ph_max': w_row['ph_max'] if (w_row and w_row['ph_max'] is not None) else None,
        'sulfides': p_row['sulfide'] if (p_row and p_row['sulfide'] and str(p_row['sulfide']).strip() != '') else (w_row['sulfides'] if (w_row and w_row['sulfides']) else ''),
        'cyanide': w_row['cyanide'] if (w_row and w_row['cyanide']) else (p_row['cyanide'] if p_row else ''),
        'free_liquids': w_row['free_liquids'] if (w_row and w_row['free_liquids']) else (p_row['free_liquids'] if p_row else ''),
        'flashpoint': w_row['flashpoint'] if (w_row and w_row['flashpoint']) else (p_row['flash_point'] if p_row else ''),
        'unloading_instructions': (
            str(w_row['unloading_instructions']).strip() if (w_row and w_row['unloading_instructions'] is not None and str(w_row['unloading_instructions']).strip() != "")
            else (str(p_row['special_handling']).strip() if (p_row and p_row['special_handling'] is not None and str(p_row['special_handling']).strip() != "")
            else ("Follow Red Folder" if (win_code in ['CCS', 'CCSM'] or win_code.startswith('CCS')) else "NO SPECIAL HANDLING REQUIRED UNLOAD TO LANDFILL NORMALLY"))
        ),
        'reactivity_codes': w_row['reactivity_codes'] if (w_row and w_row['reactivity_codes']) else 'NONE',
        'approved_date': w_row['approved_date'] if (w_row and w_row['approved_date']) else '',
        'expiration_date': w_row['expiration_date'] if (w_row and w_row['expiration_date']) else (p_row['expiration_date'] if p_row else ''),
        'lab_num': w_row['lab_num'] if (w_row and w_row['lab_num']) else (p_row['lab_number'] if p_row else ''),
        'voc_ppm': w_row['voc_ppm'] if (w_row and w_row['voc_ppm'] is not None) else None,
        'treatment_information': w_row['treatment_information'] if (w_row and w_row['treatment_information']) else 'Not applicable',
        'notes_revisions': w_row['notes_revisions'] if (w_row and w_row['notes_revisions']) else '',
        'ldr_option': int(p_row['ldr_option']) if (p_row and 'ldr_option' in p_row.keys() and p_row['ldr_option'] is not None and str(p_row['ldr_option']).isdigit()) else None
    }

    # Convert all string values in combined dictionary to uppercase
    for k, v in combined.items():
        if isinstance(v, str):
            combined[k] = v.upper()

    # Fill in fallback parsed range/voc
    if combined['ph_min'] is None or combined['ph_max'] is None:
        ph_source = (p_row['ph_range'] if (p_row and p_row['ph_range']) else None)
        if not ph_source and w_row:
            from database import parse_ph_from_any_text
            ph_source = parse_ph_from_any_text(
                w_row['notes_revisions'],
                w_row['verification_procedures'],
                w_row['physical_description'],
                w_row['handling_instruction']
            )
        if ph_source:
            ph_min_parsed, ph_max_parsed = parse_ph_range(ph_source)
            if combined['ph_min'] is None:
                combined['ph_min'] = ph_min_parsed
            if combined['ph_max'] is None:
                combined['ph_max'] = ph_max_parsed

    voc_from_profile = get_voc_ppm(p_row['voc_percentage']) if (p_row and p_row['voc_percentage'] is not None) else None
    voc_from_log = None
    try:
        voc_analyzer_row = conn.execute('''
            SELECT voc_analyzer_value FROM voc_analyzer_logs
            WHERE TRIM(UPPER(profile_number)) = ? AND voc_analyzer_value IS NOT NULL AND voc_analyzer_value > 0
            ORDER BY test_date DESC LIMIT 1
        ''', (profile_clean,)).fetchone()
        if voc_analyzer_row and voc_analyzer_row['voc_analyzer_value'] is not None:
            voc_from_log = float(voc_analyzer_row['voc_analyzer_value'])
        else:
            truck_log_row = conn.execute('''
                SELECT voc_ppm FROM truck_logs
                WHERE TRIM(UPPER(profile_number)) = ? AND voc_ppm IS NOT NULL AND voc_ppm > 0
                ORDER BY id DESC LIMIT 1
            ''', (profile_clean,)).fetchone()
            if truck_log_row and truck_log_row['voc_ppm'] is not None:
                voc_from_log = float(truck_log_row['voc_ppm'])
    except Exception as ex:
        print(f"Error fetching logged VOC for profile {profile_clean}: {ex}")

    if voc_from_log is not None and voc_from_log > 0:
        combined['voc_ppm'] = voc_from_log
    elif voc_from_profile is not None and voc_from_profile > 0:
        combined['voc_ppm'] = voc_from_profile
    elif combined['voc_ppm'] is None or combined['voc_ppm'] == 0:
        if voc_from_profile is not None:
            combined['voc_ppm'] = voc_from_profile
        elif voc_from_log is not None:
            combined['voc_ppm'] = voc_from_log
        elif w_row and w_row['voc_ppm'] is not None:
            combined['voc_ppm'] = w_row['voc_ppm']

    # Format Sulfides/Cyanide/Free Liquids to NEG/POS style or YES/NO style
    def to_neg_pos(val):
        if not val:
            return 'NEG'
        v_str = str(val).strip().upper()
        if v_str in ['YES', 'TRUE', 'POS', 'POSITIVE', 'Y']:
            return 'POS'
        if v_str in ['NO', 'FALSE', 'NEG', 'NEGATIVE', 'N']:
            return 'NEG'
        return val

    def to_yes_no(val):
        if not val:
            return 'NO'
        v_str = str(val).strip().upper()
        if v_str in ['YES', 'TRUE', 'POS', 'POSITIVE', 'Y']:
            return 'YES'
        if v_str in ['NO', 'FALSE', 'NEG', 'NEGATIVE', 'N']:
            return 'NO'
        return val

    combined['sulfides'] = to_neg_pos(combined['sulfides'])
    combined['cyanide'] = to_neg_pos(combined['cyanide'])
    combined['free_liquids'] = to_yes_no(combined['free_liquids'])

    # Format date outputs
    def format_date_serial_or_string(val):
        if not val:
            return ''
        return str(val)

    combined['approved_date'] = format_date_serial_or_string(combined['approved_date'])
    combined['expiration_date'] = format_date_serial_or_string(combined['expiration_date'])
    
    if combined['expiration_date']:
        try:
            import datetime
            exp_str = combined['expiration_date']
            exp_date = None
            if '-' in exp_str:
                # might be YYYY-MM-DD
                parts = exp_str.split('-')
                if len(parts[0]) == 4:
                    exp_date = datetime.datetime.strptime(exp_str, '%Y-%m-%d')
                else:
                    exp_date = datetime.datetime.strptime(exp_str, '%m-%d-%Y')
            elif '/' in exp_str:
                parts = exp_str.split('/')
                if len(parts[2]) == 2:
                    exp_date = datetime.datetime.strptime(exp_str, '%m/%d/%y')
                else:
                    exp_date = datetime.datetime.strptime(exp_str, '%m/%d/%Y')
            
            if exp_date:
                try:
                    app_date = exp_date.replace(year=exp_date.year - 1)
                except ValueError:
                    app_date = exp_date.replace(year=exp_date.year - 1, day=28)
                combined['approved_date'] = app_date.strftime('%m/%d/%Y')
        except Exception:
            pass

    # Ensure both dates are explicitly formatted as MM/DD/YYYY if they are valid
    import datetime
    for date_key in ['approved_date', 'expiration_date']:
        d_val = combined.get(date_key, '')
        if d_val:
            try:
                if '-' in d_val:
                    # check if YYYY-MM-DD
                    parts = d_val.split('-')
                    if len(parts[0]) == 4:
                        parsed = datetime.datetime.strptime(d_val.split(' ')[0], '%Y-%m-%d')
                        combined[date_key] = parsed.strftime('%m/%d/%Y')
            except Exception:
                pass

    # ------------------ OPTION C REDESIGN ------------------
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    ws = wb.create_sheet(title=f"WVI {combined['profile']}")
    ws.views.sheetView[0].showGridLines = False
    
    # Modern Styling Variables
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.drawing.image import Image
    import os
    
    header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid") # Crimson
    section_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Light Gray
    alt_row_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    
    header_font = Font(color="FFFFFF", bold=True, size=14)
    section_font = Font(color="1E293B", bold=True, size=12)
    label_font = Font(bold=True, color="475569")
    data_font = Font(color="000000")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    # Title
    ws.merge_cells('A1:D2')
    title_cell = ws['A1']
    title_cell.value = f"Waste Verification Instructions (WVI)"
    title_cell.font = Font(color="000000", bold=True, size=16)
    title_cell.fill = PatternFill(fill_type=None)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Logo
    try:
        from flask import current_app
        logo_path = os.path.join(current_app.root_path, 'static', 'CLH_BIG.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            img.height = 40
            img.width = 135
            ws.add_image(img, 'A1')
    except Exception as e:
        print(f"Could not load logo: {e}")
        pass
    
    # Section 1: Profile Info
    ws['A4'] = "Profile Number:"
    ws['B4'] = combined['profile']
    ws['C4'] = "Approved Date:"
    ws['D4'] = combined['approved_date']
    
    ws['A5'] = "Generator Name:"
    ws['B5'] = combined['generator_name']
    ws['C5'] = "Expiration Date:"
    ws['D5'] = combined['expiration_date']

    ws['A6'] = "Waste Name:"
    ws['B6'] = combined['waste_name']
    ws['C6'] = "Lab Number:"
    ws['D6'] = combined['lab_num']
    
    for row in range(4, 7):
        ws[f'A{row}'].font = label_font
        ws[f'C{row}'].font = label_font
        ws[f'B{row}'].font = data_font
        ws[f'D{row}'].font = data_font
    
    # Section 2: Physical & Chemical Properties
    row = 8
    ws.merge_cells(f'A{row}:D{row}')
    sec1 = ws[f'A{row}']
    sec1.value = "Physical & Chemical Properties"
    sec1.font = section_font
    sec1.fill = section_fill
    sec1.alignment = Alignment(horizontal='left', vertical='center')
    
    ph_str = f"{combined['ph_min']} - {combined['ph_max']}" if combined['ph_min'] and combined['ph_max'] else (combined['ph_min'] or combined['ph_max'] or "")

    properties = [
        ("Physical Description:", combined['physical_description'], "pH Range:", ph_str),
        ("Flash Point:", combined['flashpoint'], "Cyanide:", combined['cyanide']),
        ("Free Liquids:", combined['free_liquids'], "Sulfides:", combined['sulfides']),
        ("VOC PPM:", combined['voc_ppm'], "", "")
    ]
    
    row += 1
    for i, prop in enumerate(properties):
        ws[f'A{row}'].value = prop[0]
        ws[f'B{row}'].value = prop[1]
        ws[f'C{row}'].value = prop[2]
        ws[f'D{row}'].value = prop[3]
        
        ws[f'A{row}'].font = label_font
        ws[f'C{row}'].font = label_font
        
        if prop[0] == "VOC PPM:":
            ws[f'B{row}'].alignment = Alignment(horizontal='left')
        
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = alt_row_fill
        row += 1
        
    # Section 3: Regulatory & Shipping
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    sec2 = ws[f'A{row}']
    sec2.value = "Regulatory & Shipping Information"
    sec2.font = section_font
    sec2.fill = section_fill
    
    ldr_display = combined['ldr']
    if combined.get('ldr_option') and combined['ldr_option'] in LDR_OPTIONS_MAP:
        ldr_display = f"{combined['ldr']} - Option {combined['ldr_option']}: {LDR_OPTIONS_MAP[combined['ldr_option']]}"
        
    reg_info = [
        ("DOT Description:", combined['dot_description'], "", ""),
        ("State Waste Codes:", combined['state_waste_codes'], "Federal Waste Codes:", combined['federal_waste_codes']),
        ("Reactivity Codes:", combined['reactivity_codes'], "", ""),
        ("LDR Required:", ldr_display, "", "")
    ]
    
    row += 1
    for i, reg in enumerate(reg_info):
        ws[f'A{row}'].value = reg[0]
        ws[f'B{row}'].value = reg[1]
        
        if reg[0] in ["DOT Description:", "LDR Required:"]:
            ws.merge_cells(f'B{row}:D{row}')
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='center')
            ws.row_dimensions[row].height = 20
        else:
            ws[f'C{row}'].value = reg[2]
            ws[f'D{row}'].value = reg[3]
            ws[f'C{row}'].font = label_font
            
        ws[f'A{row}'].font = label_font
        
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = alt_row_fill
        row += 1
        
    # Section 4: Processing & Instructions
    row += 1
    ws.merge_cells(f'A{row}:D{row}')
    sec3 = ws[f'A{row}']
    sec3.value = "Processing & Instructions"
    sec3.font = section_font
    sec3.fill = section_fill
    
    win_code = str(p_row['win_code'] if p_row else '').strip().upper()
    
    if win_code in ['CBP', 'CNIA', 'CBPR']:
        disposal_loc = "DIRECT TO LANDFILL / ACTIVE CELL"
    elif win_code in ['CBPS', 'CNOS']:
        disposal_loc = "UNIT 31"
    elif win_code in ['CCS', 'CCSM'] or win_code.startswith('CCS'):
        disposal_loc = "DIRECT TO BAYS / STU / TTB"
    else:
        disposal_loc = f"DISPOSAL LOCATION DETERMINED BY WIN ID: {win_code}" if win_code else "TBD"
        
    handling_inst = str(combined.get('handling_instruction') or '').strip().upper()
    if not handling_inst:
        if win_code == 'CNIA':
            handling_inst = "CONTAINS ASBESTOS DO NOT SAMPLE"
        else:
            handling_inst = "A MINIMUM OF SAFETY GLASSES, GLOVES, BOOTS, TYVEK, & RESPIRATOR"
            
    treatment_info = str(combined.get('treatment_information') or '').strip().upper()
    if not treatment_info or treatment_info in ['NONE', 'NOT APPLICABLE', 'SEE TREATMENT INFORMATION']:
        if p_row and 'treatment_recipe' in p_row.keys() and p_row['treatment_recipe'] is not None and str(p_row['treatment_recipe']).strip():
            treatment_info = str(p_row['treatment_recipe']).strip().upper()
        else:
            if win_code in ['CCS', 'CCSM']:
                treatment_info = "SEE TREATMENT INFORMATION"
            else:
                treatment_info = "NONE"
        
    sample_proc = str(combined.get('sample_procedures') or '').strip().upper()
    if not sample_proc:
        if win_code in ['CBPS', 'CNOS']:
            sample_proc = "COLLECT SAMPLE WITH COLIWASA"
        elif win_code == 'CNIA':
            sample_proc = "NO SAMPLE REQUIRED"
        else:
            sample_proc = "COLLECT SAMPLE WITH SCOOP"

    proc_info = [
        ("Disposal Location:", disposal_loc),
        ("Sample Procedures:", sample_proc),
        ("Handling Instructions:", handling_inst),
        ("Unloading Instructions:", combined['unloading_instructions']),
        ("Treatment Information:", treatment_info),
        ("Verification Procedures:", combined['verification_procedures']),
        ("Notes / Revisions:", combined['notes_revisions'])
    ]
    
    row += 1
    for i, proc in enumerate(proc_info):
        ws[f'A{row}'].value = proc[0]
        ws[f'B{row}'].value = proc[1]
        ws.merge_cells(f'B{row}:D{row}')
        
        ws[f'A{row}'].font = label_font
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 20
        
        if i % 2 == 0:
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].fill = alt_row_fill
        row += 1
        
    # Column Widths & Print Setup
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 51
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 32
    
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = False
    ws.page_setup.fitToHeight = False
    
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    
    # Save to a memory buffer
    output_io = io.BytesIO()
    wb.save(output_io)
    output_io.seek(0)

    filename = f"{combined['profile']}.xlsx"
    return send_file(
        output_io,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )


@approvals_bp.route('/api/profile/<path:profile_number>/audit_export')
def export_profile_audit(profile_number):
    import io
    import datetime
    from flask import request, render_template_string, send_file, Response
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    profile_clean = str(profile_number).strip().upper()
    export_format = request.args.get('format', 'html').lower()

    with closing(get_db_connection()) as conn:
        p_row = conn.execute('SELECT * FROM profiles WHERE TRIM(UPPER(profile_number)) = ?', (profile_clean,)).fetchone()
        w_row = conn.execute('SELECT * FROM profile_wvi WHERE TRIM(UPPER(profile)) = ?', (profile_clean,)).fetchone()
        
        # Load shipment history from drum_inventory and truck_logs
        drum_loads = conn.execute('''
            SELECT 
                manifest, 
                job_id, 
                import_date, 
                SUM(COALESCE(weight, 0)) AS weight, 
                COALESCE(process_type, 'Containerized') AS process_type, 
                MAX(location) AS cell_location,
                '' AS grid_location,
                'Drum' AS load_source
            FROM drum_inventory
            WHERE TRIM(UPPER(inb_prof)) = ?
            GROUP BY COALESCE(job_id, manifest), import_date
        ''', (profile_clean,)).fetchall()

        truck_loads = conn.execute('''
            SELECT 
                COALESCE(manifest_number, '---') AS manifest,
                COALESCE(truck_id, load_number, '---') AS job_id,
                COALESCE(date_received, time_in, '---') AS import_date,
                COALESCE(net_weight, CASE WHEN (gross_weight > 0 AND exit_weight > 0) THEN (gross_weight - exit_weight) ELSE gross_weight END, 0) AS weight,
                COALESCE(container_type, shipping_mode, 'Yellow Entry') AS process_type,
                cell_location,
                grid_location,
                'Yellow / Bulk' AS load_source
            FROM truck_logs
            WHERE TRIM(UPPER(profile_number)) = ?
              AND UPPER(COALESCE(test_status, '')) != 'VOID'
        ''', (profile_clean,)).fetchall()

        def format_location(cell, grid):
            c_str = str(cell).strip() if cell and str(cell).strip().upper() not in ['NONE', 'NULL', ''] else ''
            g_str = str(grid).strip() if grid and str(grid).strip().upper() not in ['NONE', 'NULL', ''] else ''
            if c_str and g_str:
                return f"Cell {c_str} | Grid {g_str}"
            elif c_str:
                return f"Cell {c_str}"
            elif g_str:
                return f"Grid {g_str}"
            return "---"

        history_rows = []
        for l in list(drum_loads) + list(truck_loads):
            d = dict(l)
            d['location_fmt'] = format_location(d.get('cell_location'), d.get('grid_location'))
            history_rows.append(d)
        history_rows.sort(key=lambda x: str(x.get('import_date') or ''), reverse=True)

        # Load drum QA queue history
        drum_rows_raw = conn.execute('''
            SELECT q.drum_id, q.job_id, q.manifest, q.tests_required, q.status, MAX(i.import_date) AS import_date
            FROM drum_lab_queue q
            LEFT JOIN drum_inventory i ON TRIM(UPPER(q.drum_id)) = TRIM(UPPER(i.track_no))
            WHERE TRIM(UPPER(q.profile)) = ?
            GROUP BY q.id
            ORDER BY import_date DESC, q.id DESC
        ''', (profile_clean,)).fetchall()
        drum_rows = [dict(d) for d in drum_rows_raw]

    if not p_row and not w_row:
        return "Profile not found in database", 404

    # Build combined data dict
    p = dict(p_row) if p_row else {}
    w = dict(w_row) if w_row else {}
    
    profile_num = p.get('profile_number') or w.get('profile') or profile_clean
    generator = p.get('generator') or p.get('generator_name') or w.get('generator_name') or '---'
    epa_id = p.get('epa_id') or '---'
    customer = p.get('customer_name') or '---'
    status = p.get('status') or 'UNKNOWN'
    exp_date = p.get('expiration_date') or w.get('expiration_date') or '---'
    app_date = w.get('approved_date') or '---'
    win_code = p.get('win_code') or '---'
    lab_num = p.get('lab_number') or w.get('lab_num') or '---'
    container_type = p.get('shipping_container_type') or '---'
    waste_desc = p.get('waste_description') or p.get('waste_name') or w.get('waste_name') or '---'

    # Chemical / Physical specs
    phys_state = p.get('physical_appearance') or w.get('physical_description') or '---'
    color = p.get('color') or w.get('color') or '---'
    ph_range = p.get('ph_range') or (f"{w.get('ph_min', '')} - {w.get('ph_max', '')}".strip(" -") if (w.get('ph_min') or w.get('ph_max')) else '---')
    flash_point = p.get('flash_point') or w.get('flashpoint') or '---'
    cyanide = p.get('cyanide') or w.get('cyanide') or 'NEG'
    sulfide = p.get('sulfide') or w.get('sulfides') or 'NEG'
    free_liquids = p.get('free_liquids') or w.get('free_liquids') or 'NO'
    voc_ppm = p.get('voc_percentage') or w.get('voc_ppm') or '0'

    # Regulatory
    fed_codes = p.get('federal_waste_code') or w.get('federal_waste_codes') or '---'
    state_codes = p.get('state_waste_code') or w.get('state_waste_codes') or '---'
    dot_desc = p.get('dot_description') or w.get('dot_description') or '---'
    ldr_req = p.get('ldr_required') or w.get('ldr') or '---'
    ldr_opt = p.get('ldr_option') or '---'
    recipe = p.get('treatment_recipe') or w.get('treatment_information') or '---'
    special_handling = p.get('special_handling') or w.get('handling_instruction') or '---'

    if export_format in ['excel', 'xlsx']:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profile Audit Summary"
        ws.views.sheetView[0].showGridLines = True

        header_fill = PatternFill(start_color="8B0000", end_color="8B0000", fill_type="solid") # Dark Crimson
        subheader_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate

        header_font = Font(color="FFFFFF", bold=True, size=14)
        subheader_font = Font(color="FFFFFF", bold=True, size=11)
        label_font = Font(bold=True, color="334155")
        data_font = Font(color="000000")
        thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                             top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

        # Title Banner
        ws.merge_cells('A1:F2')
        title_cell = ws['A1']
        title_cell.value = f"CLEAN HARBORS BUTTONWILLOW — PROFILE AUDIT EXPORT ({profile_num})"
        title_cell.font = header_font
        title_cell.fill = header_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        curr_row = 4
        def write_section(title):
            nonlocal curr_row
            ws.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=6)
            c = ws.cell(row=curr_row, column=1, value=title)
            c.font = subheader_font
            c.fill = subheader_fill
            c.alignment = Alignment(horizontal='left', vertical='center')
            curr_row += 1

        def write_kv(label, value, label2="", value2=""):
            nonlocal curr_row
            ws.cell(row=curr_row, column=1, value=label).font = label_font
            ws.cell(row=curr_row, column=2, value=str(value)).font = data_font
            if label2:
                ws.cell(row=curr_row, column=4, value=label2).font = label_font
                ws.cell(row=curr_row, column=5, value=str(value2)).font = data_font
            for col in range(1, 7):
                ws.cell(row=curr_row, column=col).border = thin_border
            curr_row += 1

        write_section("1. GENERAL PROFILE IDENTIFICATION")
        write_kv("Profile Number:", profile_num, "Status:", status)
        write_kv("Generator Name:", generator, "EPA ID:", epa_id)
        write_kv("Shipping Container Type:", container_type, "WIN Code:", win_code)
        write_kv("Lab Tracking #:", lab_num, "Approved Date:", app_date)
        write_kv("Expiration Date:", exp_date)
        write_kv("Waste Description:", waste_desc)

        curr_row += 1
        write_section("2. PHYSICAL & CHEMICAL CHARACTERISTICS")
        write_kv("Physical Appearance:", phys_state, "Color:", color)
        write_kv("pH Range:", ph_range, "Flash Point (°F):", flash_point)
        write_kv("Cyanide:", cyanide, "Sulfide:", sulfide)
        write_kv("Free Liquids:", free_liquids, "VOC Level (PPM):", voc_ppm)

        curr_row += 1
        write_section("3. REGULATORY CODES & LAND DISPOSAL RESTRICTIONS (LDR)")
        write_kv("Federal Waste Codes:", fed_codes, "State Waste Codes:", state_codes)
        write_kv("LDR Status:", ldr_req, "LDR Option:", ldr_opt)
        write_kv("DOT Description:", dot_desc)
        write_kv("Treatment Recipe:", recipe)
        write_kv("Special Handling:", special_handling)

        curr_row += 1
        write_section("4. RECENT SHIPMENT LOG SUMMARY")
        ws.cell(row=curr_row, column=1, value="Import Date").font = label_font
        ws.cell(row=curr_row, column=2, value="Manifest").font = label_font
        ws.cell(row=curr_row, column=3, value="Job ID").font = label_font
        ws.cell(row=curr_row, column=4, value="Net Weight (Lbs)").font = label_font
        ws.cell(row=curr_row, column=5, value="Cell / Grid Location").font = label_font
        ws.cell(row=curr_row, column=6, value="Source").font = label_font
        curr_row += 1
        for h in history_rows[:15]:
            ws.cell(row=curr_row, column=1, value=h.get('import_date', '')).font = data_font
            ws.cell(row=curr_row, column=2, value=h.get('manifest', '')).font = data_font
            ws.cell(row=curr_row, column=3, value=h.get('job_id', '')).font = data_font
            ws.cell(row=curr_row, column=4, value=h.get('weight', 0)).font = data_font
            ws.cell(row=curr_row, column=5, value=h.get('location_fmt', '---')).font = data_font
            ws.cell(row=curr_row, column=6, value=h.get('load_source', '')).font = data_font
            for col in range(1, 7):
                ws.cell(row=curr_row, column=col).border = thin_border
            curr_row += 1

        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 24
        ws.column_dimensions['E'].width = 28
        ws.column_dimensions['F'].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Profile_{profile_num}_Audit_Report.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )

    # HTML format (default)
    html_template = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Profile Audit Certificate - {{ profile_num }}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    <style>
        body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f8fafc; color: #1e293b; }
        .audit-container { max-width: 1000px; margin: 20px auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 20px rgba(0,0,0,0.08); }
        .audit-header { border-bottom: 3px solid #8b0000; padding-bottom: 15px; margin-bottom: 25px; }
        .banner-title { color: #8b0000; font-weight: 800; font-size: 1.6rem; letter-spacing: 0.5px; }
        .badge-status { font-size: 0.9rem; padding: 6px 14px; border-radius: 20px; font-weight: 700; }
        .section-header { background: #1e293b; color: #ffffff; padding: 8px 15px; font-weight: 700; border-radius: 6px; font-size: 1rem; margin-top: 25px; margin-bottom: 15px; }
        .info-label { font-weight: 700; color: #475569; font-size: 0.85rem; text-transform: uppercase; }
        .info-value { font-weight: 600; color: #0f172a; font-size: 0.95rem; }
        .table-custom { font-size: 0.88rem; }
        .table-custom th { background-color: #f1f5f9; font-weight: 700; color: #334155; }
        @media print {
            .no-print { display: none !important; }
            body { background: #ffffff; }
            .audit-container { box-shadow: none; padding: 0; max-width: 100%; }
        }
    </style>
</head>
<body>
    <div class="container no-print mt-3 max-w-1000" style="max-width: 1000px;">
        <div class="d-flex justify-content-between align-items-center bg-white p-3 rounded shadow-sm border">
            <div>
                <span class="fw-bold text-dark me-2">📄 Regulatory Audit Document View</span>
                <span class="badge bg-secondary">Profile: {{ profile_num }}</span>
            </div>
            <div class="d-flex gap-2">
                <button class="btn btn-sm btn-primary fw-bold px-3" onclick="window.print()">🖨️ Print / Save as PDF</button>
                <a class="btn btn-sm btn-success fw-bold px-3" href="?format=excel">📊 Export to Excel (.xlsx)</a>
                <button class="btn btn-sm btn-outline-secondary fw-bold px-3" onclick="window.close()">❌ Close</button>
            </div>
        </div>
    </div>

    <div class="audit-container">
        <!-- Header -->
        <div class="audit-header d-flex justify-content-between align-items-start">
            <div>
                <div class="banner-title">CLEAN HARBORS BUTTONWILLOW</div>
                <div class="fw-bold text-muted small">FACILITY WASTE ACCEPTANCE & REGULATORY PROFILE AUDIT SHEET</div>
                <div class="text-secondary small mt-1">Generated: {{ current_time }}</div>
            </div>
            <div class="text-end">
                <div class="fs-4 fw-bold text-dark">{{ profile_num }}</div>
                <span class="badge badge-status {% if status == 'APPROVED' %}bg-success{% elif status == 'REJECTED' or status == 'EXPIRED' %}bg-danger{% else %}bg-warning text-dark{% endif %}">
                    {{ status }}
                </span>
            </div>
        </div>

        <!-- 1. Profile Identification -->
        <div class="section-header">1. GENERAL PROFILE IDENTIFICATION</div>
        <div class="row g-3 px-2">
            <div class="col-md-6">
                <div class="info-label">Generator Name</div>
                <div class="info-value">{{ generator }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Generator EPA ID</div>
                <div class="info-value">{{ epa_id }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Shipping Container Type</div>
                <div class="info-value">{{ container_type }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">BL Waste Code (WIN)</div>
                <div class="info-value">{{ win_code }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Lab Tracking Number</div>
                <div class="info-value">{{ lab_num }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Approved Date</div>
                <div class="info-value">{{ app_date }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Expiration Date</div>
                <div class="info-value">{{ exp_date }}</div>
            </div>
            <div class="col-md-12">
                <div class="info-label">Waste Description</div>
                <div class="info-value">{{ waste_desc }}</div>
            </div>
        </div>

        <!-- 2. Physical & Chemical Characteristics -->
        <div class="section-header">2. PHYSICAL & CHEMICAL CHARACTERISTICS</div>
        <div class="row g-3 px-2">
            <div class="col-md-3">
                <div class="info-label">Physical State / Appearance</div>
                <div class="info-value">{{ phys_state }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Color</div>
                <div class="info-value">{{ color }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">pH Range</div>
                <div class="info-value">{{ ph_range }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Flash Point (°F)</div>
                <div class="info-value">{{ flash_point }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Cyanide</div>
                <div class="info-value">{{ cyanide }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Reactive Sulfide</div>
                <div class="info-value">{{ sulfide }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">Free Liquids</div>
                <div class="info-value">{{ free_liquids }}</div>
            </div>
            <div class="col-md-3">
                <div class="info-label">VOC Level (PPM)</div>
                <div class="info-value">{{ voc_ppm }}</div>
            </div>
        </div>

        <!-- 3. Regulatory & LDR -->
        <div class="section-header">3. REGULATORY CODES & LAND DISPOSAL RESTRICTIONS (LDR)</div>
        <div class="row g-3 px-2">
            <div class="col-md-6">
                <div class="info-label">Federal RCRA Waste Codes</div>
                <div class="info-value">{{ fed_codes }}</div>
            </div>
            <div class="col-md-6">
                <div class="info-label">State Waste Codes</div>
                <div class="info-value">{{ state_codes }}</div>
            </div>
            <div class="col-md-4">
                <div class="info-label">LDR Determination Status</div>
                <div class="info-value">{{ ldr_req }}</div>
            </div>
            <div class="col-md-4">
                <div class="info-label">LDR Category / Option</div>
                <div class="info-value">{{ ldr_opt }}</div>
            </div>
            <div class="col-md-4">
                <div class="info-label">Assigned Treatment Recipe</div>
                <div class="info-value">{{ recipe }}</div>
            </div>
            <div class="col-md-12">
                <div class="info-label">DOT Proper Shipping Name & Description</div>
                <div class="info-value">{{ dot_desc }}</div>
            </div>
            <div class="col-md-12">
                <div class="info-label">Special Handling & Operating Instructions</div>
                <div class="info-value">{{ special_handling }}</div>
            </div>
        </div>

        <!-- 4. Shipment History Summary -->
        <div class="section-header">4. SHIPMENT & QA VERIFICATION LOG (RECENT 50)</div>
        {% if history_rows %}
        <div class="table-responsive">
            <table class="table table-bordered table-sm table-custom mt-2">
                <thead>
                    <tr>
                        <th>Import Date</th>
                        <th>Manifest #</th>
                        <th>Job ID</th>
                        <th class="text-end">Weight (Lbs)</th>
                        <th>Cell / Grid Location</th>
                        <th>Entry Source</th>
                    </tr>
                </thead>
                <tbody>
                    {% for h in history_rows %}
                    <tr>
                        <td>{{ h.import_date or '---' }}</td>
                        <td class="fw-bold">{{ h.manifest or '---' }}</td>
                        <td>{{ h.job_id or '---' }}</td>
                        <td class="text-end">{{ h.weight or '0' }}</td>
                        <td><span class="badge bg-light text-dark border font-monospace">{{ h.location_fmt or '---' }}</span></td>
                        <td><span class="badge bg-light text-dark border">{{ h.load_source or 'Log' }}</span></td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% else %}
        <div class="text-muted small px-2 py-2">No recorded shipment history found for this profile.</div>
        {% endif %}

        {% if drum_rows %}
        <div class="fw-bold text-dark mt-3 mb-2 small">DRUM TESTING & QA/QC LOGS</div>
        <div class="table-responsive">
            <table class="table table-bordered table-sm table-custom">
                <thead>
                    <tr>
                        <th>Date</th>
                        <th>Drum ID</th>
                        <th>Manifest</th>
                        <th>Job ID</th>
                        <th>Tests Required</th>
                        <th>Status</th>
                    </tr>
                </thead>
                <tbody>
                    {% for d in drum_rows %}
                    <tr>
                        <td>{{ d.import_date or '---' }}</td>
                        <td class="font-monospace fw-bold">{{ d.drum_id or '---' }}</td>
                        <td>{{ d.manifest or '---' }}</td>
                        <td>{{ d.job_id or '---' }}</td>
                        <td>{{ d.tests_required or '---' }}</td>
                        <td><span class="badge bg-secondary">{{ d.status or 'PENDING' }}</span></td>
                    </tr>
                    {% endfor %}
                </tbody>
            </table>
        </div>
        {% endif %}

        <div class="border-top pt-3 mt-4 text-center text-muted small">
            <div>Clean Harbors Environmental Services — Buttonwillow Facility Compliance Portal</div>
            <div>This document is generated for regulatory inspection, audit, and waste analysis plan (WAP) compliance review.</div>
        </div>
    </div>
</body>
</html>'''

    current_time_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    rendered_html = render_template_string(
        html_template,
        profile_num=profile_num,
        generator=generator,
        epa_id=epa_id,
        customer=customer,
        status=status,
        exp_date=exp_date,
        app_date=app_date,
        win_code=win_code,
        lab_num=lab_num,
        container_type=container_type,
        waste_desc=waste_desc,
        phys_state=phys_state,
        color=color,
        ph_range=ph_range,
        flash_point=flash_point,
        cyanide=cyanide,
        sulfide=sulfide,
        free_liquids=free_liquids,
        voc_ppm=voc_ppm,
        fed_codes=fed_codes,
        state_codes=state_codes,
        ldr_req=ldr_req,
        ldr_opt=ldr_opt,
        recipe=recipe,
        dot_desc=dot_desc,
        special_handling=special_handling,
        history_rows=history_rows,
        drum_rows=drum_rows,
        current_time=current_time_str
    )

    return Response(rendered_html, mimetype='text/html')


