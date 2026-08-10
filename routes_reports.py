import os, glob, re, difflib, openpyxl, json
from flask import Blueprint, render_template, request, redirect, url_for, send_file, jsonify
from datetime import date, datetime, timedelta
from contextlib import closing
from io import BytesIO
import pandas as pd
from collections import defaultdict
from database import get_db_connection


reports_bp = Blueprint('reports_bp', __name__)

def build_2026voc_excel(trucks, selected_date):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    
    # Setup exact 5 sheets matching I:\Buttonwillow\VOCs and TONNAGE by YEAR\2026VOC
    ws35 = wb.active
    ws35.title = '35'
    ws_non = wb.create_sheet(title='NON-VOC')
    ws31 = wb.create_sheet(title='31 GALLONS ')
    ws_s34 = wb.create_sheet(title='STORAGE TO 34')
    ws_tw = wb.create_sheet(title='TREATED WASTE (34 TO 35)')

    # Sheet 1: '35'
    ws35['A1'] = 'DAILY WMU 35 VOC TRACKING'
    ws35['A1'].font = Font(bold=True, size=12)
    ws35.append([]) # Row 2 blank
    headers_35 = ['RECEIVED', 'VOCs (ppm)', 'Weight (Pounds)', 'Weight (TONs)', 'VOCSxWt.', 'MANIFEST #', 'APPROVAL #', 'WMU', 'Weighmaster Load No.']
    ws35.append(headers_35)
    for col in range(1, 10):
        ws35.cell(row=3, column=col).font = Font(bold=True)

    # Sheet 2: 'NON-VOC'
    ws_non['A1'] = "DAILY NON-VOC'S REPORT"
    ws_non['A1'].font = Font(bold=True, size=12)
    headers_non = ['RECEIVED', 'VOCs ppm', 'Pounds', 'TONs', 'VOCS/Wt.', 'MANIFEST #', 'APPROVAL #', 'WMU', 'Weighmaster Load No.']
    ws_non.append(headers_non)
    for col in range(1, 10):
        ws_non.cell(row=2, column=col).font = Font(bold=True)

    # Sheet 3: '31 GALLONS '
    ws31.append([])
    ws31['A2'] = 'GALLONS PER LOAD CALCULATOR'
    ws31['A2'].font = Font(bold=True, size=12)
    ws31.append([])
    headers_31 = ['SPECIFIC GRAVITY', 'NET (LBS) WEIGHT OF LOAD', 'GALLONS PER LOAD', 'WEIGHT TICKET #', 'WMU #']
    ws31.append(headers_31)
    for col in range(1, 6):
        ws31.cell(row=4, column=col).font = Font(bold=True)

    # Sheet 4: 'STORAGE TO 34'
    ws_s34['A1'] = 'TREATED WASTE TRANSFER TO WMU 34 (for storage) '
    ws_s34['A1'].font = Font(bold=True, size=12)
    headers_s34 = ['DATE', 'STU Batch #s', 'Waste (Tons)', 'Additive (Tons)', 'Waste + Additive (Tons)', 'Water (Tons)', 'VOCs (ppm)', 'VOC/Wt. (PPM)', 'Final Grid', 'Date Moved', 'Cell 35-?', 'Note:', 'ZERO VOC']
    ws_s34.append(headers_s34)
    for col in range(1, 14):
        ws_s34.cell(row=2, column=col).font = Font(bold=True)

    # Sheet 5: 'TREATED WASTE (34 TO 35)'
    ws_tw['A1'] = 'TREATED WASTE TRANSFER TO WMU 35 from STORAGE (from map)'
    ws_tw['A1'].font = Font(bold=True, size=12)
    headers_tw = ['Date ', 'STU Batch #s', 'Waste (Tons)', 'Additive (Tons)', 'Waste + Additive (Tons)', 'Water (Tons) ', 'VOCs (ppm)', 'VOC/Wt.', 'Final Grid', 'Date', 'Cell 35-?']
    ws_tw.append(headers_tw)
    for col in range(1, 12):
        ws_tw.cell(row=2, column=col).font = Font(bold=True)

    row_35_idx = 4
    row_non_idx = 3
    row_31_idx = 5

    try:
        dt_received = datetime.strptime(selected_date, '%Y-%m-%d').date()
    except Exception:
        dt_received = selected_date

    active_trucks = [t for t in trucks if t.get('test_status') != 'VOID']
    voided_trucks = [t for t in trucks if t.get('test_status') == 'VOID']

    for t in active_trucks:
        wmu_cell = str(t.get('cell_location', '') or '').strip().upper()
        wmu_grid = str(t.get('grid_location', '') or '').strip().upper()
        
        if wmu_cell and wmu_grid:
            wmu = f"{wmu_cell}/{wmu_grid}"
        elif wmu_cell:
            wmu = wmu_cell
        elif wmu_grid:
            wmu = wmu_grid
        else:
            wmu = '35-7' if '35' in str(t.get('win_code', '')) else ('31' if '31' in str(t.get('win_code', '')) else '35-7')

        try: voc = float(t.get('measured_voc') if t.get('measured_voc') is not None else (t.get('voc_percentage') or 0.0))
        except (ValueError, TypeError): voc = 0.0

        try:
            gross = float(t.get('gross_weight') or 0)
            tare = float(t.get('exit_weight') or 0)
            lbs = gross - tare if gross > tare else gross
        except (ValueError, TypeError):
            lbs = 0.0

        manifest = str(t.get('manifest_number', '') or '').strip()
        profile = str(t.get('profile_number', '') or '').strip()
        ticket = t.get('truck_id') or t.get('load_number', '')
        try:
            ticket = int(ticket)
        except Exception:
            pass

        # Sheet 1: '35'
        if wmu.startswith('35') or '35' in wmu:
            ws35.cell(row=row_35_idx, column=1, value=dt_received)
            ws35.cell(row=row_35_idx, column=2, value=voc)
            ws35.cell(row=row_35_idx, column=3, value=lbs)
            ws35.cell(row=row_35_idx, column=4, value=f"=C{row_35_idx}/2000")
            ws35.cell(row=row_35_idx, column=5, value=f"=D{row_35_idx}*B{row_35_idx}")
            ws35.cell(row=row_35_idx, column=6, value=manifest)
            ws35.cell(row=row_35_idx, column=7, value=profile)
            ws35.cell(row=row_35_idx, column=8, value=wmu)
            ws35.cell(row=row_35_idx, column=9, value=ticket)
            row_35_idx += 1

        # Sheet 2: 'NON-VOC' (Only Unit 31 or STU loads)
        is_unit_31 = '31' in wmu or wmu.startswith('31') or wmu.startswith('U31')
        is_stu = '34' in wmu or 'STU' in wmu or 'BAY' in wmu or wmu in ['CCS', 'CCSF', 'CCSM']
        if is_unit_31 or is_stu:
            ws_non.cell(row=row_non_idx, column=1, value=dt_received)
            ws_non.cell(row=row_non_idx, column=2, value=voc)
            ws_non.cell(row=row_non_idx, column=3, value=lbs)
            ws_non.cell(row=row_non_idx, column=4, value=f"=C{row_non_idx}/2000")
            ws_non.cell(row=row_non_idx, column=5, value=f"=D{row_non_idx}*B{row_non_idx}")
            ws_non.cell(row=row_non_idx, column=6, value=manifest)
            ws_non.cell(row=row_non_idx, column=7, value=profile)
            ws_non.cell(row=row_non_idx, column=8, value=wmu)
            ws_non.cell(row=row_non_idx, column=9, value=ticket)
            row_non_idx += 1

        # Sheet 3: '31 GALLONS '
        if '31' in wmu or str(t.get('win_code', '')).upper() in ['CBPS', 'CNOS']:
            try: sg = float(t.get('specific_gravity') or 1.0)
            except Exception: sg = 1.0
            if sg <= 0: sg = 1.0

            ws31.cell(row=row_31_idx, column=1, value=sg)
            ws31.cell(row=row_31_idx, column=2, value=lbs)
            ws31.cell(row=row_31_idx, column=3, value=f"=IFERROR(B{row_31_idx}/(A{row_31_idx}*8.33), 0)")
            ws31.cell(row=row_31_idx, column=4, value=ticket)
            ws31.cell(row=row_31_idx, column=5, value=wmu)
            row_31_idx += 1

    # Totals for Sheet '35'
    if row_35_idx > 4:
        ws35.cell(row=row_35_idx, column=1, value="TOTAL").font = Font(bold=True)
        ws35.cell(row=row_35_idx, column=3, value=f"=SUM(C4:C{row_35_idx-1})").font = Font(bold=True)
        ws35.cell(row=row_35_idx, column=4, value=f"=SUM(D4:D{row_35_idx-1})").font = Font(bold=True)
        ws35.cell(row=row_35_idx, column=5, value=f"=SUM(E4:E{row_35_idx-1})").font = Font(bold=True)

    # Totals for Sheet 'NON-VOC'
    if row_non_idx > 3:
        ws_non.cell(row=row_non_idx, column=1, value="TOTAL").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=3, value=f"=SUM(C3:C{row_non_idx-1})").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=4, value=f"=SUM(D3:D{row_non_idx-1})").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=5, value=f"=SUM(E3:E{row_non_idx-1})").font = Font(bold=True)

    # Totals for Sheet '31 GALLONS '
    if row_31_idx > 5:
        ws31.cell(row=row_31_idx, column=1, value="TOTAL").font = Font(bold=True)
        ws31.cell(row=row_31_idx, column=2, value=f"=SUM(B5:B{row_31_idx-1})").font = Font(bold=True)
        ws31.cell(row=row_31_idx, column=3, value=f"=SUM(C5:C{row_31_idx-1})").font = Font(bold=True)

    # Section for VOID Tickets on Sheet 'NON-VOC'
    if voided_trucks:
        row_non_idx += 2
        ws_non.cell(row=row_non_idx, column=1, value="VOID").font = Font(bold=True, size=12)
        row_non_idx += 1
        headers_non = ['RECEIVED', 'VOCs ppm', 'Pounds', 'TONs', 'VOCS/Wt.', 'MANIFEST #', 'APPROVAL #', 'WMU', 'Weighmaster Load No.']
        for col_idx, h_text in enumerate(headers_non, start=1):
            ws_non.cell(row=row_non_idx, column=col_idx, value=h_text).font = Font(bold=True)
        row_non_idx += 1
        start_void_row = row_non_idx

        for t in voided_trucks:
            wmu_cell = str(t.get('cell_location', '') or '').strip().upper()
            wmu_grid = str(t.get('grid_location', '') or '').strip().upper()
            if wmu_cell and wmu_grid:
                wmu = f"{wmu_cell}/{wmu_grid}"
            elif wmu_cell:
                wmu = wmu_cell
            elif wmu_grid:
                wmu = wmu_grid
            else:
                wmu = '35-7' if '35' in str(t.get('win_code', '')) else ('31' if '31' in str(t.get('win_code', '')) else '35-7')

            try: voc = float(t.get('measured_voc') if t.get('measured_voc') is not None else (t.get('voc_percentage') or 0.0))
            except (ValueError, TypeError): voc = 0.0

            try:
                gross = float(t.get('gross_weight') or 0)
                tare = float(t.get('exit_weight') or 0)
                lbs = gross - tare if gross > tare else gross
            except (ValueError, TypeError):
                lbs = 0.0

            manifest = str(t.get('manifest_number', '') or '').strip()
            profile = str(t.get('profile_number', '') or '').strip()
            ticket = t.get('truck_id') or t.get('load_number', '')
            try:
                ticket = int(ticket)
            except Exception:
                pass

            ws_non.cell(row=row_non_idx, column=1, value=dt_received)
            ws_non.cell(row=row_non_idx, column=2, value=voc)
            ws_non.cell(row=row_non_idx, column=3, value=lbs)
            ws_non.cell(row=row_non_idx, column=4, value=f"=C{row_non_idx}/2000")
            ws_non.cell(row=row_non_idx, column=5, value=f"=D{row_non_idx}*B{row_non_idx}")
            ws_non.cell(row=row_non_idx, column=6, value=manifest)
            ws_non.cell(row=row_non_idx, column=7, value=profile)
            ws_non.cell(row=row_non_idx, column=8, value=wmu)
            ws_non.cell(row=row_non_idx, column=9, value=ticket)
            row_non_idx += 1

        ws_non.cell(row=row_non_idx, column=1, value="TOTAL VOID").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=3, value=f"=SUM(C{start_void_row}:C{row_non_idx-1})").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=4, value=f"=SUM(D{start_void_row}:D{row_non_idx-1})").font = Font(bold=True)
        ws_non.cell(row=row_non_idx, column=5, value=f"=SUM(E{start_void_row}:E{row_non_idx-1})").font = Font(bold=True)

    # Auto-fit column widths generously across all sheets
    from openpyxl.utils import get_column_letter
    for ws in [ws35, ws_non, ws31, ws_s34, ws_tw]:
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if val.startswith('='):
                    max_len = max(max_len, 14)
                else:
                    max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = max(max_len + 5, 18)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@reports_bp.route('/export')
def export_excel():
    selected_date = request.args.get('date', date.today().isoformat())

    with closing(get_db_connection()) as conn:
        trucks_raw = conn.execute('''
            SELECT t.*, p.voc_percentage, p.win_code
            FROM truck_logs t
            LEFT JOIN profiles p ON TRIM(UPPER(t.profile_number)) = TRIM(UPPER(p.profile_number))
            WHERE t.date_received = ? AND t.test_status != 'REJECTED'
            ORDER BY CAST(t.load_number AS INTEGER) ASC
        ''', (selected_date,)).fetchall()
        
    trucks = [dict(r) for r in trucks_raw]
    excel_stream = build_2026voc_excel(trucks, selected_date)
    
    try:
        dt = datetime.strptime(selected_date, '%Y-%m-%d')
        file_name = f"{dt.strftime('%m%d%Y')}VOC.xlsx"
    except Exception:
        file_name = f"Daily_VOC_Tracking_{selected_date}.xlsx"

    return send_file(
        excel_stream, 
        as_attachment=True, 
        download_name=file_name, 
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@reports_bp.route('/export_daily_styled_excel')
def export_daily_styled_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    selected_date = request.args.get('date', date.today().isoformat()).strip()

    with closing(get_db_connection()) as conn:
        # Inbound truck logs for selected date
        trucks_raw = conn.execute('''
            SELECT t.*, p.generator, p.win_code, p.waste_description, p.shipping_container_type
            FROM truck_logs t
            LEFT JOIN profiles p ON TRIM(UPPER(t.profile_number)) = TRIM(UPPER(p.profile_number))
            WHERE t.date_received = ? AND (t.test_status IS NULL OR t.test_status != 'REJECTED')
            ORDER BY t.id DESC
        ''', (selected_date,)).fetchall()
        
        # Scheduled loads per profile for selected date
        sched_rows = conn.execute('''
            SELECT TRIM(UPPER(profile_number)) as profile, SUM(load_count) as scheduled
            FROM daily_schedule
            WHERE schedule_date = ?
            GROUP BY TRIM(UPPER(profile_number))
        ''', (selected_date,)).fetchall()
        sched_map = {r['profile']: (r['scheduled'] or 0) for r in sched_rows if r['profile']}

        # Actual received loads per profile for selected date
        rec_rows = conn.execute('''
            SELECT TRIM(UPPER(t.profile_number)) as profile, p.generator, p.win_code,
                   COUNT(t.id) as received_count,
                   SUM(t.gross_weight) as total_lbs,
                   AVG(t.measured_voc) as avg_voc
            FROM truck_logs t
            LEFT JOIN profiles p ON TRIM(UPPER(t.profile_number)) = TRIM(UPPER(p.profile_number))
            WHERE t.date_received = ? AND (t.test_status IS NULL OR t.test_status != 'REJECTED')
            GROUP BY TRIM(UPPER(t.profile_number))
        ''', (selected_date,)).fetchall()
        rec_map = {r['profile']: dict(r) for r in rec_rows if r['profile']}

    trucks = [dict(r) for r in trucks_raw]
    
    # Merge distinct profiles from schedule & receiving
    all_prof_keys = list(set(sched_map.keys()).union(set(rec_map.keys())))
    missing_keys = [k for k in all_prof_keys if k not in rec_map]
    missing_prof_info = {}
    if missing_keys:
        with closing(get_db_connection()) as conn:
            placeholders = ','.join(['?'] * len(missing_keys))
            p_info_rows = conn.execute(f'''
                SELECT profile_number, generator, win_code
                FROM profiles
                WHERE TRIM(UPPER(profile_number)) IN ({placeholders})
            ''', missing_keys).fetchall()
            for pr in p_info_rows:
                missing_prof_info[str(pr['profile_number']).strip().upper()] = dict(pr)

    combined_profiles = []
    for prof_key in all_prof_keys:
        r_info = rec_map.get(prof_key, {})
        m_info = missing_prof_info.get(prof_key, {})
        
        gen_name = r_info.get('generator') or m_info.get('generator') or '---'
        win_code = r_info.get('win_code') or m_info.get('win_code') or '---'
        sched_cnt = sched_map.get(prof_key, 0)
        rec_cnt = r_info.get('received_count', 0)
        variance = rec_cnt - sched_cnt
        total_lbs = float(r_info.get('total_lbs') or 0)
        avg_voc = float(r_info.get('avg_voc') or 0)
        
        combined_profiles.append({
            'profile_number': prof_key,
            'generator': gen_name,
            'win_code': win_code,
            'scheduled': sched_cnt,
            'received': rec_cnt,
            'variance': variance,
            'total_lbs': total_lbs,
            'avg_voc': avg_voc
        })

    def get_variance_sort_key(p):
        sched = p['scheduled']
        rec = p['received']
        var = rec - sched
        if sched > 0 and rec == 0:
            # Priority 1: Zero Arrival (100% Deficit) — largest scheduled volume first
            return (1, -sched, p['profile_number'])
        elif sched > 0 and rec < sched:
            # Priority 2: Partial Deficit — largest deficit first, then lowest fulfillment rate
            fulfillment_rate = rec / sched
            return (2, var, fulfillment_rate, -sched, p['profile_number'])
        elif sched > 0 and rec == sched:
            # Priority 3: Fully Fulfilled (100%) — largest scheduled volume first
            return (3, -sched, p['profile_number'])
        elif sched > 0 and rec > sched:
            # Priority 4: Over Schedule (+N) — largest excess volume first
            return (4, -var, p['profile_number'])
        else:
            # Priority 5: Unscheduled Arrival — largest received volume first
            return (5, -rec, p['profile_number'])

    combined_profiles.sort(key=get_variance_sort_key)

    wb = openpyxl.Workbook()
    
    # --- STYLING TOKENS ---
    dark_red_fill = PatternFill(start_color="970000", end_color="970000", fill_type="solid")
    slate_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    card_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    card_green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    card_amber_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    card_red_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    red_badge_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_badge_font = Font(name="Calibri", size=11, bold=True, color="991B1B")

    orange_badge_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    orange_badge_font = Font(name="Calibri", size=11, bold=True, color="9A3412")

    yellow_badge_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    yellow_badge_font = Font(name="Calibri", size=11, bold=True, color="92400E")

    green_badge_fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
    green_badge_font = Font(name="Calibri", size=11, bold=True, color="065F46")

    blue_badge_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
    blue_badge_font = Font(name="Calibri", size=11, bold=True, color="1E40AF")

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Calibri", size=10, bold=True, color="475569")
    meta_val_font = Font(name="Calibri", size=10, color="0F172A")

    kpi_lbl_font = Font(name="Calibri", size=9, bold=True, color="475569")
    kpi_val_font = Font(name="Calibri", size=16, bold=True, color="0F172A")

    tbl_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_side = Side(border_style="thin", color="CBD5E1")
    double_bottom_side = Side(border_style="double", color="0F172A")
    thick_top_side = Side(border_style="thin", color="0F172A")

    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    summary_border = Border(top=thick_top_side, bottom=double_bottom_side, left=thin_side, right=thin_side)

    # ----------------------------------------------------
    # TAB 1: DETAILED INBOUND LOGS
    # ----------------------------------------------------
    ws1 = wb.active
    ws1.title = "Inbound Truck Logs"
    ws1.views.sheetView[0].showGridLines = True

    ws1.merge_cells("A1:J1")
    ws1["A1"] = "CLEAN HARBORS - BUTTONWILLOW FACILITY"
    ws1["A1"].font = title_font
    ws1["A1"].fill = dark_red_fill
    ws1["A1"].alignment = align_center
    ws1.row_dimensions[1].height = 30

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"DAILY INBOUND TRUCK LOG REPORT — {selected_date}"
    ws1["A2"].font = subtitle_font
    ws1["A2"].fill = slate_fill
    ws1["A2"].alignment = align_center
    ws1.row_dimensions[2].height = 22

    ws1["A4"] = "Report Scope:"
    ws1["A4"].font = meta_label_font
    ws1["B4"] = f"Inbound Receiving & Chemist Logs for {selected_date}"
    ws1["B4"].font = meta_val_font

    ws1["F4"] = "Generated Date:"
    ws1["F4"].font = meta_label_font
    ws1["G4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws1["G4"].font = meta_val_font

    ws1["A5"] = "Data Source:"
    ws1["A5"].font = meta_label_font
    ws1["B5"] = "Truck Log App Master Database"
    ws1["B5"].font = meta_val_font

    def build_kpi_card(ws, col_start, col_end, label, formula_or_val, num_fmt, fill_color):
        ws.merge_cells(f"{col_start}7:{col_end}7")
        ws.merge_cells(f"{col_start}8:{col_end}8")
        ws[f"{col_start}7"] = label
        ws[f"{col_start}7"].font = kpi_lbl_font
        ws[f"{col_start}7"].fill = fill_color
        ws[f"{col_start}7"].alignment = align_center
        ws[f"{col_start}8"] = formula_or_val
        ws[f"{col_start}8"].font = kpi_val_font
        ws[f"{col_start}8"].fill = fill_color
        ws[f"{col_start}8"].alignment = align_center
        ws[f"{col_start}8"].number_format = num_fmt
        for row in range(7, 9):
            for col in range(openpyxl.utils.column_index_from_string(col_start), openpyxl.utils.column_index_from_string(col_end) + 1):
                ws.cell(row=row, column=col).border = cell_border

    end_row1 = 9 + (len(trucks) if trucks else 1)
    total_scheduled_cnt = sum(sched_map.values()) if sched_map else 0

    u35_trucks = [
        t for t in trucks 
        if '35' in str(t.get('cell_location', '') or '').strip().upper() 
        or (not t.get('cell_location') and '35' in str(t.get('win_code', '') or ''))
    ]
    u35_voc_x_tons = 0.0
    u35_total_tons = 0.0
    for t in u35_trucks:
        voc_val = float(t.get('measured_voc') if t.get('measured_voc') is not None else (t.get('voc_percentage') or 0.0))
        gross = float(t.get('gross_weight') or 0)
        tare = float(t.get('exit_weight') or 0)
        net_val = float(t.get('net_weight') or 0)
        if net_val > 500:
            tons = net_val / 2000.0
        elif net_val == 0 and gross > 0:
            tons = (gross - tare) / 2000.0
        elif net_val > 0:
            tons = net_val
        else:
            tons = gross / 2000.0 if gross > 0 else 0.0
        
        if tons > 0:
            u35_voc_x_tons += (voc_val * tons)
            u35_total_tons += tons

    u35_weighted_voc = (u35_voc_x_tons / u35_total_tons) if u35_total_tons > 0 else 0.0

    build_kpi_card(ws1, "A", "B", "TOTAL SCHEDULED LOADS", total_scheduled_cnt, "#,##0", card_blue_fill)
    build_kpi_card(ws1, "C", "D", "TOTAL INBOUND LOADS", f"=COUNTA(A10:A{end_row1})" if trucks else 0, "#,##0", card_green_fill)
    build_kpi_card(ws1, "E", "F", "UNIT 35 AVG VOC (PPM)", round(u35_weighted_voc, 2), "#,##0.0", card_amber_fill)
    build_kpi_card(ws1, "G", "H", "UNIT 35 NET TONNAGE", round(u35_total_tons, 2), "#,##0.00", card_red_fill)

    ws1.row_dimensions[7].height = 18
    ws1.row_dimensions[8].height = 26

    headers1 = [
        ("Ticket / Load #", align_center),
        ("Manifest #", align_center),
        ("Profile #", align_center),
        ("Generator Name", align_left),
        ("Net Weight (Lbs)", align_right),
        ("Net Weight (Tons)", align_right),
        ("VOC (PPM)", align_center),
        ("SG", align_center),
        ("Cell Unit", align_center),
        ("Grid", align_center)
    ]

    ws1.row_dimensions[9].height = 26
    for col_idx, (hdr_text, alignment) in enumerate(headers1, start=1):
        cell = ws1.cell(row=9, column=col_idx, value=hdr_text)
        cell.font = tbl_hdr_font
        cell.fill = dark_red_fill
        cell.alignment = alignment
        cell.border = cell_border

    if trucks:
        for idx, t in enumerate(trucks):
            row_num = 10 + idx
            ws1.row_dimensions[row_num].height = 20
            
            lbs = float(t.get('gross_weight') or 0)
            voc = t.get('measured_voc')
            voc_val = float(voc) if voc is not None else None
            
            c_tick = ws1.cell(row=row_num, column=1, value=t.get('truck_id') or t.get('load_number') or 'N/A')
            c_tick.alignment = align_center
            c_tick.font = bold_font
            c_tick.border = cell_border

            c_man = ws1.cell(row=row_num, column=2, value=t.get('manifest_number') or '---')
            c_man.alignment = align_center
            c_man.font = bold_font
            c_man.border = cell_border

            c_prof = ws1.cell(row=row_num, column=3, value=t.get('profile_number') or '---')
            c_prof.alignment = align_center
            c_prof.font = bold_font
            c_prof.border = cell_border

            c_gen = ws1.cell(row=row_num, column=4, value=t.get('generator') or '---')
            c_gen.alignment = align_left
            c_gen.font = regular_font
            c_gen.border = cell_border

            c_lbs = ws1.cell(row=row_num, column=5, value=lbs)
            c_lbs.alignment = align_right
            c_lbs.font = regular_font
            c_lbs.number_format = '#,##0'
            c_lbs.border = cell_border

            c_tons = ws1.cell(row=row_num, column=6, value=f"=E{row_num}/2000")
            c_tons.alignment = align_right
            c_tons.font = bold_font
            c_tons.number_format = '#,##0.00'
            c_tons.border = cell_border

            c_voc = ws1.cell(row=row_num, column=7, value=voc_val if voc_val is not None else 0)
            c_voc.alignment = align_center
            c_voc.number_format = '#,##0'
            c_voc.border = cell_border
            if voc_val is not None and voc_val >= 50.0:
                c_voc.fill = red_badge_fill
                c_voc.font = red_badge_font
            else:
                c_voc.font = regular_font

            c_sg = ws1.cell(row=row_num, column=8, value=t.get('specific_gravity') or '-')
            c_sg.alignment = align_center
            c_sg.font = regular_font
            c_sg.border = cell_border

            c_cell = ws1.cell(row=row_num, column=9, value=t.get('cell_location') or '-')
            c_cell.alignment = align_center
            c_cell.font = regular_font
            c_cell.border = cell_border

            c_grid = ws1.cell(row=row_num, column=10, value=t.get('grid_location') or '-')
            c_grid.alignment = align_center
            c_grid.font = regular_font
            c_grid.border = cell_border

        tot_row1 = end_row1 + 1
        ws1.row_dimensions[tot_row1].height = 24
        c_tot_lbl1 = ws1.cell(row=tot_row1, column=4, value="TOTAL DAILY INBOUNDS")
        c_tot_lbl1.font = bold_font
        c_tot_lbl1.alignment = align_right
        c_tot_lbl1.border = summary_border

        c_tot_lbs = ws1.cell(row=tot_row1, column=5, value=f"=SUM(E10:E{end_row1})")
        c_tot_lbs.font = bold_font
        c_tot_lbs.alignment = align_right
        c_tot_lbs.number_format = '#,##0'
        c_tot_lbs.border = summary_border

        c_tot_tons = ws1.cell(row=tot_row1, column=6, value=f"=SUM(F10:F{end_row1})")
        c_tot_tons.font = bold_font
        c_tot_tons.alignment = align_right
        c_tot_tons.number_format = '#,##0.00'
        c_tot_tons.border = summary_border

        for col in [1, 2, 3, 7, 8, 9, 10]:
            ws1.cell(row=tot_row1, column=col).border = summary_border

    # ----------------------------------------------------
    # TAB 2: PROFILE BREAKDOWN & SCHEDULE VARIANCE
    # ----------------------------------------------------
    ws2 = wb.create_sheet(title="Profile Breakdown & Variance")
    ws2.views.sheetView[0].showGridLines = True

    ws2.merge_cells("A1:K1")
    ws2["A1"] = "CLEAN HARBORS - BUTTONWILLOW FACILITY"
    ws2["A1"].font = title_font
    ws2["A1"].fill = dark_red_fill
    ws2["A1"].alignment = align_center
    ws2.row_dimensions[1].height = 30

    ws2.merge_cells("A2:K2")
    ws2["A2"] = f"PROFILE SUMMARY & SCHEDULED VARIANCE TRACKER — {selected_date}"
    ws2["A2"].font = subtitle_font
    ws2["A2"].fill = slate_fill
    ws2["A2"].alignment = align_center
    ws2.row_dimensions[2].height = 22

    headers2 = [
        ("Rank", align_center),
        ("Profile #", align_center),
        ("Generator Name", align_left),
        ("WIN Code", align_center),
        ("Scheduled Loads", align_right),
        ("Received Loads", align_right),
        ("Variance (+/-)", align_right),
        ("Fulfillment Rate", align_right),
        ("Total Net Tonnage", align_right),
        ("Average VOC PPM", align_center),
        ("Variance Status", align_center)
    ]

    ws2.row_dimensions[4].height = 26
    for col_idx, (hdr_text, alignment) in enumerate(headers2, start=1):
        cell = ws2.cell(row=4, column=col_idx, value=hdr_text)
        cell.font = tbl_hdr_font
        cell.fill = dark_red_fill
        cell.alignment = alignment
        cell.border = cell_border

    if combined_profiles:
        for idx, p in enumerate(combined_profiles):
            row_num = 5 + idx
            ws2.row_dimensions[row_num].height = 20
            
            sched_cnt = p['scheduled']
            rec_cnt = p['received']
            lbs = p['total_lbs']
            tons = lbs / 2000.0
            avg_voc = p['avg_voc']
            
            formula_variance = f"=F{row_num}-E{row_num}"
            formula_fulfillment = f"=IF(E{row_num}>0, F{row_num}/E{row_num}, 0)"

            if sched_cnt > 0 and rec_cnt == 0:
                status_str = "Zero Arrival (100% Deficit)"
                badge_fill = red_badge_fill
                badge_font = red_badge_font
            elif sched_cnt > 0 and rec_cnt < sched_cnt:
                pct = round((rec_cnt / sched_cnt) * 100, 1)
                status_str = f"Partial Deficit ({pct}%)"
                badge_fill = orange_badge_fill
                badge_font = orange_badge_font
            elif sched_cnt > 0 and rec_cnt == sched_cnt:
                status_str = "Fully Fulfilled (100%)"
                badge_fill = green_badge_fill
                badge_font = green_badge_font
            elif rec_cnt > sched_cnt and sched_cnt > 0:
                diff = rec_cnt - sched_cnt
                status_str = f"Over Schedule (+{diff})"
                badge_fill = blue_badge_fill
                badge_font = blue_badge_font
            else:
                status_str = "Unscheduled Arrival"
                badge_fill = yellow_badge_fill
                badge_font = yellow_badge_font

            ws2.cell(row=row_num, column=1, value=idx+1).alignment = align_center
            ws2.cell(row=row_num, column=1).font = bold_font
            ws2.cell(row=row_num, column=1).border = cell_border

            ws2.cell(row=row_num, column=2, value=p['profile_number'] or '---').alignment = align_center
            ws2.cell(row=row_num, column=2).font = bold_font
            ws2.cell(row=row_num, column=2).border = cell_border

            ws2.cell(row=row_num, column=3, value=p['generator'] or '---').alignment = align_left
            ws2.cell(row=row_num, column=3).font = regular_font
            ws2.cell(row=row_num, column=3).border = cell_border

            ws2.cell(row=row_num, column=4, value=p['win_code'] or '---').alignment = align_center
            ws2.cell(row=row_num, column=4).font = regular_font
            ws2.cell(row=row_num, column=4).border = cell_border

            c_sched = ws2.cell(row=row_num, column=5, value=sched_cnt)
            c_sched.alignment = align_right
            c_sched.font = regular_font
            c_sched.number_format = '#,##0'
            c_sched.border = cell_border

            c_rec = ws2.cell(row=row_num, column=6, value=rec_cnt)
            c_rec.alignment = align_right
            c_rec.font = regular_font
            c_rec.number_format = '#,##0'
            c_rec.border = cell_border

            c_var = ws2.cell(row=row_num, column=7, value=formula_variance)
            c_var.alignment = align_right
            c_var.font = bold_font
            c_var.number_format = '+#,##0;-#,##0;0'
            c_var.border = cell_border

            c_ful = ws2.cell(row=row_num, column=8, value=formula_fulfillment)
            c_ful.alignment = align_right
            c_ful.font = regular_font
            c_ful.number_format = '0.0%'
            c_ful.border = cell_border

            c_ton = ws2.cell(row=row_num, column=9, value=tons)
            c_ton.alignment = align_right
            c_ton.font = bold_font
            c_ton.number_format = '#,##0.00'
            c_ton.border = cell_border

            c_avoc = ws2.cell(row=row_num, column=10, value=avg_voc)
            c_avoc.alignment = align_center
            c_avoc.font = regular_font
            c_avoc.number_format = '#,##0.0'
            c_avoc.border = cell_border

            c_stat = ws2.cell(row=row_num, column=11, value=status_str)
            c_stat.alignment = align_center
            c_stat.font = badge_font
            c_stat.fill = badge_fill
            c_stat.border = cell_border

        tot_row2 = 4 + len(combined_profiles) + 1
        ws2.row_dimensions[tot_row2].height = 24
        c_tot_lbl2 = ws2.cell(row=tot_row2, column=4, value="TOTAL PROFILE SUMMARY")
        c_tot_lbl2.font = bold_font
        c_tot_lbl2.alignment = align_right
        c_tot_lbl2.border = summary_border

        c_tot_sched2 = ws2.cell(row=tot_row2, column=5, value=f"=SUM(E5:E{tot_row2-1})")
        c_tot_sched2.font = bold_font
        c_tot_sched2.alignment = align_right
        c_tot_sched2.number_format = '#,##0'
        c_tot_sched2.border = summary_border

        c_tot_rec2 = ws2.cell(row=tot_row2, column=6, value=f"=SUM(F5:F{tot_row2-1})")
        c_tot_rec2.font = bold_font
        c_tot_rec2.alignment = align_right
        c_tot_rec2.number_format = '#,##0'
        c_tot_rec2.border = summary_border

        c_tot_var2 = ws2.cell(row=tot_row2, column=7, value=f"=SUM(G5:G{tot_row2-1})")
        c_tot_var2.font = bold_font
        c_tot_var2.alignment = align_right
        c_tot_var2.number_format = '+#,##0;-#,##0;0'
        c_tot_var2.border = summary_border

        c_tot_ful2 = ws2.cell(row=tot_row2, column=8, value=f"=IF(E{tot_row2}>0, F{tot_row2}/E{tot_row2}, 0)")
        c_tot_ful2.font = bold_font
        c_tot_ful2.alignment = align_right
        c_tot_ful2.number_format = '0.0%'
        c_tot_ful2.border = summary_border

        c_tot_ton2 = ws2.cell(row=tot_row2, column=9, value=f"=SUM(I5:I{tot_row2-1})")
        c_tot_ton2.font = bold_font
        c_tot_ton2.alignment = align_right
        c_tot_ton2.number_format = '#,##0.00'
        c_tot_ton2.border = summary_border

        for col in [1, 2, 3, 10, 11]:
            ws2.cell(row=tot_row2, column=col).border = summary_border

    # Auto-fit column widths for both sheets
    for sheet in [ws1, ws2]:
        for col in sheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.coordinate in sheet.merged_cells:
                    continue
                if len(val) > max_len:
                    max_len = len(val)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Daily_Operations_Report_{selected_date}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

def is_valid_profile_entry(prof):
    import re
    if not prof: return False
    p = str(prof).strip().upper()
    if p in ['TOTAL', 'TOTALS', 'APPROVAL #', 'APPROVAL#', 'NONE', 'N/A', '']:
        return False
    if any(k in p for k in ['TONNAGE', 'FOAM', 'INFO', 'REPORT', 'RECEIVED', 'PPM', 'MANIFEST', 'WEIGHT', 'POUNDS', 'TONS', 'AVERAGE']):
        return False
    if re.match(r'^\d+(\.\d+)?$', p):
        return False
    if len(p) < 4:
        return False
    return True

@reports_bp.route('/voc_audit')
def voc_audit():
    import openpyxl, glob, os, re, difflib

    selected_date = request.args.get('date', '')
    filter_typos_only = request.args.get('typos_only', 'false').lower() == 'true'

    with closing(get_db_connection()) as conn:
        master_rows = conn.execute('SELECT profile_number, generator, win_code FROM profiles').fetchall()
        master_profiles = {str(r['profile_number']).strip().upper(): dict(r) for r in master_rows if r['profile_number']}
        master_list = list(master_profiles.keys())

        sched_query = 'SELECT schedule_date, profile_number, load_count FROM daily_schedule'
        params = ()
        if selected_date:
            sched_query += ' WHERE schedule_date = ?'
            params = (selected_date,)
            
        sched_rows = conn.execute(sched_query, params).fetchall()
        schedule_by_date = {}
        for r in sched_rows:
            sdate = str(r['schedule_date']).strip()
            prof = str(r['profile_number'] or '').strip().upper()
            if not sdate or not prof: continue
            if sdate not in schedule_by_date:
                schedule_by_date[sdate] = {}
            schedule_by_date[sdate][prof] = schedule_by_date[sdate].get(prof, 0) + (r['load_count'] or 1)

    voc_folder = r'I:\Buttonwillow\VOCs and TONNAGE  by YEAR\2026VOC'
    if os.path.exists(voc_folder):
        voc_files = glob.glob(os.path.join(voc_folder, '*VOC.xlsx'))
    else:
        voc_files = []

    audit_records = []
    total_files = len(voc_files)
    total_typos = 0

    for fpath in sorted(voc_files, reverse=True):
        fname = os.path.basename(fpath)
        match = re.match(r'(\d{2})(\d{2})(\d{4})VOC\.xlsx', fname, re.IGNORECASE)
        if not match: continue
        mm, dd, yyyy = match.groups()
        iso_date = f"{yyyy}-{mm}-{dd}"

        if selected_date and iso_date != selected_date:
            continue

        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        except Exception:
            continue

        file_actual_counts = {}

        for sheet_name in ['35', 'NON-VOC']:
            if sheet_name not in wb.sheetnames: continue
            ws = wb[sheet_name]
            
            header_found = False
            approval_col_idx = None
            
            for row in ws.iter_rows(values_only=True):
                if not row: continue
                row_str = [str(c or '').strip().upper() for c in row]
                if 'APPROVAL #' in row_str or 'APPROVAL#' in row_str:
                    header_found = True
                    approval_col_idx = row_str.index('APPROVAL #') if 'APPROVAL #' in row_str else row_str.index('APPROVAL#')
                    continue
                
                if header_found and approval_col_idx is not None and len(row) > approval_col_idx:
                    raw_prof = row[approval_col_idx]
                    if not raw_prof or str(raw_prof).strip().upper() in ['TOTAL', 'APPROVAL #', 'NONE', '']:
                        continue
                    
                    prof_clean = str(raw_prof).strip().upper()
                    if is_valid_profile_entry(prof_clean):
                        file_actual_counts[prof_clean] = file_actual_counts.get(prof_clean, 0) + 1

        wb.close()
        scheduled_for_date = schedule_by_date.get(iso_date, {})

        for prof, actual_cnt in file_actual_counts.items():
            if not is_valid_profile_entry(prof):
                continue

            if prof in master_profiles:
                is_exact = True
                suggested = prof
                score = 100.0
                gen_name = master_profiles[prof].get('generator', '---')
            else:
                is_exact = False
                matches = difflib.get_close_matches(prof, master_list, n=1, cutoff=0.55)
                if matches:
                    suggested = matches[0]
                    score = round(difflib.SequenceMatcher(None, prof, suggested).ratio() * 100, 1)
                    gen_name = master_profiles[suggested].get('generator', '---')
                else:
                    suggested = None
                    score = 0.0
                    gen_name = '---'

            if not is_exact:
                total_typos += 1

            if filter_typos_only and is_exact:
                continue

            sched_cnt = scheduled_for_date.get(prof, 0) or (scheduled_for_date.get(suggested, 0) if suggested else 0)

            audit_records.append({
                'date': iso_date,
                'file_name': fname,
                'file_profile': prof,
                'suggested_master': suggested or 'UNKNOWN',
                'is_exact': is_exact,
                'score': score,
                'generator': gen_name,
                'actual_count': actual_cnt,
                'scheduled_count': sched_cnt,
                'variance': actual_cnt - sched_cnt
            })

    return render_template(
        'voc_audit.html',
        records=audit_records,
        selected_date=selected_date,
        filter_typos_only=filter_typos_only,
        total_files=total_files,
        total_typos=total_typos
    )

@reports_bp.route('/reports')
def reports():
    today_str = date.today().isoformat()
    selected_date = request.args.get('date', today_str)
    
    with closing(get_db_connection()) as conn:
        all_profiles = conn.execute('SELECT profile_number, voc_percentage FROM profiles').fetchall()
        voc_dict = {str(p['profile_number'] or '').strip().upper(): p['voc_percentage'] for p in all_profiles}
        
        sched_data = conn.execute("SELECT profile_number, SUM(load_count) as sched_loads FROM daily_schedule WHERE schedule_date = ? GROUP BY profile_number", (selected_date,)).fetchall()
        schedule_dict = {}
        for row in sched_data:
            prof_clean = str(row['profile_number'] or '').strip().upper()
            schedule_dict[prof_clean] = schedule_dict.get(prof_clean, 0) + row['sched_loads']
        total_scheduled = sum(schedule_dict.values())
        
        # Pulls truck counts and exact tonnage sent to Unit 35
        actual_data = conn.execute('''
            SELECT 
                TRIM(UPPER(t.profile_number)) as profile_number, 
                COUNT(t.id) as total_trucks, 
                SUM(CASE WHEN t.cell_location LIKE '35%' THEN (CASE WHEN t.net_weight > 500 THEN t.net_weight / 2000.0 ELSE t.net_weight END) ELSE 0 END) as total_tons, 
                MAX(p.voc_percentage) as profile_voc 
            FROM truck_logs t 
            LEFT JOIN profiles p ON TRIM(UPPER(t.profile_number)) = TRIM(UPPER(p.profile_number)) 
            WHERE t.test_status != 'REJECTED' AND t.date_received = ? 
            GROUP BY TRIM(UPPER(t.profile_number))
        ''', (selected_date,)).fetchall()
        
        # Pulls logs details, joining case-insensitively
        truck_logs_raw = conn.execute('''
            SELECT t.*, p.generator
            FROM truck_logs t
            LEFT JOIN profiles p ON TRIM(UPPER(t.profile_number)) = TRIM(UPPER(p.profile_number))
            WHERE t.date_received = ? AND t.test_status != 'REJECTED'
            ORDER BY CAST(t.load_number AS INTEGER) ASC
        ''', (selected_date,)).fetchall()
        
        truck_logs = []
        for row in truck_logs_raw:
            t = dict(row)
            gross = float(t.get('gross_weight') or 0)
            tare = float(t.get('exit_weight') or 0)
            net_val = float(t.get('net_weight') or 0)
            
            # Ensure net_weight is always in Tons
            if net_val > 500:
                t['net_weight'] = net_val / 2000.0
            elif net_val == 0 and gross > 0:
                t['net_weight'] = (gross - tare) / 2000.0
                
            truck_logs.append(t)
        
        # Check for matching 2026VOC file for selected_date to supplement variance tracking
        voc_actual_dict = {}
        try:
            dt_obj = datetime.strptime(selected_date, '%Y-%m-%d')
            file_pattern = f"{dt_obj.strftime('%m%d%Y')}VOC.xlsx"
            voc_folder = r'I:\Buttonwillow\VOCs and TONNAGE  by YEAR\2026VOC'
            target_fpath = os.path.join(voc_folder, file_pattern)
            if os.path.exists(target_fpath):
                wb = openpyxl.load_workbook(target_fpath, data_only=True, read_only=True)
                for sheet_name in ['35', 'NON-VOC']:
                    if sheet_name not in wb.sheetnames: continue
                    ws = wb[sheet_name]
                    header_found = False
                    approval_col_idx = None
                    for r_vals in ws.iter_rows(values_only=True):
                        if not r_vals: continue
                        r_str = [str(c or '').strip().upper() for c in r_vals]
                        if 'APPROVAL #' in r_str or 'APPROVAL#' in r_str:
                            header_found = True
                            approval_col_idx = r_str.index('APPROVAL #') if 'APPROVAL #' in r_str else r_str.index('APPROVAL#')
                            continue
                        if header_found and approval_col_idx is not None and len(r_vals) > approval_col_idx:
                            raw_prof = r_vals[approval_col_idx]
                            prof_clean = str(raw_prof or '').strip().upper()
                            if is_valid_profile_entry(prof_clean):
                                voc_actual_dict[prof_clean] = voc_actual_dict.get(prof_clean, 0) + 1
                wb.close()
        except Exception:
            pass

    master_report, grand_trucks, grand_tons = [], 0, 0.0
    
    # --- TONNAGE-WEIGHTED TRACKING VARIABLES ---
    voc_x_tons_sum = 0.0
    unit_35_total_tons = 0.0
    
    actual_profiles = set()
    for a in actual_data:
        prof = str(a['profile_number'] or '').strip().upper()
        if not is_valid_profile_entry(prof):
            continue

        actual_profiles.add(prof)
        sched = schedule_dict.get(prof, 0)
        
        # Use truck_logs count or fallback to VOC file count if truck_logs has 0
        actual_cnt = a['total_trucks'] or voc_actual_dict.get(prof, 0)

        try:
            safe_voc = float(a['profile_voc'])
        except (ValueError, TypeError):
            safe_voc = 0.0
            
        unit_35_tons = a['total_tons'] or 0.0
        
        master_report.append({
            'profile_number': prof, 
            'total_trucks': actual_cnt, 
            'total_tons': unit_35_tons, 
            'avg_voc_ppm': safe_voc, 
            'scheduled': sched, 
            'variance': actual_cnt - sched
        })
        
        grand_trucks += actual_cnt
        grand_tons += unit_35_tons
        
        if unit_35_tons > 0: 
            voc_x_tons_sum += (safe_voc * unit_35_tons) 
            unit_35_total_tons += unit_35_tons

    # Check profiles present in VOC file that were not in truck_logs
    for prof, voc_cnt in voc_actual_dict.items():
        if prof not in actual_profiles and is_valid_profile_entry(prof):
            actual_profiles.add(prof)
            sched = schedule_dict.get(prof, 0)
            try: safe_voc = float(voc_dict.get(prof, 0.0))
            except (ValueError, TypeError): safe_voc = 0.0

            master_report.append({
                'profile_number': prof,
                'total_trucks': voc_cnt,
                'total_tons': 0.0,
                'avg_voc_ppm': safe_voc,
                'scheduled': sched,
                'variance': voc_cnt - sched
            })
            grand_trucks += voc_cnt
            
    for prof, sched in schedule_dict.items():
        if prof not in actual_profiles and is_valid_profile_entry(prof): 
            try:
                safe_voc = float(voc_dict.get(prof, 0.0))
            except (ValueError, TypeError):
                safe_voc = 0.0
                
            master_report.append({
                'profile_number': prof, 
                'total_trucks': 0, 
                'total_tons': 0.0, 
                'avg_voc_ppm': safe_voc, 
                'scheduled': sched, 
                'variance': -sched
            })
            
    master_report.sort(key=lambda x: (x['scheduled'], x['total_trucks']), reverse=True)
            
    # Calculate Final Tonnage-Weighted Average
    final_avg_voc = (voc_x_tons_sum / unit_35_total_tons) if unit_35_total_tons > 0 else 0.0
    
    grand_totals = {
        'grand_trucks': grand_trucks, 
        'grand_tons': grand_tons, 
        'grand_avg_voc_ppm': final_avg_voc, 
        'grand_scheduled': total_scheduled
    }
    
    trucks_by_profile = defaultdict(list)
    for t in truck_logs:
        prof_clean = str(t['profile_number'] or '').strip().upper()
        trucks_by_profile[prof_clean].append(t)

    # ---------------------------------------------------------
    # GAP FILLING LOGIC FOR PRINTED COVER SHEET (SAFEGUARDED)
    # ---------------------------------------------------------
    filled_trucks = []
    if truck_logs:
        def get_load_int(t):
            try: return int(t['load_number'])
            except: return -1
            
        numbered_trucks = [t for t in truck_logs if get_load_int(t) > 0]
        
        if numbered_trucks:
            truck_dict = {get_load_int(t): dict(t) for t in numbered_trucks}
            sorted_load_nums = sorted(truck_dict.keys())
            
            for idx, current_num in enumerate(sorted_load_nums):
                # 1. Add the actual truck
                filled_trucks.append(truck_dict[current_num])
                
                # 2. Check the gap to the NEXT truck
                if idx < len(sorted_load_nums) - 1:
                    next_num = sorted_load_nums[idx + 1]
                    gap = next_num - current_num - 1
                    
                    if 0 < gap <= 20:
                        # Safe to fill small missing gaps individually
                        for i in range(current_num + 1, next_num):
                            filled_trucks.append({
                                'load_number': str(i),
                                'manifest_number': '---',
                                'profile_number': 'VOID / NOT ISSUED',
                                'sales_order': '',
                                'generator': 'Awaiting Entry / Missing Ticket',
                                'gross_weight': None,
                                'exit_weight': None,
                                'net_weight': None,
                                'cell_location': '---',
                                'grid_location': '---',
                                'time_in': '',
                                'time_out': '',
                                'test_status': 'VOID'
                            })
                    elif gap > 20:
                        # DANGER: Huge gap detected (likely a test, typo, or late entry). 
                        # Print ONE summary row to prevent crashing the server.
                        filled_trucks.append({
                            'load_number': f"{current_num + 1} ➔ {next_num - 1}",
                            'manifest_number': '⚠️ LARGE GAP',
                            'profile_number': 'SYSTEM SAFEGUARD',
                            'sales_order': '',
                            'generator': 'Multiple load numbers skipped to prevent system lag.',
                            'gross_weight': None,
                            'exit_weight': None,
                            'net_weight': None,
                            'cell_location': '---',
                            'grid_location': '---',
                            'time_in': '',
                            'time_out': '',
                            'test_status': 'VOID'
                        })
                        
        # Add any letters/non-numbers (like "12A") at the very end just in case
        for t in truck_logs:
            if get_load_int(t) == -1:
                filled_trucks.append(dict(t))

    return render_template('reports.html', 
                           report_data=master_report, 
                           grand_totals=grand_totals, 
                           selected_date=selected_date, 
                           today_str=today_str, 
                           trucks_by_profile=trucks_by_profile, 
                           all_trucks=filled_trucks)


@reports_bp.route('/compliance')
def compliance():
    today = date.today()
    # Default date range: Jan 1st of current year to today
    start_date = date(today.year, 1, 1).isoformat()
    end_date = today.isoformat()
    return render_template('compliance.html', start_date=start_date, end_date=end_date)


def parse_voc_filename_date(fname):
    if fname.startswith('~$') or fname.startswith('.'):
        return None
    base = re.sub(r'\.+[xX][lL][sS][xX]?$', '', fname).strip().upper()
    base = re.sub(r'[\s_]*VOC$', '', base).strip()
    
    m_emb = re.search(r'(\d{1,2})[\-_/](\d{1,2})[\-_/](20\d{2})', base)
    if m_emb:
        m, d, yyyy = m_emb.groups()
        return f"{yyyy}-{int(m):02d}-{int(d):02d}"

    m_emb8 = re.search(r'(\d{2})(\d{2})(20\d{2})', base)
    if m_emb8:
        mm, dd, yyyy = m_emb8.groups()
        return f"{yyyy}-{mm}-{dd}"

    m2 = re.match(r'^(\d{1})(\d{2})(20\d{2})$', base)
    if m2:
        m, dd, yyyy = m2.groups()
        return f"{yyyy}-0{m}-{dd}"
        
    m3 = re.match(r'^(\d{2})(\d{2})(\d{2})$', base)
    if m3:
        mm, dd, yy = m3.groups()
        return f"20{yy}-{mm}-{dd}"

    return None

_VOC_FILE_CACHE = {}
_VOC_CACHE_TIME = None

def get_voc_file_counts(force_refresh=False):
    global _VOC_FILE_CACHE, _VOC_CACHE_TIME
    now = datetime.now()
    cache_path = os.path.join(os.path.dirname(__file__), 'voc_cache.json')
    
    if not force_refresh and _VOC_FILE_CACHE:
        return _VOC_FILE_CACHE
        
    if not force_refresh and os.path.exists(cache_path):
        try:
            with open(cache_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                _VOC_FILE_CACHE = data.get('counts_by_date', {})
                _VOC_CACHE_TIME = now
                if _VOC_FILE_CACHE:
                    return _VOC_FILE_CACHE
        except Exception:
            pass

    counts_by_date = {}
    voc_folder = r'I:\Buttonwillow\VOCs and TONNAGE  by YEAR\2026VOC'
    if os.path.exists(voc_folder):
        voc_files = glob.glob(os.path.join(voc_folder, '*'))
        for fpath in voc_files:
            fname = os.path.basename(fpath)
            if fname.startswith('~$') or not (fname.lower().endswith('.xlsx') or fname.lower().endswith('.xls')):
                continue
            iso_date = parse_voc_filename_date(fname)
            if not iso_date:
                continue
            try:
                wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
                cnt = 0
                for sname in ['35', 'NON-VOC']:
                    if sname not in wb.sheetnames: continue
                    ws = wb[sname]
                    h_found = False
                    app_idx = None
                    for r in ws.iter_rows(values_only=True):
                        if not r: continue
                        r_str = [str(cell or '').strip().upper() for cell in r]
                        if 'APPROVAL #' in r_str or 'APPROVAL#' in r_str:
                            h_found = True
                            app_idx = r_str.index('APPROVAL #') if 'APPROVAL #' in r_str else r_str.index('APPROVAL#')
                            continue
                        if h_found and app_idx is not None and len(r) > app_idx:
                            p = str(r[app_idx] or '').strip().upper()
                            if is_valid_profile_entry(p):
                                cnt += 1
                wb.close()
                if cnt > 0:
                    counts_by_date[iso_date] = cnt
            except Exception:
                pass

    _VOC_FILE_CACHE = counts_by_date
    _VOC_CACHE_TIME = now
    try:
        with open(cache_path, 'w', encoding='utf-8') as f:
            json.dump({'counts_by_date': _VOC_FILE_CACHE, 'last_updated': now.isoformat()}, f, indent=2)
    except Exception:
        pass
    return _VOC_FILE_CACHE

@reports_bp.route('/api/compliance/data')
def compliance_data():
    report_type = request.args.get('report_type', 'variance')
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    include_weekends = request.args.get('include_weekends', 'false').lower() == 'true'
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    
    if not start_str or not end_str:
        return jsonify({'error': 'Missing dates'}), 400
        
    try:
        start_dt = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
        
    # Generate all dates in range
    date_list = []
    curr = start_dt
    while curr <= end_dt:
        date_list.append(curr.isoformat())
        curr += timedelta(days=1)
        
    with closing(get_db_connection()) as conn:
        if report_type == 'variance':
            # 1. Scheduled
            sched_rows = conn.execute('''
                SELECT schedule_date, SUM(load_count) as scheduled
                FROM daily_schedule
                WHERE schedule_date BETWEEN ? AND ?
                GROUP BY schedule_date
            ''', (start_str, end_str)).fetchall()
            sched_map = {r['schedule_date']: r['scheduled'] for r in sched_rows}
            
            # 2. Actual
            actual_rows = conn.execute('''
                SELECT date_received, COUNT(id) as actual
                FROM truck_logs
                WHERE date_received BETWEEN ? AND ? AND test_status != 'REJECTED'
                GROUP BY date_received
            ''', (start_str, end_str)).fetchall()
            actual_map = {r['date_received']: r['actual'] for r in actual_rows}
            
            # Merge 2026VOC file counts — use the HIGHER of truck_logs vs VOC file
            voc_file_counts = get_voc_file_counts(force_refresh=force_refresh)
            for iso_date in date_list:
                voc_cnt = voc_file_counts.get(iso_date, 0)
                db_cnt = actual_map.get(iso_date, 0)
                if voc_cnt > db_cnt:
                    actual_map[iso_date] = voc_cnt
            
            # Known facility holidays (0/0 days that are NOT weekends)
            FACILITY_HOLIDAYS = {
                '2026-01-01',  # New Year's Day
                '2026-01-02',  # New Year's (observed)
                '2026-05-25',  # Memorial Day
                '2026-05-26',  # Memorial Day (observed)
                '2026-07-03',  # Independence Day (observed)
                '2026-07-04',  # Independence Day
                '2026-09-07',  # Labor Day
                '2026-11-26',  # Thanksgiving
                '2026-11-27',  # Day after Thanksgiving
                '2026-12-24',  # Christmas Eve
                '2026-12-25',  # Christmas Day
                '2026-12-31',  # New Year's Eve
            }
            exclude_holidays = request.args.get('exclude_holidays', 'false').lower() == 'true'
            
            raw_labels = date_list
            raw_scheduled = [sched_map.get(d, 0) for d in raw_labels]
            raw_actual = [actual_map.get(d, 0) for d in raw_labels]
            raw_variance = [raw_actual[i] - raw_scheduled[i] for i in range(len(raw_labels))]
            
            labels = []
            scheduled_data = []
            actual_data = []
            variance_data = []
            table_data = []

            for i, d in enumerate(raw_labels):
                s = raw_scheduled[i]
                a = raw_actual[i]
                v = raw_variance[i]
                
                # Exclude 0/0 weekend or non-operating days if include_weekends is false
                if not include_weekends and s == 0 and a == 0:
                    continue
                
                # Exclude holidays if toggle is on
                if exclude_holidays and d in FACILITY_HOLIDAYS:
                    continue
                    
                labels.append(d)
                scheduled_data.append(s)
                actual_data.append(a)
                variance_data.append(v)
                table_data.append({
                    'date': d, 'scheduled': s, 'actual': a, 'variance': f"{v:+d}"
                })

            # Summary Metrics
            total_sched = sum(raw_scheduled)
            total_act = sum(raw_actual)
            net_variance = total_act - total_sched
            
            avg_variance = round(net_variance / len(labels), 1) if labels else 0
            
            summary = {
                'kpi1_val': total_sched, 'kpi1_label': 'Total Scheduled',
                'kpi2_val': total_act, 'kpi2_label': 'Total Actual Weighed',
                'kpi3_val': f"{net_variance:+d}", 'kpi3_label': 'Net Variance',
                'kpi4_val': avg_variance, 'kpi4_label': 'Avg Daily Variance' + ('' if include_weekends else ' (Active Days)')
            }
            
            return jsonify({
                'labels': labels,
                'datasets': [
                    {'label': 'Scheduled Loads', 'data': scheduled_data},
                    {'label': 'Actual Loads', 'data': actual_data},
                    {'label': 'Variance', 'data': variance_data}
                ],
                'summary': summary,
                'table_data': table_data
            })
            
        elif report_type == 'traffic':
            traffic_rows = conn.execute('''
                SELECT date_received, COUNT(id) as traffic
                FROM truck_logs
                WHERE date_received BETWEEN ? AND ? AND test_status != 'REJECTED'
                GROUP BY date_received
            ''', (start_str, end_str)).fetchall()
            traffic_map = {r['date_received']: r['traffic'] for r in traffic_rows}
            
            labels = date_list
            counts = [traffic_map.get(d, 0) for d in labels]
            
            # Summary metrics
            total_trucks = sum(counts)
            avg_daily = round(total_trucks / len(labels), 1) if labels else 0
            max_daily = max(counts) if counts else 0
            
            # Find busiest day
            busiest_day = 'N/A'
            if counts:
                busiest_idx = counts.index(max_daily)
                busiest_day = labels[busiest_idx]
                
            summary = {
                'kpi1_val': total_trucks, 'kpi1_label': 'Total Trucks Weighed',
                'kpi2_val': avg_daily, 'kpi2_label': 'Avg Daily Trucks',
                'kpi3_val': max_daily, 'kpi3_label': 'Busiest Day Traffic',
                'kpi4_val': busiest_day, 'kpi4_label': 'Busiest Day Date'
            }
            
            table_data = [{'date': d, 'count': counts[i]} for i, d in enumerate(labels)]
            
            return jsonify({
                'labels': labels,
                'datasets': [
                    {'label': 'Total Checked-in Trucks', 'data': counts}
                ],
                'summary': summary,
                'table_data': table_data
            })
            
        elif report_type == 'tonnage_ytd':
            # Aggregation query
            rows = conn.execute('''
                SELECT date_received, cell_location, SUM(net_weight) as tons
                FROM truck_logs
                WHERE date_received BETWEEN ? AND ? AND exit_weight IS NOT NULL AND test_status != 'REJECTED'
                GROUP BY date_received, cell_location
                ORDER BY date_received ASC
            ''', (start_str, end_str)).fetchall()
            
            labels = date_list
            # Pre-populate daily bins
            daily_wmu35 = {d: 0.0 for d in labels}
            daily_wmu31 = {d: 0.0 for d in labels}
            daily_stu = {d: 0.0 for d in labels}
            daily_landfill = {d: 0.0 for d in labels}
            
            for r in rows:
                d = r['date_received']
                if d not in daily_wmu35:
                    continue
                wmu = str(r['cell_location'] or '').strip().upper()
                tons = float(r['tons'] or 0.0)
                
                is_stu = wmu.startswith('34') or wmu.startswith('STU') or 'BAY' in wmu or wmu in ['CCS', 'CCSF', 'CCSM']
                
                if wmu.startswith('35'):
                    daily_wmu35[d] += tons
                elif wmu.startswith('31'):
                    daily_wmu31[d] += tons
                elif is_stu:
                    daily_stu[d] += tons
                else:
                    daily_landfill[d] += tons
            
            # Compute running cumulative sums (YTD Curves)
            cum_wmu35, cum_wmu31, cum_stu, cum_landfill = [], [], [], []
            sum35, sum31, sum_s, sum_l = 0.0, 0.0, 0.0, 0.0
            
            for d in labels:
                sum35 += daily_wmu35[d]
                sum31 += daily_wmu31[d]
                sum_s += daily_stu[d]
                sum_l += daily_landfill[d]
                cum_wmu35.append(round(sum35, 2))
                cum_wmu31.append(round(sum31, 2))
                cum_stu.append(round(sum_s, 2))
                cum_landfill.append(round(sum_l, 2))
                
            grand_total = round(sum35 + sum31 + sum_s + sum_l, 2)
            
            summary = {
                'kpi1_val': f"{grand_total:,.2f} Tons", 'kpi1_label': 'Total YTD Tonnage',
                'kpi2_val': f"{sum35:,.2f} Tons", 'kpi2_label': 'WMU 35 Total',
                'kpi3_val': f"{sum31:,.2f} Tons", 'kpi3_label': 'WMU 31 Total',
                'kpi4_val': f"{sum_s:,.2f} Tons", 'kpi4_label': 'STU / Decon Total'
            }
            
            table_data = []
            for i, d in enumerate(labels):
                table_data.append({
                    'date': d,
                    'wmu35': f"{cum_wmu35[i]:,.2f}",
                    'wmu31': f"{cum_wmu31[i]:,.2f}",
                    'stu': f"{cum_stu[i]:,.2f}",
                    'landfill': f"{cum_landfill[i]:,.2f}",
                    'total': f"{(cum_wmu35[i] + cum_wmu31[i] + cum_stu[i] + cum_landfill[i]):,.2f}"
                })
                
            return jsonify({
                'labels': labels,
                'datasets': [
                    {'label': 'WMU 35 Cumulative', 'data': cum_wmu35},
                    {'label': 'WMU 31 Cumulative', 'data': cum_wmu31},
                    {'label': 'STU / Decon Cumulative', 'data': cum_stu},
                    {'label': 'Landfill / Other Cumulative', 'data': cum_landfill}
                ],
                'summary': summary,
                'table_data': table_data
            })
            
        elif report_type == 'timings':
            rows = conn.execute('''
                SELECT time_in, COUNT(id) as count
                FROM truck_logs
                WHERE date_received BETWEEN ? AND ? AND time_in IS NOT NULL AND time_in != ''
                GROUP BY time_in
            ''', (start_str, end_str)).fetchall()
            
            # Map hours 00 to 23
            hourly_counts = {f"{h:02d}": 0 for h in range(24)}
            for r in rows:
                time_str = r['time_in']
                try:
                    hour_part = time_str.split(':')[0].strip()
                    if hour_part.isdigit():
                        h_int = int(hour_part)
                        if 0 <= h_int < 24:
                            hourly_counts[f"{h_int:02d}"] += r['count']
                except:
                    pass
            
            # Select working hours range to show in chart for clean design (e.g. 06:00 to 18:00)
            hours_keys = [f"{h:02d}" for h in range(6, 19)]
            labels = [f"{h}:00" for h in hours_keys]
            counts = [hourly_counts[h] for h in hours_keys]
            
            total_weighed = sum(hourly_counts.values())
            
            # Find peak hour
            peak_val = max(counts) if counts else 0
            peak_hour = 'N/A'
            if peak_val > 0:
                peak_idx = counts.index(peak_val)
                peak_hour = labels[peak_idx]
                
            summary = {
                'kpi1_val': total_weighed, 'kpi1_label': 'Total Checked-in Trucks',
                'kpi2_val': peak_hour, 'kpi2_label': 'Busiest Hour (Peak)',
                'kpi3_val': peak_val, 'kpi3_label': 'Peak Hour Truck Volume',
                'kpi4_val': f"{round((peak_val/total_weighed*100) if total_weighed > 0 else 0)}%", 'kpi4_label': 'Peak Hour Traffic Ratio'
            }
            
            table_data = [{'hour': labels[i], 'count': counts[i]} for i, h in enumerate(labels)]
            
            return jsonify({
                'labels': labels,
                'datasets': [
                    {'label': 'Truck Entry Count', 'data': counts}
                ],
                'summary': summary,
                'table_data': table_data
            })
            
        elif report_type == 'norm':
            # Retrieve NORM loads (profiles with win_code = 'CNON')
            rows = conn.execute('''
                SELECT t.id, t.date_received, t.profile_number, t.load_number, t.manifest_number, 
                       t.net_weight, t.cell_location, p.generator
                FROM truck_logs t
                INNER JOIN profiles p ON t.profile_number = p.profile_number
                WHERE p.win_code = 'CNON'
                  AND t.date_received BETWEEN ? AND ?
                ORDER BY t.date_received ASC, CAST(t.load_number AS INTEGER) ASC
            ''', (start_str, end_str)).fetchall()

            labels = date_list
            daily_counts = {d: 0 for d in labels}
            total_tons = 0.0
            unique_profiles = set()
            
            table_data = []
            for r in rows:
                d = r['date_received']
                if d in daily_counts:
                    daily_counts[d] += 1
                unique_profiles.add(r['profile_number'])
                total_tons += float(r['net_weight'] or 0.0)
                
                table_data.append({
                    'date': d,
                    'profile_number': r['profile_number'],
                    'load_number': r['load_number'],
                    'manifest_number': r['manifest_number'] or '---',
                    'generator': r['generator'] or 'Unknown Generator',
                    'weight': f"{float(r['net_weight'] or 0.0):.2f}",
                    'cell_location': r['cell_location'] or '---'
                })

            counts = [daily_counts[d] for d in labels]
            total_norm_loads = len(rows)
            
            busiest_day = 'N/A'
            max_daily = max(counts) if counts else 0
            if max_daily > 0:
                busiest_idx = counts.index(max_daily)
                busiest_day = labels[busiest_idx]

            summary = {
                'kpi1_val': total_norm_loads, 'kpi1_label': 'Total NORM Loads',
                'kpi2_val': len(unique_profiles), 'kpi2_label': 'Unique Profiles',
                'kpi3_val': f"{total_tons:,.2f} Tons", 'kpi3_label': 'Total NORM Tonnage',
                'kpi4_val': busiest_day if max_daily > 0 else 'None', 'kpi4_label': 'Busiest NORM Day'
            }

            return jsonify({
                'labels': labels,
                'datasets': [
                    {'label': 'NORM Loads Checked-in', 'data': counts}
                ],
                'summary': summary,
                'table_data': table_data
            })
            
    return jsonify({'error': 'Invalid report type'}), 400


@reports_bp.route('/api/compliance/export_profile_variance_excel')
def export_profile_variance_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    limit = request.args.get('limit', 20, type=int)

    with closing(get_db_connection()) as conn:
        sched_rows = conn.execute("""
            SELECT 
                UPPER(TRIM(profile_number)) as profile_number,
                MAX(generator) as generator_name,
                SUM(COALESCE(load_count, 1)) as total_scheduled_loads
            FROM daily_schedule
            WHERE profile_number IS NOT NULL AND TRIM(profile_number) != ''
            GROUP BY UPPER(TRIM(profile_number))
        """).fetchall()

        tl_rows = conn.execute("""
            SELECT UPPER(TRIM(profile_number)) as profile_number, COUNT(*) as received_tl_count
            FROM truck_logs
            WHERE profile_number IS NOT NULL AND TRIM(profile_number) != ''
            GROUP BY UPPER(TRIM(profile_number))
        """).fetchall()

        voc_rows = conn.execute("""
            SELECT UPPER(TRIM(profile_number)) as profile_number, COUNT(*) as received_voc_count
            FROM voc_analyzer_logs
            WHERE profile_number IS NOT NULL AND TRIM(profile_number) != ''
            GROUP BY UPPER(TRIM(profile_number))
        """).fetchall()

    sched_dict = {r['profile_number']: dict(r) for r in sched_rows}
    tl_dict = {r['profile_number']: r['received_tl_count'] for r in tl_rows}
    voc_dict = {r['profile_number']: r['received_voc_count'] for r in voc_rows}

    combined = []
    for prof, sdata in sched_dict.items():
        s_cnt = sdata['total_scheduled_loads'] or 1
        t_cnt = tl_dict.get(prof, 0)
        v_cnt = voc_dict.get(prof, 0)
        r_cnt = max(t_cnt, v_cnt)
        var = r_cnt - s_cnt
        missing = s_cnt - r_cnt
        ful_pct = round((r_cnt / s_cnt * 100), 1) if s_cnt > 0 else 0.0
        combined.append({
            'profile_number': prof,
            'generator_name': sdata['generator_name'] or 'N/A',
            'scheduled': s_cnt,
            'received': r_cnt,
            'variance': var,
            'missing': missing,
            'fulfillment_pct': ful_pct
        })

    # Sort by largest missing loads (most negative variance)
    combined.sort(key=lambda x: (x['variance'], -x['scheduled']))
    top_profiles = combined[:limit]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profile Schedule Variance"
    ws.views.sheetView[0].showGridLines = True

    dark_red_fill = PatternFill(start_color="970000", end_color="970000", fill_type="solid")
    slate_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    card_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
    card_green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
    card_amber_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
    card_red_fill = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")

    red_badge_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    red_badge_font = Font(name="Calibri", size=11, bold=True, color="991B1B")

    orange_badge_fill = PatternFill(start_color="FFEDD5", end_color="FFEDD5", fill_type="solid")
    orange_badge_font = Font(name="Calibri", size=11, bold=True, color="9A3412")

    yellow_badge_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    yellow_badge_font = Font(name="Calibri", size=11, bold=True, color="92400E")

    title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
    subtitle_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_label_font = Font(name="Calibri", size=10, bold=True, color="475569")
    meta_val_font = Font(name="Calibri", size=10, color="0F172A")

    kpi_lbl_font = Font(name="Calibri", size=9, bold=True, color="475569")
    kpi_val_font = Font(name="Calibri", size=16, bold=True, color="0F172A")

    tbl_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    thin_side = Side(border_style="thin", color="CBD5E1")
    double_bottom_side = Side(border_style="double", color="0F172A")
    thick_top_side = Side(border_style="thin", color="0F172A")

    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    summary_border = Border(top=thick_top_side, bottom=double_bottom_side, left=thin_side, right=thin_side)

    ws.merge_cells("A1:H1")
    ws["A1"] = "CLEAN HARBORS - BUTTONWILLOW FACILITY"
    ws["A1"].font = title_font
    ws["A1"].fill = dark_red_fill
    ws["A1"].alignment = align_center
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = f"TOP {limit} PROFILES SCHEDULED VARIANCE TRACKER (SCHEDULED LOADS NOT ARRIVED)"
    ws["A2"].font = subtitle_font
    ws["A2"].fill = slate_fill
    ws["A2"].alignment = align_center
    ws.row_dimensions[2].height = 22

    ws["A4"] = "Report Scope:"
    ws["A4"].font = meta_label_font
    ws["B4"] = f"Top {limit} Profiles by Largest Unfulfilled Scheduled Volume"
    ws["B4"].font = meta_val_font

    ws["E4"] = "Generated Date:"
    ws["E4"].font = meta_label_font
    ws["F4"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["F4"].font = meta_val_font

    ws["A5"] = "Data Source:"
    ws["A5"].font = meta_label_font
    ws["B5"] = "Daily Schedule vs. Receiving Truck Logs & VOC Analyzer History"
    ws["B5"].font = meta_val_font

    def build_kpi_card(ws, col_start, col_end, label, formula_or_val, is_pct, fill_color):
        ws.merge_cells(f"{col_start}7:{col_end}7")
        ws.merge_cells(f"{col_start}8:{col_end}8")
        ws[f"{col_start}7"] = label
        ws[f"{col_start}7"].font = kpi_lbl_font
        ws[f"{col_start}7"].fill = fill_color
        ws[f"{col_start}7"].alignment = align_center
        ws[f"{col_start}8"] = formula_or_val
        ws[f"{col_start}8"].font = kpi_val_font
        ws[f"{col_start}8"].fill = fill_color
        ws[f"{col_start}8"].alignment = align_center
        if is_pct:
            ws[f"{col_start}8"].number_format = '0.0%'
        else:
            ws[f"{col_start}8"].number_format = '#,##0'
        for row in range(7, 9):
            for col in range(openpyxl.utils.column_index_from_string(col_start), openpyxl.utils.column_index_from_string(col_end) + 1):
                ws.cell(row=row, column=col).border = cell_border

    end_row_idx = 9 + len(top_profiles)
    build_kpi_card(ws, "A", "B", f"TOP {limit} SCHEDULED LOADS", f"=SUM(D10:D{end_row_idx})", False, card_blue_fill)
    build_kpi_card(ws, "C", "D", f"TOP {limit} RECEIVED LOADS", f"=SUM(E10:E{end_row_idx})", False, card_green_fill)
    build_kpi_card(ws, "E", "F", "TOTAL MISSING / UNFULFILLED LOADS", f"=SUM(F10:F{end_row_idx})", False, card_red_fill)
    build_kpi_card(ws, "G", "H", "OVERALL FULFILLMENT RATE", f"=E{end_row_idx+1}/D{end_row_idx+1}", True, card_amber_fill)

    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 26

    headers = [
        ("Rank", align_center),
        ("Profile #", align_center),
        ("Generator Name", align_left),
        ("Scheduled Loads", align_right),
        ("Received Loads", align_right),
        ("Missing Loads", align_right),
        ("Fulfillment Rate", align_right),
        ("Variance Status", align_center)
    ]

    ws.row_dimensions[9].height = 26
    for col_idx, (hdr_text, alignment) in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=hdr_text)
        cell.font = tbl_hdr_font
        cell.fill = dark_red_fill
        cell.alignment = alignment
        cell.border = cell_border

    for idx, rdata in enumerate(top_profiles):
        row_num = 10 + idx
        ws.row_dimensions[row_num].height = 20
        
        prof_no = rdata['profile_number']
        gen_name = rdata['generator_name']
        sched_loads = rdata['scheduled']
        rec_loads = rdata['received']
        
        formula_missing = f"=D{row_num}-E{row_num}"
        formula_fulfillment = f"=IF(D{row_num}>0, E{row_num}/D{row_num}, 0)"
        
        if rec_loads == 0:
            status_str = "Zero Arrival (100% Deficit)"
            badge_fill = red_badge_fill
            badge_font = red_badge_font
        elif rdata['fulfillment_pct'] < 10.0:
            status_str = "Critical Deficit (<10% Fulfilled)"
            badge_fill = orange_badge_fill
            badge_font = orange_badge_font
        else:
            status_str = f"Partial Deficit ({rdata['fulfillment_pct']}%)"
            badge_fill = yellow_badge_fill
            badge_font = yellow_badge_font

        c_rank = ws.cell(row=row_num, column=1, value=idx+1)
        c_rank.alignment = align_center
        c_rank.font = bold_font
        c_rank.border = cell_border

        c_prof = ws.cell(row=row_num, column=2, value=prof_no)
        c_prof.alignment = align_center
        c_prof.font = bold_font
        c_prof.border = cell_border

        c_gen = ws.cell(row=row_num, column=3, value=gen_name)
        c_gen.alignment = align_left
        c_gen.font = regular_font
        c_gen.border = cell_border

        c_sched = ws.cell(row=row_num, column=4, value=sched_loads)
        c_sched.alignment = align_right
        c_sched.font = regular_font
        c_sched.number_format = '#,##0'
        c_sched.border = cell_border

        c_rec = ws.cell(row=row_num, column=5, value=rec_loads)
        c_rec.alignment = align_right
        c_rec.font = regular_font
        c_rec.number_format = '#,##0'
        c_rec.border = cell_border

        c_miss = ws.cell(row=row_num, column=6, value=formula_missing)
        c_miss.alignment = align_right
        c_miss.font = bold_font
        c_miss.number_format = '#,##0'
        c_miss.border = cell_border

        c_ful = ws.cell(row=row_num, column=7, value=formula_fulfillment)
        c_ful.alignment = align_right
        c_ful.font = regular_font
        c_ful.number_format = '0.0%'
        c_ful.border = cell_border

        c_stat = ws.cell(row=row_num, column=8, value=status_str)
        c_stat.alignment = align_center
        c_stat.font = badge_font
        c_stat.fill = badge_fill
        c_stat.border = cell_border

    tot_row = end_row_idx + 1
    ws.row_dimensions[tot_row].height = 24
    c_tot_lbl = ws.cell(row=tot_row, column=3, value=f"TOP {limit} TOTAL / AVERAGE")
    c_tot_lbl.font = bold_font
    c_tot_lbl.alignment = align_right
    c_tot_lbl.border = summary_border

    c_tot_sched = ws.cell(row=tot_row, column=4, value=f"=SUM(D10:D{end_row_idx})")
    c_tot_sched.font = bold_font
    c_tot_sched.alignment = align_right
    c_tot_sched.number_format = '#,##0'
    c_tot_sched.border = summary_border

    c_tot_rec = ws.cell(row=tot_row, column=5, value=f"=SUM(E10:E{end_row_idx})")
    c_tot_rec.font = bold_font
    c_tot_rec.alignment = align_right
    c_tot_rec.number_format = '#,##0'
    c_tot_rec.border = summary_border

    c_tot_miss = ws.cell(row=tot_row, column=6, value=f"=SUM(F10:F{end_row_idx})")
    c_tot_miss.font = bold_font
    c_tot_miss.alignment = align_right
    c_tot_miss.number_format = '#,##0'
    c_tot_miss.border = summary_border

    c_tot_ful = ws.cell(row=tot_row, column=7, value=f"=E{tot_row}/D{tot_row}")
    c_tot_ful.font = bold_font
    c_tot_ful.alignment = align_right
    c_tot_ful.number_format = '0.0%'
    c_tot_ful.border = summary_border

    for col in [1, 2, 8]:
        ws.cell(row=tot_row, column=col).border = summary_border

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 46
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 30

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"Profile_Schedule_Variance_Top{limit}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@reports_bp.route('/api/compliance/export_excel')
def compliance_export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    report_type = request.args.get('report_type', 'variance')
    start_str = request.args.get('start_date')
    end_str = request.args.get('end_date')
    include_weekends = request.args.get('include_weekends', 'false').lower() == 'true'
    exclude_holidays = request.args.get('exclude_holidays', 'false').lower() == 'true'
    force_refresh = request.args.get('refresh', 'false').lower() == 'true'
    
    if not start_str or not end_str:
        return jsonify({'error': 'Missing dates'}), 400
        
    try:
        start_dt = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_dt = datetime.strptime(end_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Invalid date format'}), 400
        
    date_list = []
    curr = start_dt
    while curr <= end_dt:
        date_list.append(curr.isoformat())
        curr += timedelta(days=1)
        
    FACILITY_HOLIDAYS = {
        '2026-01-01', '2026-01-02', '2026-05-25', '2026-05-26',
        '2026-07-03', '2026-07-04', '2026-09-07', '2026-11-26',
        '2026-11-27', '2026-12-24', '2026-12-25', '2026-12-31'
    }

    with closing(get_db_connection()) as conn:
        # 1. Scheduled
        sched_rows = conn.execute('''
            SELECT schedule_date, SUM(load_count) as scheduled
            FROM daily_schedule
            WHERE schedule_date BETWEEN ? AND ?
            GROUP BY schedule_date
        ''', (start_str, end_str)).fetchall()
        sched_map = {r['schedule_date']: r['scheduled'] for r in sched_rows}
        
        # 2. Actual
        actual_rows = conn.execute('''
            SELECT date_received, COUNT(id) as actual
            FROM truck_logs
            WHERE date_received BETWEEN ? AND ? AND test_status != 'REJECTED'
            GROUP BY date_received
        ''', (start_str, end_str)).fetchall()
        actual_map = {r['date_received']: r['actual'] for r in actual_rows}
        
        voc_file_counts = get_voc_file_counts(force_refresh=force_refresh)
        for iso_date in date_list:
            voc_cnt = voc_file_counts.get(iso_date, 0)
            db_cnt = actual_map.get(iso_date, 0)
            if voc_cnt > db_cnt:
                actual_map[iso_date] = voc_cnt

        rows_data = []
        for d in date_list:
            s = sched_map.get(d, 0)
            a = actual_map.get(d, 0)
            v = a - s
            dt = datetime.strptime(d, '%Y-%m-%d').date()
            day_name = dt.strftime('%A')
            is_holl = d in FACILITY_HOLIDAYS
            
            if not include_weekends and s == 0 and a == 0 and dt.weekday() >= 5:
                continue
            if exclude_holidays and is_holl:
                continue
            
            status = f"Surpass (+{v})" if v > 0 else (f"Deficit ({v})" if v < 0 else "On Target (0)")
            note = ""
            if is_holl:
                note = "Holiday"
            elif dt.weekday() >= 5:
                note = "Weekend"
            
            rows_data.append({
                'date': d,
                'day_name': day_name,
                'scheduled': s,
                'actual': a,
                'variance': v,
                'status': status,
                'note': note
            })

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Variance Tracker"
        ws.views.sheetView[0].showGridLines = True

        # Custom Styling Palette
        navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        slate_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        card_blue_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        card_green_fill = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        card_amber_fill = PatternFill(start_color="FFFBEB", end_color="FFFBEB", fill_type="solid")
        card_purple_fill = PatternFill(start_color="F5F3FF", end_color="F5F3FF", fill_type="solid")

        green_badge_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        green_badge_font = Font(name="Calibri", size=11, bold=True, color="166534")

        red_badge_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        red_badge_font = Font(name="Calibri", size=11, bold=True, color="991B1B")

        gray_badge_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        gray_badge_font = Font(name="Calibri", size=11, bold=True, color="475569")

        title_font = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
        subtitle_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        meta_label_font = Font(name="Calibri", size=10, bold=True, color="475569")
        meta_val_font = Font(name="Calibri", size=10, color="0F172A")

        kpi_lbl_font = Font(name="Calibri", size=9, bold=True, color="475569")
        kpi_val_font = Font(name="Calibri", size=16, bold=True, color="0F172A")

        tbl_hdr_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Calibri", size=11)
        bold_font = Font(name="Calibri", size=11, bold=True)

        align_center = Alignment(horizontal="center", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        thin_side = Side(border_style="thin", color="CBD5E1")
        double_bottom_side = Side(border_style="double", color="0F172A")
        thick_top_side = Side(border_style="thin", color="0F172A")

        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        summary_border = Border(top=thick_top_side, bottom=double_bottom_side)

        # 1. Header Banner
        ws.merge_cells("A1:G1")
        ws["A1"] = "CLEAN HARBORS - BUTTONWILLOW FACILITY"
        ws["A1"].font = title_font
        ws["A1"].fill = navy_fill
        ws["A1"].alignment = align_center
        ws.row_dimensions[1].height = 28

        ws.merge_cells("A2:G2")
        ws["A2"] = "COMPLIANCE REPORT - VARIANCE TRACKER (SCHEDULED VS. ACTUAL TRUCK LOADS)"
        ws["A2"].font = subtitle_font
        ws["A2"].fill = slate_fill
        ws["A2"].alignment = align_center
        ws.row_dimensions[2].height = 22

        # 2. Metadata Block
        ws["A4"] = "Date Range:"
        ws["A4"].font = meta_label_font
        ws["B4"] = f"{start_str} to {end_str}"
        ws["B4"].font = meta_val_font

        ws["D4"] = "Generated:"
        ws["D4"].font = meta_label_font
        ws["E4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        ws["E4"].font = meta_val_font

        ws["A5"] = "Filters:"
        ws["A5"].font = meta_label_font
        ws["B5"] = f"Include 0/0 Weekends: {'Yes' if include_weekends else 'No'} | Exclude Holidays: {'Yes' if exclude_holidays else 'No'}"
        ws["B5"].font = meta_val_font

        # 3. KPI Summary Block (Rows 7-8)
        kpis = [
            ("B", "C", "TOTAL SCHEDULED", "=SUM(C12:C{end_row})", card_blue_fill),
            ("D", "D", "TOTAL ACTUAL WEIGHED", "=SUM(D12:D{end_row})", card_green_fill),
            ("E", "E", "NET VARIANCE", "=SUM(E12:E{end_row})", card_amber_fill),
            ("F", "G", "AVG DAILY VARIANCE", "=AVERAGE(E12:E{end_row})", card_purple_fill),
        ]
        
        start_row = 12
        end_row = 11 + len(rows_data)

        # Draw KPI cards
        ws["B7"] = "TOTAL SCHEDULED"
        ws["B7"].font = kpi_lbl_font
        ws["B7"].fill = card_blue_fill
        ws["B7"].alignment = align_center
        ws["B8"] = f"=SUM(C{start_row}:C{end_row})" if rows_data else 0
        ws["B8"].font = kpi_val_font
        ws["B8"].fill = card_blue_fill
        ws["B8"].alignment = align_center
        ws["B8"].number_format = "#,##0"

        ws["C7"] = "TOTAL ACTUAL WEIGHED"
        ws["C7"].font = kpi_lbl_font
        ws["C7"].fill = card_green_fill
        ws["C7"].alignment = align_center
        ws["C8"] = f"=SUM(D{start_row}:D{end_row})" if rows_data else 0
        ws["C8"].font = kpi_val_font
        ws["C8"].fill = card_green_fill
        ws["C8"].alignment = align_center
        ws["C8"].number_format = "#,##0"

        ws["D7"] = "NET VARIANCE"
        ws["D7"].font = kpi_lbl_font
        ws["D7"].fill = card_amber_fill
        ws["D7"].alignment = align_center
        ws["D8"] = f"=SUM(E{start_row}:E{end_row})" if rows_data else 0
        ws["D8"].font = kpi_val_font
        ws["D8"].fill = card_amber_fill
        ws["D8"].alignment = align_center
        ws["D8"].number_format = "+#,##0;-#,##0;0"

        ws["E7"] = "AVG DAILY VARIANCE"
        ws["E7"].font = kpi_lbl_font
        ws["E7"].fill = card_purple_fill
        ws["E7"].alignment = align_center
        ws["E8"] = f"=AVERAGE(E{start_row}:E{end_row})" if rows_data else 0
        ws["E8"].font = kpi_val_font
        ws["E8"].fill = card_purple_fill
        ws["E8"].alignment = align_center
        ws["E8"].number_format = "+0.0;-0.0;0.0"

        ws.row_dimensions[7].height = 18
        ws.row_dimensions[8].height = 26

        # 4. Data Table Headers (Row 11)
        headers = ["Date", "Day of Week", "Scheduled Loads", "Actual Loads", "Variance", "Performance Status", "Notes"]
        ws.row_dimensions[11].height = 26
        for col_idx, text in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_idx, value=text)
            cell.font = tbl_hdr_font
            cell.fill = slate_fill
            cell.alignment = align_center if col_idx not in (3, 4, 5) else align_right
            cell.border = cell_border

        # 5. Populate Data Rows
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        for i, r in enumerate(rows_data, 12):
            ws.row_dimensions[i].height = 20
            row_fill = zebra_fill if i % 2 == 1 else PatternFill(fill_type=None)

            c1 = ws.cell(row=i, column=1, value=r['date'])
            c1.alignment = align_center
            
            c2 = ws.cell(row=i, column=2, value=r['day_name'])
            c2.alignment = align_center

            c3 = ws.cell(row=i, column=3, value=r['scheduled'])
            c3.alignment = align_right
            c3.number_format = "#,##0"

            c4 = ws.cell(row=i, column=4, value=r['actual'])
            c4.alignment = align_right
            c4.number_format = "#,##0"

            c5 = ws.cell(row=i, column=5, value=r['variance'])
            c5.alignment = align_right
            c5.number_format = "+#,##0;-#,##0;0"

            c6 = ws.cell(row=i, column=6, value=r['status'])
            c6.alignment = align_center
            if r['variance'] > 0:
                c6.fill = green_badge_fill
                c6.font = green_badge_font
            elif r['variance'] < 0:
                c6.fill = red_badge_fill
                c6.font = red_badge_font
            else:
                c6.fill = gray_badge_fill
                c6.font = gray_badge_font

            c7 = ws.cell(row=i, column=7, value=r['note'])
            c7.alignment = align_center

            for c in (c1, c2, c3, c4, c5, c7):
                c.font = regular_font
                if c != c6 and row_fill.fill_type:
                    c.fill = row_fill
                c.border = cell_border
            c6.border = cell_border

        # 6. Total / Summary Row
        tot_row = end_row + 1
        ws.row_dimensions[tot_row].height = 24

        ws.cell(row=tot_row, column=1, value="TOTAL / AVERAGE").font = bold_font
        ws.cell(row=tot_row, column=1).alignment = align_left

        c_ts = ws.cell(row=tot_row, column=3, value=f"=SUM(C{start_row}:C{end_row})" if rows_data else 0)
        c_ts.font = bold_font
        c_ts.alignment = align_right
        c_ts.number_format = "#,##0"

        c_ta = ws.cell(row=tot_row, column=4, value=f"=SUM(D{start_row}:D{end_row})" if rows_data else 0)
        c_ta.font = bold_font
        c_ta.alignment = align_right
        c_ta.number_format = "#,##0"

        c_tv = ws.cell(row=tot_row, column=5, value=f"=SUM(E{start_row}:E{end_row})" if rows_data else 0)
        c_tv.font = bold_font
        c_tv.alignment = align_right
        c_tv.number_format = "+#,##0;-#,##0;0"

        c_avg = ws.cell(row=tot_row, column=6, value=f"Avg: =AVERAGE(E{start_row}:E{end_row})" if rows_data else "Avg: 0")
        c_avg.font = bold_font
        c_avg.alignment = align_center

        for col_idx in range(1, 8):
            cell = ws.cell(row=tot_row, column=col_idx)
            cell.border = summary_border

        # Auto-fit Column Widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or '')
                if cell.number_format and val.startswith('='):
                    val = "FormulaValue"
                if len(val) > max_len and cell.row > 2: # Ignore long title banner
                    max_len = len(val)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)
        ws.column_dimensions["F"].width = 24 # Status column width

        # Return file stream
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"Variance_Tracker_{start_str}_to_{end_str}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename
        )


@reports_bp.route('/api/compliance/export_profile_variance')
def export_profile_variance():
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime

    limit = request.args.get('limit', 20, type=int)
    start_date = request.args.get('start_date', '2026-05-11')
    end_date = request.args.get('end_date', datetime.now().strftime('%Y-%m-%d'))

    with closing(get_db_connection()) as conn:
        sched_sql = """
            SELECT 
                TRIM(UPPER(s.profile_number)) as profile,
                MAX(s.generator) as generator,
                MAX(s.routing_code) as win_code,
                SUM(COALESCE(s.load_count, 1)) as scheduled_loads,
                COUNT(DISTINCT s.schedule_date) as scheduled_days
            FROM daily_schedule s
            WHERE s.profile_number IS NOT NULL AND TRIM(s.profile_number) != ''
              AND TRIM(UPPER(s.profile_number)) NOT IN ('PENDING', 'TBD', 'CHDRUMLOAD', 'EMDRUMLOAD', 'GEDRUMLOAD', 'BLCBPNONEB')
              AND TRIM(UPPER(s.profile_number)) NOT LIKE '%DRUM%'
              AND s.schedule_date >= ? AND s.schedule_date <= ?
            GROUP BY TRIM(UPPER(s.profile_number))
        """
        sched_df = pd.read_sql(sched_sql, conn, params=(start_date, end_date))

        # Build normalized profile map for typo resolution
        known_profiles = set(sched_df['profile'].unique())
        prof_sql_all = "SELECT TRIM(UPPER(profile_number)) FROM profiles WHERE profile_number IS NOT NULL"
        for p_row in conn.execute(prof_sql_all):
            if p_row[0]: known_profiles.add(p_row[0])

        norm_to_official = {}
        for p in known_profiles:
            n = re.sub(r'[^A-Z0-9]', '', str(p).strip().upper())
            if n and n not in norm_to_official:
                norm_to_official[n] = p

        # Load VOC profile counts
        cache_path = os.path.join(os.path.dirname(__file__), 'voc_profile_cache.json')
        raw_voc_counts = {}
        if os.path.exists(cache_path):
            try:
                with open(cache_path, 'r', encoding='utf-8') as f:
                    raw_voc_counts = json.load(f).get('profile_counts', {})
            except Exception:
                pass

        resolved_voc = {}
        for raw, count in raw_voc_counts.items():
            raw_u = str(raw).strip().upper()
            target = raw_u
            if raw_u in known_profiles:
                target = raw_u
            else:
                n = re.sub(r'[^A-Z0-9]', '', raw_u)
                if n in norm_to_official:
                    target = norm_to_official[n]
                elif raw_u.startswith("CH") and raw_u[2:] in known_profiles:
                    target = raw_u[2:]
                elif ("CH" + raw_u) in known_profiles:
                    target = "CH" + raw_u
            resolved_voc[target] = resolved_voc.get(target, 0) + count

        voc_df = pd.DataFrame(list(resolved_voc.items()), columns=['profile', 'voc_received'])

        rcv_sql = """
            SELECT 
                TRIM(UPPER(t.profile_number)) as profile,
                COUNT(*) as db_received
            FROM truck_logs t
            WHERE t.profile_number IS NOT NULL AND TRIM(t.profile_number) != ''
              AND (t.test_status IS NULL OR t.test_status != 'REJECTED')
              AND TRIM(UPPER(t.profile_number)) NOT LIKE '%DRUM%'
              AND t.date_received >= ? AND t.date_received <= ?
            GROUP BY TRIM(UPPER(t.profile_number))
        """
        rcv_df = pd.read_sql(rcv_sql, conn, params=(start_date, end_date))

        df = pd.merge(sched_df, voc_df, on='profile', how='left')
        df['voc_received'] = df['voc_received'].fillna(0).astype(int)

        df = pd.merge(df, rcv_df, on='profile', how='left')
        df['db_received'] = df['db_received'].fillna(0).astype(int)

        df['received_loads'] = df[['voc_received', 'db_received']].max(axis=1)
        df['missed_loads'] = df['scheduled_loads'] - df['received_loads']
        df['show_up_rate'] = (df['received_loads'] / df['scheduled_loads'] * 100).round(1)

        prof_sql = "SELECT TRIM(UPPER(profile_number)) as profile, generator as p_generator, waste_name, win_code as p_win FROM profiles"
        prof_df = pd.read_sql(prof_sql, conn)
        df = pd.merge(df, prof_df, on='profile', how='left')

        df['generator'] = df['generator'].fillna(df['p_generator']).fillna('UNKNOWN GENERATOR')
        df['waste_name'] = df['waste_name'].fillna('N/A')
        df['win_code'] = df['win_code'].fillna(df['p_win']).fillna('N/A')

        top20 = df[df['missed_loads'] > 0].sort_values(by=['missed_loads', 'scheduled_loads'], ascending=[False, False]).head(limit).copy()

    # Create Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Top {limit} Schedule Variance"
    ws.views.sheetView[0].showGridLines = True

    # Styling Palette
    DARK_NAVY = "1E293B"      # Title Banner
    SLATE_BLUE = "334155"     # Subtitle & Table Headers
    LIGHT_GRAY_FILL = "F8FAFC" # Alternating row fill
    WHITE = "FFFFFF"

    CARD1_BG, CARD1_BORDER = "EFF6FF", "3B82F6"
    CARD2_BG, CARD2_BORDER = "ECFDF5", "10B981"
    CARD3_BG, CARD3_BORDER = "FEF2F2", "EF4444"
    CARD4_BG, CARD4_BORDER = "F5F3FF", "8B5CF6"

    font_title = Font(name="Calibri", size=16, bold=True, color=WHITE)
    font_subtitle = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_meta_label = Font(name="Calibri", size=10, bold=True, color="475569")
    font_meta_val = Font(name="Calibri", size=10, color="1E293B")
    font_kpi_hdr = Font(name="Calibri", size=9, bold=True, color="475569")
    font_kpi_val = Font(name="Calibri", size=16, bold=True, color="0F172A")
    font_tbl_hdr = Font(name="Calibri", size=11, bold=True, color=WHITE)
    font_cell = Font(name="Calibri", size=10, color="1E293B")
    font_cell_bold = Font(name="Calibri", size=10, bold=True, color="1E293B")

    thin_gray = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    double_bottom = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=Side(border_style="double", color="1E293B"))

    # Title Banner
    ws.merge_cells("A1:K1")
    cell_a1 = ws["A1"]
    cell_a1.value = "CLEAN HARBORS - BUTTONWILLOW FACILITY"
    cell_a1.font = font_title
    cell_a1.fill = PatternFill(start_color=DARK_NAVY, end_color=DARK_NAVY, fill_type="solid")
    cell_a1.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:K2")
    cell_a2 = ws["A2"]
    cell_a2.value = f"PROFILE SCHEDULE VARIANCE REPORT - TOP {limit} PROFILES WITH HIGHEST UNRECEIVED / MISSED LOADS"
    cell_a2.font = font_subtitle
    cell_a2.fill = PatternFill(start_color=SLATE_BLUE, end_color=SLATE_BLUE, fill_type="solid")
    cell_a2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    # Metadata
    ws["A4"] = "Report Scope:"
    ws["A4"].font = font_meta_label
    ws["B4"] = "Active Scheduled Profiles vs Received Truck Logs"
    ws["B4"].font = font_meta_val

    ws["D4"] = "Report Generated:"
    ws["D4"].font = font_meta_label
    ws["E4"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["E4"].font = font_meta_val

    ws["A5"] = "Ranking Metric:"
    ws["A5"].font = font_meta_label
    ws["B5"] = f"Top {limit} Profiles by Total Missed Loads (Scheduled minus Actual Received)"
    ws["B5"].font = font_meta_val

    # KPI Summary Cards
    last_row = 11 + len(top20)
    kpi_configs = [
        ("B7", "B8", "C8", "TOTAL SCHEDULED LOADS", f"=SUM(G12:G{last_row})", CARD1_BG, CARD1_BORDER),
        ("D7", "D8", "E8", "TOTAL ACTUAL RECEIVED", f"=SUM(H12:H{last_row})", CARD2_BG, CARD2_BORDER),
        ("F7", "F8", "G8", "TOTAL MISSED LOADS", f"=SUM(I12:I{last_row})", CARD3_BG, CARD3_BORDER),
        ("H7", "H8", "I8", "AVG SHOW-UP RATE", f"=AVERAGE(J12:J{last_row})", CARD4_BG, CARD4_BORDER),
    ]

    ws.row_dimensions[7].height = 18
    ws.row_dimensions[8].height = 28

    for top_left, val_start, val_end, label, formula, bg_color, b_color in kpi_configs:
        col_v1, col_v2 = val_start[0], val_end[0]
        r_hdr, r_val = int(top_left[1:]), int(val_start[1:])
        cell_h = ws[top_left]
        cell_h.value = label
        cell_h.font = font_kpi_hdr
        cell_h.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell_h.alignment = Alignment(horizontal="center", vertical="center")
        
        ws.merge_cells(f"{val_start}:{val_end}")
        cell_v = ws[val_start]
        cell_v.value = formula
        cell_v.font = font_kpi_val
        cell_v.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell_v.alignment = Alignment(horizontal="center", vertical="center")
        cell_v.number_format = "0.0%" if "AVERAGE" in formula else "#,##0"
            
        card_side = Side(border_style="medium", color=b_color)
        for row in range(r_hdr, r_val + 1):
            for col_letter in [col_v1, col_v2]:
                c_item = ws[f"{col_letter}{row}"]
                c_item.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                c_item.border = Border(
                    top=card_side if row == r_hdr else None,
                    bottom=card_side if row == r_val else None,
                    left=card_side if col_letter == col_v1 else None,
                    right=card_side if col_letter == col_v2 else None
                )

    # Table Headers
    headers = [
        "Rank", "Profile Number", "Generator Name", "Waste Description", 
        "WIN Code", "Scheduled Days", "Scheduled Loads", "Actual Received", 
        "Missed Loads", "Show-Up Rate", "Performance Status"
    ]
    ws.row_dimensions[11].height = 24
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col_num, value=h_text)
        cell.font = font_tbl_hdr
        cell.fill = PatternFill(start_color=SLATE_BLUE, end_color=SLATE_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center" if col_num in [1, 5, 6, 7, 8, 9, 10, 11] else "left", vertical="center")
        cell.border = cell_border

    # Data Rows
    start_row = 12
    for idx, (_, row) in enumerate(top20.iterrows(), start=1):
        r_idx = start_row + idx - 1
        ws.row_dimensions[r_idx].height = 20
        is_zebra = (idx % 2 == 0)
        row_fill = PatternFill(start_color=LIGHT_GRAY_FILL, end_color=LIGHT_GRAY_FILL, fill_type="solid") if is_zebra else None
        
        ws.cell(row=r_idx, column=1, value=idx)
        ws.cell(row=r_idx, column=2, value=row['profile'])
        ws.cell(row=r_idx, column=3, value=row['generator'])
        ws.cell(row=r_idx, column=4, value=row['waste_name'])
        ws.cell(row=r_idx, column=5, value=row['win_code'])
        ws.cell(row=r_idx, column=6, value=int(row['scheduled_days']))
        ws.cell(row=r_idx, column=7, value=int(row['scheduled_loads']))
        ws.cell(row=r_idx, column=8, value=int(row['received_loads']))
        ws.cell(row=r_idx, column=9, value=f"=G{r_idx}-H{r_idx}")
        ws.cell(row=r_idx, column=10, value=f"=IF(G{r_idx}>0, H{r_idx}/G{r_idx}, 0)")
        
        su_rate = row['show_up_rate']
        rcv_cnt = row['received_loads']
        if rcv_cnt == 0:
            status_text, bg_badge, txt_badge = "No Show (0%)", "FEE2E2", "991B1B"
        elif su_rate < 25.0:
            status_text, bg_badge, txt_badge = f"Critical Shortfall ({su_rate:.1f}%)", "FFEDD5", "C2410C"
        elif su_rate < 75.0:
            status_text, bg_badge, txt_badge = f"Partial Show ({su_rate:.1f}%)", "FEF9C3", "854D0E"
        else:
            status_text, bg_badge, txt_badge = f"On Track ({su_rate:.1f}%)", "DCFCE7", "166534"
            
        badge_cell = ws.cell(row=r_idx, column=11, value=status_text)
        badge_cell.font = Font(name="Calibri", size=10, bold=True, color=txt_badge)
        badge_cell.fill = PatternFill(start_color=bg_badge, end_color=bg_badge, fill_type="solid")
        badge_cell.alignment = Alignment(horizontal="center", vertical="center")

        for c_idx in range(1, 12):
            c_cell = ws.cell(row=r_idx, column=c_idx)
            if c_idx != 11 and row_fill:
                c_cell.fill = row_fill
            c_cell.border = cell_border
            
            if c_idx in [1, 2]:
                c_cell.alignment = Alignment(horizontal="center" if c_idx == 1 else "left", vertical="center")
                c_cell.font = font_cell_bold
            elif c_idx in [3, 4]:
                c_cell.alignment = Alignment(horizontal="left", vertical="center")
                c_cell.font = font_cell
            elif c_idx in [5, 6]:
                c_cell.alignment = Alignment(horizontal="center", vertical="center")
                c_cell.font = font_cell
                if c_idx == 6: c_cell.number_format = "#,##0"
            elif c_idx in [7, 8, 9]:
                c_cell.alignment = Alignment(horizontal="right", vertical="center")
                c_cell.font = font_cell_bold if c_idx == 9 else font_cell
                c_cell.number_format = "#,##0"
            elif c_idx == 10:
                c_cell.alignment = Alignment(horizontal="right", vertical="center")
                c_cell.font = font_cell_bold
                c_cell.number_format = "0.0%"

    # Total Row
    tot_row = start_row + len(top20)
    ws.row_dimensions[tot_row].height = 24
    ws.cell(row=tot_row, column=2, value=f"TOP {len(top20)} TOTALS")
    ws.cell(row=tot_row, column=6, value=f"=SUM(F12:F{tot_row-1})")
    ws.cell(row=tot_row, column=7, value=f"=SUM(G12:G{tot_row-1})")
    ws.cell(row=tot_row, column=8, value=f"=SUM(H12:H{tot_row-1})")
    ws.cell(row=tot_row, column=9, value=f"=SUM(I12:I{tot_row-1})")
    ws.cell(row=tot_row, column=10, value=f"=IF(G{tot_row}>0, H{tot_row}/G{tot_row}, 0)")

    for c_idx in range(1, 12):
        t_cell = ws.cell(row=tot_row, column=c_idx)
        t_cell.font = Font(name="Calibri", size=11, bold=True, color="0F172A")
        t_cell.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        t_cell.border = double_bottom
        if c_idx in [6, 7, 8, 9]:
            t_cell.alignment = Alignment(horizontal="right", vertical="center")
            t_cell.number_format = "#,##0"
        elif c_idx == 10:
            t_cell.alignment = Alignment(horizontal="right", vertical="center")
            t_cell.number_format = "0.0%"
        elif c_idx == 2:
            t_cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column Widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 42
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 24

    ws.freeze_panes = "A12"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"Top_{len(top20)}_Scheduled_Profiles_No_Show_Variance.xlsx"
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename
    )



