from flask import Blueprint, render_template, request, redirect, url_for, send_file, jsonify
from datetime import date, datetime
from contextlib import closing
from io import BytesIO
import pandas as pd
import uuid
import json
import os
import zipfile
from openpyxl.styles import Font, Alignment, Border, Side
import stu_services
from database import get_db_connection

stu_bp = Blueprint('stu_bp', __name__)

@stu_bp.route('/stu/hub')
def stu_hub():
    view = request.args.get('view', 'pipeline')  # 'pipeline' or 'inventory'
    category = request.args.get('category', 'All')
    
    with closing(get_db_connection()) as conn:
        # Get active pipeline loads
        active_jobs = conn.execute('''
            SELECT DISTINCT job_id 
            FROM drum_lab_queue 
            WHERE status != 'FINAL CODED'
            UNION
            SELECT DISTINCT job_id
            FROM drum_inventory
            WHERE process_type = 'PENDING SAMPLING' AND job_id IS NOT NULL
        ''').fetchall()
        
        pipeline_loads = []
        for row in active_jobs:
            job_id = row['job_id']
            if not job_id: continue
            
            # Fetch total drums in inventory for this job
            total_drums = conn.execute("SELECT COUNT(*) FROM drum_inventory WHERE job_id = ?", (job_id,)).fetchone()[0]
            
            # Fetch labs status
            total_labs = conn.execute("SELECT COUNT(*) FROM drum_lab_queue WHERE job_id = ?", (job_id,)).fetchone()[0]
            completed_labs = conn.execute("SELECT COUNT(*) FROM drum_lab_queue WHERE job_id = ? AND status = 'COMPLETED'", (job_id,)).fetchone()[0]
            coded_labs = conn.execute("SELECT COUNT(*) FROM drum_lab_queue WHERE job_id = ? AND coded_in_win = 1", (job_id,)).fetchone()[0]
            
            # Determine stage
            if total_drums == 0:
                stage = "Empty Load"
            elif completed_labs < total_labs:
                stage = f"Pending Lab Analysis ({completed_labs}/{total_labs} tested)"
            elif coded_labs < total_labs:
                stage = f"Ready for WIN Coding ({coded_labs}/{total_labs} coded)"
            else:
                stage = "Ready for Finalization"
                
            pipeline_loads.append({
                'job_id': job_id,
                'total_drums': total_drums,
                'total_labs': total_labs,
                'completed_labs': completed_labs,
                'coded_labs': coded_labs,
                'stage': stage
            })
            
        # Get all active inventory drums (including pending sampling)
        drums_raw = conn.execute("SELECT * FROM drum_inventory ORDER BY age DESC").fetchall()
        drums = []
        from datetime import datetime
        today = date.today()
        for row in drums_raw:
            d = dict(row)
            d['last_scan_date_display'] = d.get('last_scan_date') or 'N/A'
            d['days_since_scanned'] = 'N/A'
            
            scan_date_str = d.get('last_scan_date')
            if scan_date_str:
                scan_date_str = str(scan_date_str).strip()
                if scan_date_str and scan_date_str.lower() != 'none' and scan_date_str.lower() != 'nan':
                    # Try parsing multiple date formats
                    for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%y'):
                        try:
                            scan_dt = datetime.strptime(scan_date_str, fmt).date()
                            delta = today - scan_dt
                            d['days_since_scanned'] = max(0, delta.days)
                            break
                        except ValueError:
                            continue
            drums.append(d)
        
        process_types = sorted(list(set([str(d['process_type']).strip().upper() for d in drums if d.get('process_type')])))

        last_upload_row = conn.execute('SELECT MAX(import_date) FROM drum_inventory').fetchone()
        last_upload = last_upload_row[0] if last_upload_row else 'NO DATA'
        
        # Get active LAS bulk trucks pending release (Load As Sample)
        las_trucks_raw = conn.execute('''
            SELECT tl.*, 
                   p.generator, p.waste_description, p.win_code, p.voc_percentage, p.special_handling,
                   p.ph_range, p.physical_appearance, p.flash_point, p.expiration_date,
                   COALESCE(w.cyanide, p.cyanide) AS wvi_cyanide,
                   COALESCE(w.sulfides, p.sulfide) AS wvi_sulfides,
                   COALESCE(w.free_liquids, p.free_liquids) AS wvi_free_liquids,
                   w.ph_min, w.ph_max, w.voc_ppm AS wvi_voc_ppm, w.flashpoint AS wvi_flashpoint,
                   w.generator_name, w.waste_name, w.handling_instruction, w.color AS w_color, p.color AS p_color
            FROM truck_logs tl
            LEFT JOIN profiles p ON TRIM(UPPER(tl.profile_number)) = TRIM(UPPER(p.profile_number))
            LEFT JOIN profile_wvi w ON TRIM(UPPER(tl.profile_number)) = TRIM(UPPER(w.profile))
            WHERE tl.test_assigned LIKE 'LAS%' AND tl.test_status = 'LAB COMPLETED'
            ORDER BY tl.id DESC
        ''').fetchall()
        
        import re
        las_trucks = []
        for row in las_trucks_raw:
            t = dict(row)
            
            # --- FALLBACK FOR SPECTABLE DISPLAY ---
            if t.get('generator_name'):
                t['generator'] = t['generator_name']
            if t.get('waste_name'):
                t['waste_description'] = t['waste_name']
            if t.get('wvi_flashpoint'):
                t['flash_point'] = t['wvi_flashpoint']
            if t.get('handling_instruction'):
                t['special_handling'] = t['handling_instruction']
                
            # Set color
            t['color'] = t.get('w_color') or t.get('p_color') or ''
                
            # If ph_min or ph_max is not set in WVI, try parsing from profile ph_range
            if t.get('ph_min') is None or t.get('ph_max') is None:
                if t.get('ph_range'):
                    ph_str = str(t['ph_range']).strip().upper()
                    if "7 (NEUTRAL)" in ph_str or ph_str == "7" or "7 NEUTRAL" in ph_str:
                        t['ph_min'] = 4.0
                        t['ph_max'] = 10.0
                    else:
                        m = re.findall(r'(\d+\.?\d*)', ph_str)
                        if len(m) >= 2:
                            try:
                                if t.get('ph_min') is None: t['ph_min'] = float(m[0])
                                if t.get('ph_max') is None: t['ph_max'] = float(m[1])
                            except:
                                pass
                        elif len(m) == 1:
                            try:
                                if t.get('ph_min') is None: t['ph_min'] = float(m[0])
                                if t.get('ph_max') is None: t['ph_max'] = float(m[0])
                            except:
                                pass
                            
            # Convert voc_percentage fallback to voc_ppm
            if t.get('wvi_voc_ppm') is not None:
                t['voc_ppm'] = t['wvi_voc_ppm']
            elif t.get('voc_percentage') is not None:
                try:
                    val = float(t['voc_percentage'])
                    t['voc_ppm'] = val
                except:
                    t['voc_ppm'] = None
            else:
                t['voc_ppm'] = None
                
            # Compute default pH for form field auto-fill
            ph_val = 7.0
            if t.get('ph_min') is not None and t.get('ph_max') is not None:
                ph_val = round((t['ph_min'] + t['ph_max']) / 2.0, 1)
            elif t.get('ph_range'):
                m = re.findall(r'(\d+\.?\d*)', str(t['ph_range']))
                if len(m) >= 2:
                    try:
                        ph_val = round((float(m[0]) + float(m[1])) / 2.0, 1)
                    except:
                        pass
                elif len(m) == 1:
                    try:
                        ph_val = float(m[0])
                    except:
                        pass
            t['default_ph'] = ph_val
            
            # Compute default VOC for form field auto-fill
            voc_val = 0.0
            if t.get('voc_ppm') is not None:
                voc_val = t['voc_ppm']
            t['default_voc'] = round(voc_val, 1)
            
            # Compute default sulfides
            t['default_sulfides'] = 'Positive' if str(t.get('wvi_sulfides') or '').strip().lower() in ['yes', 'true', 'pos', 'positive', 'y', 'neg or pos'] else 'Negative'
            
            # Compute default cyanide
            t['default_cyanide'] = 'Positive' if str(t.get('wvi_cyanide') or '').strip().lower() in ['yes', 'true', 'pos', 'positive', 'y', 'neg or pos'] else 'Negative'
            
            # Compute default free liquids
            t['default_free_liquids'] = 'Yes' if str(t.get('wvi_free_liquids') or '').strip().lower() in ['yes', 'true', 'y', 'free'] else 'No'
            
            # Compute default flashpoint
            fp = str(t.get('wvi_flashpoint') or t.get('flash_point') or 'Not Required').strip()
            t['default_flashpoint'] = fp
            
            las_trucks.append(t)
        
    return render_template('stu_hub.html', 
                           pipeline_loads=pipeline_loads, 
                           drums=drums, 
                           view=view, 
                           category=category, 
                           last_upload=last_upload,
                           las_trucks=las_trucks,
                           process_types=process_types)

@stu_bp.route('/stu/inventory')
def stu_inventory():
    return redirect(url_for('stu_bp.stu_hub', view='inventory', category=request.args.get('category', 'All')))

@stu_bp.route('/stu/sampling', methods=['GET', 'POST'])
def stu_sampling():
    if request.method == 'POST':
        load_number = request.form.get('load_number', '').strip()
        if not load_number: return "No Load Number entered", 400
        
        if 'label_pdf' not in request.files: return "No file uploaded", 400
        file = request.files['label_pdf']
        if file.filename == '': return "No file selected", 400

        file_bytes = file.read()
        
        raw_drums = stu_services.parse_drum_labels_from_pdf(BytesIO(file_bytes))
        if not raw_drums:
            return "<script>alert('No valid drums found in PDF.'); window.history.back();</script>"

        if not os.path.exists('temp_uploads'):
            os.makedirs('temp_uploads')
            
        temp_filename = f"{uuid.uuid4().hex}.pdf"
        temp_filepath = os.path.join('temp_uploads', temp_filename)
        
        with open(temp_filepath, 'wb') as f:
            f.write(file_bytes)
        
        return render_template('stu_sampling.html', drums=raw_drums, pdf_filename=temp_filename, raw_drums_json=json.dumps(raw_drums), load_number=load_number)
        
    return render_template('stu_sampling.html', drums=None)

@stu_bp.route('/generate_sampling_packet', methods=['POST'])
def generate_sampling_packet():
    pdf_filename = request.form.get('pdf_filename')
    selected_drums_json = request.form.get('selected_drums')
    load_number = request.form.get('load_number', '').strip()

    if not pdf_filename or not selected_drums_json or not load_number:
        return f"Error: Data dropped.", 400

    temp_filepath = os.path.join('temp_uploads', pdf_filename)
    try:
        with open(temp_filepath, 'rb') as f:
            file_bytes = f.read()
    except FileNotFoundError:
        return "Temporary file lost. Please start over and re-upload the PDF.", 400

    raw_drums = json.loads(selected_drums_json)
    job_name = load_number

    with closing(get_db_connection()) as conn:
        picklist_data, total_samples = stu_services.process_drums(conn, raw_drums)
        
        for d in raw_drums:
            existing = conn.execute("SELECT id FROM drum_inventory WHERE track_no = ?", (d['drum_id'],)).fetchone()
            if not existing:
                conn.execute("INSERT INTO drum_inventory (track_no, inb_prof, manifest, process_type, weight, ph, age, voc_ppm, voc_weight, import_date, job_id, status) VALUES (?, ?, ?, 'PENDING SAMPLING', 0, 0, 0, 0, 0, ?, ?, 'PLANT RECEIVED')",
                             (d['drum_id'], d['profile'], d['manifest'], date.today().isoformat(), job_name))
                
        for p in picklist_data:
            if p.get('is_sampled') == 'Yes':
                
                existing_lab = conn.execute("SELECT id FROM drum_lab_queue WHERE drum_id = ? AND status != 'COMPLETED'", (p['drum_id'],)).fetchone()
                if not existing_lab:
                    conn.execute('''
                        INSERT INTO drum_lab_queue (job_id, drum_id, profile, manifest, tests_required, status)
                        VALUES (?, ?, ?, ?, ?, 'PENDING')
                    ''', (job_name, p['drum_id'], p['profile'], p['manifest'], p.get('voc_testing_trigger', 'FingerPrint')))
                
        conn.commit()

    memory_zip = BytesIO()
    with zipfile.ZipFile(memory_zip, 'w', zipfile.ZIP_DEFLATED) as zf:
        picklist_buffer = BytesIO()
        stu_services.create_pdf_report(picklist_buffer, f"Drum Sampling Pick List - {job_name}", picklist_data, total_samples)
        zf.writestr(f"Picklist_{job_name}.pdf", picklist_buffer.getvalue())

        lab_buffer = BytesIO()
        stu_services.create_lab_sheet_pdf(lab_buffer, job_name, picklist_data)
        zf.writestr(f"LabBenchSheet_{job_name}.pdf", lab_buffer.getvalue())

        df = pd.DataFrame(picklist_data)
        excel_io = BytesIO()
        with pd.ExcelWriter(excel_io, engine='openpyxl') as writer:
            cols = ["sample_num", "drum_id", "manifest", "manifest_line", "profile", "display_profile", "waste_code", "accumulation_date", "is_sampled"]
            exist_cols = [c for c in cols if c in df.columns]
            df[exist_cols].to_excel(writer, index=False, sheet_name='Data')
            ws = writer.sheets['Data']
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            ws.page_setup.orientation = 'landscape'
        zf.writestr(f"Data_{job_name}.xlsx", excel_io.getvalue())

        annotated_buffer = BytesIO()
        stu_services.create_annotated_pdf(BytesIO(file_bytes), annotated_buffer, picklist_data)
        if annotated_buffer.getvalue():
            zf.writestr(f"LABELS_{job_name}_Marked.pdf", annotated_buffer.getvalue())

    if os.path.exists(temp_filepath):
        os.remove(temp_filepath)

    memory_zip.seek(0)
    return send_file(memory_zip, mimetype='application/zip', as_attachment=True, download_name=f"Drum_Sampling_{job_name}.zip")

@stu_bp.route('/upload_vpi', methods=['POST'])
def upload_vpi():
    if 'vpi_file' not in request.files: return "No file uploaded", 400
    file = request.files['vpi_file']
    if file.filename == '': return "No file selected", 400
    try:
        df = pd.read_csv(file)
        df.columns = [str(c).strip() for c in df.columns]
        
        required = ['Track No', 'Process Type', 'Weight', 'pH', 'Inb Prof', 'Age', 'Type']
        if not all(col in df.columns for col in required): 
            return "Error: Uploaded CSV is missing required columns from WIN.", 400
            
        df['Process Type'] = df['Process Type'].astype(str).str.strip().str.lower()
        df['Inb Prof'] = df['Inb Prof'].astype(str).str.strip().str.lower()
        df['Type'] = df['Type'].astype(str).str.strip().str.lower() 
        
        with closing(get_db_connection()) as conn:
            df['Weight'] = pd.to_numeric(df['Weight'], errors='coerce').fillna(0)
            
            # Now apply exclusions and filtering
            df = df[~df['Process Type'].isin(['=', 'nan', ''])]
            
            # Exclude bulk loads (weight > 5000 lbs) EXCEPT if it is a put pile
            is_heavy = df['Weight'] > 5000
            is_put_pile = df['Process Type'].str.contains('put', na=False)
            
            df = df[~(is_heavy & ~is_put_pile)]
            
            df['pH'] = pd.to_numeric(df['pH'], errors='coerce').fillna(0)
            df['Age'] = pd.to_numeric(df['Age'], errors='coerce').fillna(0)
            
            df = df.drop_duplicates(subset=['Track No'], keep='last')
            
            # Preserve existing statuses, pending sampling, and rejected/failed/resample drums
            try:
                preserved_drums_raw = conn.execute("""
                    SELECT * FROM drum_inventory 
                    WHERE status NOT IN ('FINAL CODED', 'ACTIVE', 'PLANT RECEIVED') 
                       OR reject_notes IS NOT NULL 
                       OR outgoing_manifest IS NOT NULL
                       OR process_type = 'PENDING SAMPLING'
                """).fetchall()
                preserved_drums = [dict(d) for d in preserved_drums_raw]
            except Exception as ex:
                print(f"Error fetching preserved drums: {ex}")
                preserved_drums = []
                
            conn.execute('DELETE FROM drum_inventory')
            
            # Query profiles database to map voc_percentage
            profiles_df = pd.read_sql_query("SELECT LOWER(profile_number) as profile_number, voc_percentage FROM profiles", conn)
            
            # Map VOC percentage
            voc_dict = dict(zip(profiles_df['profile_number'], profiles_df['voc_percentage']))
            df['voc_ppm'] = pd.to_numeric(df['Inb Prof'].map(voc_dict), errors='coerce').fillna(0.0)
            df['voc_weight'] = df['Weight'] * df['voc_ppm']
            
            if 'Area' in df.columns:
                df['location'] = df['Area'].astype(str).str.strip().str.replace("'", "")
                df['location'] = df['location'].replace({'nan': None, 'None': None, '': None})
            else:
                df['location'] = None

            if 'Last Scan Date' in df.columns:
                df['last_scan_date'] = df['Last Scan Date'].astype(str).str.strip().str.replace("nan", "")
            else:
                df['last_scan_date'] = None
                
            statuses = ['PLANT RECEIVED' if str(pt).strip().lower() == 'pending sampling' else 'FINAL CODED' for pt in df['Process Type']]
            cleaned_data = list(zip(df['Track No'], df['Inb Prof'], df['Process Type'], df['Weight'], df['pH'], df['Age'], df['voc_ppm'], df['voc_weight'], [date.today().isoformat()]*len(df), statuses, df['location'], df['last_scan_date']))
            conn.executemany("INSERT INTO drum_inventory (track_no, inb_prof, process_type, weight, ph, age, voc_ppm, voc_weight, import_date, status, location, last_scan_date) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", cleaned_data)
            
            # Re-apply preserved statuses and re-insert missing ones
            imported_tracks_upper = {str(t).strip().upper() for t in df['Track No']}
            
            for p in preserved_drums:
                track_upper = str(p['track_no']).strip().upper()
                if track_upper in imported_tracks_upper:
                    # Update status, reject_notes, outgoing_manifest for the imported drum case-insensitively
                    conn.execute("""
                        UPDATE drum_inventory 
                        SET status = ?, reject_notes = ?, outgoing_manifest = ? 
                        WHERE TRIM(UPPER(track_no)) = ?
                    """, (p['status'], p['reject_notes'], p['outgoing_manifest'], track_upper))
                else:
                    # Re-insert the full row because it was not in the new VPI feed
                    loc_val = p['location']
                    status_val = p['status'] if p['status'] is not None else 'PLANT RECEIVED'
                    conn.execute("""
                        INSERT INTO drum_inventory (
                            track_no, inb_prof, manifest, process_type, weight, ph, age, 
                            voc_ppm, voc_weight, import_date, job_id, status, reject_notes, 
                            outgoing_manifest, location, last_scan_date
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        p['track_no'], p['inb_prof'], p['manifest'], p['process_type'], 
                        p['weight'], p['ph'], p['age'], p['voc_ppm'], p['voc_weight'], 
                        p['import_date'], p['job_id'], status_val, p['reject_notes'], 
                        p['outgoing_manifest'], loc_val, p.get('last_scan_date')
                    ))
                
            conn.commit()
    except Exception as e: 
        return f"Critical Error processing file: {e}", 500
    
    return redirect(url_for('stu_bp.stu_hub', view='inventory'))

@stu_bp.route('/export_stu', methods=['POST'])
def export_stu():
    selected_ids = request.form.getlist('selected_drums')
    category = request.form.get('category_name', 'Custom')
    if not selected_ids: return "No drums selected for export", 400
        
    placeholders = ','.join('?' for _ in selected_ids)
    with closing(get_db_connection()) as conn:
        df = pd.read_sql_query(f"SELECT track_no, inb_prof, process_type, location, weight, ph, age, voc_ppm, voc_weight FROM drum_inventory WHERE id IN ({placeholders}) ORDER BY age DESC", conn, params=selected_ids)
        
    if df.empty: return "No data found", 400
    df.rename(columns={'track_no': 'Track No', 'inb_prof': 'Inb Prof', 'process_type': 'Process Type', 'location': 'Location', 'weight': 'Weight', 'ph': 'pH', 'age': 'Age', 'voc_ppm': 'VOC', 'voc_weight': 'VOC Weight'}, inplace=True)
    df.insert(0, ' ', '')
    
    output = BytesIO()
    clean_category = category.replace('_', ' ')
    date_str = date.today().strftime('%m-%d-%Y')
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startrow=1, sheet_name=f"{category[:31]}")
        worksheet = writer.sheets[f"{category[:31]}"]
        worksheet.page_setup.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        worksheet.page_setup.orientation = 'landscape'
        last_row = len(df) + 2  
        
        title_cell = worksheet.cell(row=1, column=1, value=f"{clean_category} - Generated On - {date.today().strftime('%m/%d/%Y')}")
        worksheet.merge_cells('A1:J1')
        title_cell.font = Font(size=18, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        for col_num in range(1, 11):
            worksheet.cell(row=2, column=col_num).font = Font(size=16, bold=True, underline='double')
            worksheet.cell(row=2, column=col_num).alignment = Alignment(horizontal='center')
            
        worksheet.column_dimensions['A'].width = 6
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for r in range(3, last_row + 1): 
            worksheet.cell(row=r, column=1).border = thin_border
            worksheet.cell(row=r, column=6).number_format = '0.00' 
            worksheet.cell(row=r, column=9).number_format = '0.00' 
            worksheet.cell(row=r, column=10).number_format = '0.00' 
            
        worksheet.cell(row=last_row + 2, column=4, value="Total Weight (lbs):").font = Font(bold=True)
        wt_tot = worksheet.cell(row=last_row + 2, column=5, value=f"=SUM(F3:F{last_row})")
        wt_tot.font = Font(bold=True)
        wt_tot.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 3, column=4, value="Total Weight (tons):").font = Font(bold=True)
        wt_tons = worksheet.cell(row=last_row + 3, column=5, value=f"=E{last_row + 2}/2000")
        wt_tons.font = Font(bold=True)
        wt_tons.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 2, column=8, value="Total VOC Weight:").font = Font(bold=True)
        voc_tot = worksheet.cell(row=last_row + 2, column=9, value=f"=SUM(J3:J{last_row})")
        voc_tot.font = Font(bold=True)
        voc_tot.number_format = '#,##0.00'
        
        worksheet.cell(row=last_row + 3, column=8, value="Average VOC:").font = Font(bold=True)
        voc_avg = worksheet.cell(row=last_row + 3, column=9, value=f"=AVERAGE(I3:I{last_row})")
        voc_avg.font = Font(bold=True)
        voc_avg.number_format = '0.00'
        
        for col_idx, col in enumerate(worksheet.columns, start=1):
            if col_idx == 1: continue 
            max_length = 0
            for cell in col:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            col_letter = worksheet.cell(row=2, column=col_idx).column_letter
            worksheet.column_dimensions[col_letter].width = max_length + 3

    output.seek(0)
    export_filename = f"{clean_category} {date_str}.xlsx"
    return send_file(output, as_attachment=True, download_name=export_filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@stu_bp.route('/drum_action', methods=['POST'])
@stu_bp.route('/stu/drum_action', methods=['POST'])
def drum_action():
    drum_id = request.form.get('drum_id')
    action = request.form.get('action') # 'RESAMPLE' or 'REJECT'
    notes = request.form.get('reject_notes', '')
    manifest = request.form.get('outgoing_manifest', '')
    
    new_status = None
    with closing(get_db_connection()) as conn:
        if action == 'RESAMPLE':
            drum = conn.execute("SELECT status, process_type FROM drum_inventory WHERE id = ?", (drum_id,)).fetchone()
            if drum:
                if drum['status'] == 'RESAMPLE':
                    new_status = 'PLANT RECEIVED' if drum['process_type'] == 'PENDING SAMPLING' else 'FINAL CODED'
                else:
                    new_status = 'RESAMPLE'
                conn.execute("UPDATE drum_inventory SET status = ? WHERE id = ?", (new_status, drum_id))
        elif action == 'REJECT':
            conn.execute("UPDATE drum_inventory SET status = 'REJECTED', reject_notes = ?, outgoing_manifest = ? WHERE id = ?", (notes, manifest, drum_id))
            new_status = 'REJECTED'
        elif action == 'UPDATE_STATUS':
            target_status = request.form.get('status')
            if target_status == 'REJECTED':
                conn.execute("UPDATE drum_inventory SET status = 'REJECTED', reject_notes = ?, outgoing_manifest = ? WHERE id = ?", 
                             (notes, manifest, drum_id))
            else:
                conn.execute("UPDATE drum_inventory SET status = ?, reject_notes = NULL, outgoing_manifest = NULL WHERE id = ?", 
                             (target_status, drum_id))
            new_status = target_status
        conn.commit()
        
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({"status": "success", "drum_id": drum_id, "action": action, "new_status": new_status})
    return redirect(url_for('stu_bp.stu_hub', view='inventory'))

@stu_bp.route('/stu/put_pile_action', methods=['POST'])
def put_pile_action():
    drum_id = request.form.get('drum_id')
    action = request.form.get('action') # 'PASS' or 'FAIL'
    recipe = request.form.get('recipe', '')
    notes = request.form.get('notes', '')
    
    with closing(get_db_connection()) as conn:
        drum = conn.execute("SELECT track_no FROM drum_inventory WHERE id = ?", (drum_id,)).fetchone()
        track_no = drum['track_no'] if drum else 'UNKNOWN'
        
        if action == 'PASS':
            conn.execute("UPDATE drum_inventory SET status = 'PASS' WHERE id = ?", (drum_id,))
        elif action == 'FAIL':
            conn.execute("UPDATE drum_inventory SET status = 'FAIL' WHERE id = ?", (drum_id,))
            conn.execute("INSERT INTO put_pile_retreats (track_no, retreat_date, recipe, notes) VALUES (?, ?, ?, ?)", 
                         (track_no, date.today().isoformat(), recipe, notes))
        conn.commit()
        
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({"status": "success", "drum_id": drum_id, "action": action})
    return redirect(url_for('stu_bp.stu_hub', view='inventory'))

@stu_bp.route('/stu/bulk_resample', methods=['POST'])
def bulk_resample():
    if request.is_json:
        data = request.get_json()
        drum_ids = data.get('selected_drums', [])
    else:
        drum_ids = request.form.getlist('selected_drums')
        
    if not drum_ids:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({"status": "success", "updated": []})
        return redirect(url_for('stu_bp.stu_hub', view='inventory'))
        
    updated = []
    with closing(get_db_connection()) as conn:
        for d_id in drum_ids:
            drum = conn.execute("SELECT status, process_type FROM drum_inventory WHERE id = ?", (d_id,)).fetchone()
            if drum:
                current_status = drum['status']
                process_type = drum['process_type']
                
                if current_status == 'RESAMPLE':
                    new_status = 'PLANT RECEIVED' if process_type == 'PENDING SAMPLING' else 'FINAL CODED'
                else:
                    new_status = 'RESAMPLE'
                    
                conn.execute("UPDATE drum_inventory SET status = ? WHERE id = ?", (new_status, d_id))
                updated.append({"id": d_id, "status": new_status})
        conn.commit()
        
    if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({"status": "success", "updated": updated})
    return redirect(url_for('stu_bp.stu_hub', view='inventory'))

@stu_bp.route('/stu/audit_trail')
def stu_audit_trail():
    with closing(get_db_connection()) as conn:
        retreats = conn.execute("SELECT * FROM put_pile_retreats ORDER BY id DESC").fetchall()
    return render_template('stu_audit_trail.html', retreats=[dict(r) for r in retreats])
